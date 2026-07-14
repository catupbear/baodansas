# -*- coding: utf-8 -*-
"""
续保任务模块 — Flask API 蓝图
提供看板统计、任务列表、详情、分配、跟进、状态流转、手动扫描、配置、导出等接口。

⚠️ 本蓝图不像 insurance_bp 那样有自己的 before_request 鉴权（Flask 蓝图间不共享
before_request），所有路由必须显式加 @login_required / @admin_required（来自
auth/decorators.py），照抄 web/api.py 的模式，不能假设 g.current_user 会被自动设置。
"""
import io
import json
import logging

import pymysql
import pymysql.cursors
from flask import Blueprint, g, jsonify, request, send_file

from auth.decorators import login_required, admin_required
from auth.db import ROLE_SUPER_ADMIN, ROLE_ENTERPRISE, ROLE_EMPLOYEE
from insurance import renewal_db as rdb
from insurance.db import get_insurance_config, set_insurance_config

logger = logging.getLogger(__name__)

renewal_bp = Blueprint("renewal", __name__, url_prefix="/api/renewal")

# 模块级依赖注入（与 insurance_bp / api_bp 同样的模式）
_db = None
_cos_storage = None

_CONFIG_KEY_WINDOW_DAYS = "renewal_window_days"      # 提醒窗口天数（未来到期），默认90
_CONFIG_KEY_QRCODE_URL = "renewal_qrcode_url"          # 客服二维码URL（平台统一一张）
_EXPORT_ROW_LIMIT = 3000  # nginx对/api/renewal/*无显式超时覆盖，默认60秒会先于gunicorn 120秒生效，导出行数设上限保护


def init_renewal_api(db, cos_storage=None):
    """注入数据库/COS依赖"""
    global _db, _cos_storage
    _db = db
    _cos_storage = cos_storage


# ============================================================
# 权限过滤辅助函数（从 insurance/api.py:276-298 复制，保持逻辑一致；
# 不跨蓝图 import 私有函数，避免加载 insurance/api.py 整个5000+行模块）
# ============================================================

def _get_user_ids_filter():
    """根据当前用户角色返回 user_ids 过滤列表，超管和企业管理员返回 None（不按用户过滤）"""
    role = g.current_user["role"]
    uid = g.current_user["user_id"]
    if role == ROLE_SUPER_ADMIN:
        return None
    if role == ROLE_ENTERPRISE:
        return None
    return [uid]


def _get_enterprise_id_filter():
    """返回当前用户的 enterprise_id，超管返回 None（不过滤），其他角色按企业过滤"""
    role = g.current_user["role"]
    if role == ROLE_SUPER_ADMIN:
        return None
    parent_id = g.current_user.get("parent_id")
    return parent_id if parent_id else g.current_user["user_id"]


def _require_enterprise_or_admin():
    """企业管理员及以上（enterprise/super_admin）才能通过；否则返回 (response, code) 供路由直接 return。"""
    if g.current_user["role"] not in (ROLE_SUPER_ADMIN, ROLE_ENTERPRISE):
        return jsonify({"code": 403, "msg": "权限不足"}), 403
    return None


def _operator_name() -> str:
    """当前登录用户的姓名（g.current_user 里没有 name 字段，需要单独查一次）。"""
    from auth.db import get_user_by_id
    user = get_user_by_id(_db, g.current_user["user_id"])
    return (user or {}).get("name") or ""


def _can_operate_task(task: dict) -> bool:
    """当前用户是否可以对该任务做 followup/status 操作：本人任务，或 enterprise/super_admin。"""
    role = g.current_user["role"]
    if role == ROLE_SUPER_ADMIN:
        return True
    if role == ROLE_ENTERPRISE:
        return task.get("enterprise_id") == g.current_user.get("parent_id")
    if role == ROLE_EMPLOYEE:
        return task.get("assignee") == g.current_user["user_id"]
    return False


def _parse_list_filters(req):
    """解析 /list 和 /export 共用的筛选参数，供两处复用避免逻辑漂移。"""
    return {
        "status": req.args.get("status", "").strip(),
        "assignee": req.args.get("assignee", "").strip(),
        "date_start": req.args.get("date_start", "").strip(),
        "date_end": req.args.get("date_end", "").strip(),
        "keyword": req.args.get("keyword", "").strip(),
    }


def _resolve_scope():
    """当前用户的 (enterprise_id, assignee_ids) 数据范围，供 list/stats/export 复用。"""
    return _get_enterprise_id_filter(), _get_user_ids_filter()


def _apply_explicit_assignee_filter(assignee_ids, explicit_assignee):
    """
    若请求显式传了 assignee 参数（前端筛选下拉选了具体人），与角色收窄范围取交集：
    - assignee_ids 为 None（超管/企业管理员不按人过滤）且显式传了人 -> 用显式的那个人
    - assignee_ids 是具体列表（员工只看自己）-> 忽略显式参数，仍按角色收窄的范围
    """
    if not explicit_assignee:
        return assignee_ids
    try:
        eid = int(explicit_assignee)
    except ValueError:
        return assignee_ids
    if assignee_ids is None:
        return [eid]
    return [eid] if eid in assignee_ids else []


# ============================================================
# 看板统计
# ============================================================

@renewal_bp.route("/stats", methods=["GET"])
@login_required
def get_stats():
    enterprise_id, assignee_ids = _resolve_scope()
    stats = rdb.get_stats(_db, enterprise_id=enterprise_id, assignee_ids=assignee_ids)
    return jsonify({"code": 0, "data": stats})


# ============================================================
# 任务列表 / 详情 / 历史保单
# ============================================================

@renewal_bp.route("/list", methods=["GET"])
@login_required
def get_list():
    enterprise_id, assignee_ids = _resolve_scope()
    filters = _parse_list_filters(request)
    assignee_ids = _apply_explicit_assignee_filter(assignee_ids, filters["assignee"])
    page = int(request.args.get("page", 1))
    page_size = min(int(request.args.get("page_size", 20)), 200)
    result = rdb.list_tasks(
        _db, page=page, page_size=page_size, enterprise_id=enterprise_id,
        assignee_ids=assignee_ids, status=filters["status"],
        date_start=filters["date_start"], date_end=filters["date_end"],
        keyword=filters["keyword"],
    )
    return jsonify({"code": 0, "data": result})


def _query_history(customer_key: str, enterprise_id):
    """按车牌或VIN查该客户的全部历史保单（不限90天窗口，实时查询保证完整性）。"""
    conn = _db.pool.connection()
    try:
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        sql = (
            "SELECT r.id AS record_id, r.filename, r.created_at, "
            "pf.policy_no, pf.company_short, pf.policy_type, pf.plate_no, pf.vin, "
            "pf.start_date, pf.end_date, pf.end_date_iso, pf.total_premium "
            "FROM insurance_records r JOIN insurance_policy_fields pf ON r.id = pf.record_id "
            "WHERE (pf.plate_no = %s OR pf.vin = %s) "
            "AND r.deleted_at IS NULL AND r.status = 'done' AND r.doc_category = '保单'"
        )
        params = [customer_key, customer_key]
        if enterprise_id is not None:
            sql += " AND r.enterprise_id = %s"
            params.append(enterprise_id)
        sql += " ORDER BY pf.end_date_iso DESC"
        cursor.execute(sql, params)
        rows = cursor.fetchall()
        for r in rows:
            r["created_at"] = str(r["created_at"])
            if r.get("end_date_iso") is not None:
                r["end_date_iso"] = str(r["end_date_iso"])
        return rows
    finally:
        conn.close()


@renewal_bp.route("/detail/<int:task_id>", methods=["GET"])
@login_required
def get_detail(task_id):
    task = rdb.get_task(_db, task_id)
    if not task:
        return jsonify({"code": 404, "msg": "任务不存在"}), 404
    if not _can_operate_task(task):
        return jsonify({"code": 403, "msg": "权限不足"}), 403
    history = _query_history(task["customer_key"], task.get("enterprise_id"))
    followups = rdb.get_followups(_db, task_id)
    task["created_at"] = str(task["created_at"])
    task["updated_at"] = str(task["updated_at"])
    for f in ("end_date", "remind_at", "hint_end_date"):
        if task.get(f):
            task[f] = str(task[f])
    return jsonify({"code": 0, "data": {"task": task, "history": history, "followups": followups}})


@renewal_bp.route("/history/<path:customer_key>", methods=["GET"])
@login_required
def get_history(customer_key):
    enterprise_id = _get_enterprise_id_filter()
    history = _query_history(customer_key, enterprise_id)
    return jsonify({"code": 0, "data": {"history": history}})


# ============================================================
# 可分配人员列表
# ============================================================

@renewal_bp.route("/assignees", methods=["GET"])
@login_required
def get_assignees():
    role = g.current_user["role"]
    if role == ROLE_SUPER_ADMIN:
        eid_raw = request.args.get("enterprise_id", "").strip()
        if not eid_raw:
            return jsonify({"code": 400, "msg": "请提供 enterprise_id"}), 400
        try:
            eid = int(eid_raw)
        except ValueError:
            return jsonify({"code": 400, "msg": "enterprise_id 非法"}), 400
    elif role == ROLE_ENTERPRISE:
        eid = g.current_user.get("parent_id")
    else:
        # 员工及其他角色：不提供转派对象列表（员工无法互相转派任务）
        return jsonify({"code": 0, "data": {"assignees": []}})

    if not eid:
        return jsonify({"code": 0, "data": {"assignees": []}})

    conn = _db.pool.connection()
    try:
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        # 显式只取 enterprise/employee 两种角色，排除 sales/channel_sales/personal_sales
        cursor.execute(
            "SELECT id, name, role FROM users WHERE parent_id = %s "
            "AND role IN ('enterprise','employee') AND enabled = 1 "
            "ORDER BY role DESC, name",
            (eid,),
        )
        rows = cursor.fetchall()
        return jsonify({"code": 0, "data": {"assignees": rows}})
    finally:
        conn.close()


# ============================================================
# 分配 / 跟进 / 状态变更
# ============================================================

@renewal_bp.route("/assign", methods=["POST"])
@login_required
def assign():
    guard = _require_enterprise_or_admin()
    if guard:
        return guard
    body = request.get_json(force=True) or {}
    task_ids = body.get("task_ids") or []
    assignee = body.get("assignee")
    if not task_ids or not assignee:
        return jsonify({"code": 400, "msg": "task_ids 和 assignee 不能为空"}), 400

    conn = _db.pool.connection()
    try:
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        cursor.execute("SELECT id, name FROM users WHERE id = %s", (assignee,))
        user = cursor.fetchone()
    finally:
        conn.close()
    if not user:
        return jsonify({"code": 400, "msg": "指定的跟进人不存在"}), 400

    # enterprise 角色只能分配企业内部任务，做一次范围校验，避免越权分配到别的企业的任务
    if g.current_user["role"] == ROLE_ENTERPRISE:
        my_eid = g.current_user.get("parent_id")
        for tid in task_ids:
            t = rdb.get_task(_db, tid)
            if not t or t.get("enterprise_id") != my_eid:
                return jsonify({"code": 403, "msg": f"任务 {tid} 不属于本企业，无权分配"}), 403

    rdb.assign_tasks(_db, task_ids, assignee, user.get("name") or "")
    for tid in task_ids:
        rdb.add_followup(
            _db, tid, g.current_user["user_id"], _operator_name(), "assign",
            content=f"分配跟进人：{user.get('name') or assignee}",
        )
    return jsonify({"code": 0, "msg": f"已分配 {len(task_ids)} 条任务"})


@renewal_bp.route("/followup", methods=["POST"])
@login_required
def followup():
    body = request.get_json(force=True) or {}
    task_id = body.get("task_id")
    content = (body.get("content") or "").strip()
    next_remind_at = body.get("next_remind_at") or None
    if not task_id or not content:
        return jsonify({"code": 400, "msg": "task_id 和 content 不能为空"}), 400

    task = rdb.get_task(_db, task_id)
    if not task:
        return jsonify({"code": 404, "msg": "任务不存在"}), 404
    if not _can_operate_task(task):
        return jsonify({"code": 403, "msg": "权限不足"}), 403

    rdb.add_followup(
        _db, task_id, g.current_user["user_id"], _operator_name(), "followup",
        content=content, next_remind_at=next_remind_at,
    )
    return jsonify({"code": 0, "msg": "已添加跟进记录"})


@renewal_bp.route("/status", methods=["POST"])
@login_required
def change_status():
    body = request.get_json(force=True) or {}
    task_id = body.get("task_id")
    status = (body.get("status") or "").strip()
    remark = (body.get("remark") or "").strip()
    if status not in ("pending", "following", "won", "lost"):
        return jsonify({"code": 400, "msg": "非法状态"}), 400

    task = rdb.get_task(_db, task_id)
    if not task:
        return jsonify({"code": 404, "msg": "任务不存在"}), 404
    if not _can_operate_task(task):
        return jsonify({"code": 403, "msg": "权限不足"}), 403

    rdb.update_status(_db, task_id, status)
    rdb.set_hint(_db, task_id, None, None)  # 状态变更（尤其标won/忽略提示）清空续保检测提示
    rdb.add_followup(
        _db, task_id, g.current_user["user_id"], _operator_name(), "status_change",
        content=remark or f"状态变更为：{status}",
    )
    return jsonify({"code": 0, "msg": "状态已更新"})


# ============================================================
# 手动触发扫描
# ============================================================

@renewal_bp.route("/scan", methods=["POST"])
@admin_required
def manual_scan():
    try:
        import renewal_task_scan
        result = renewal_task_scan.run_scan(_db)
        return jsonify({"code": 0, "msg": "扫描完成", "data": result})
    except Exception as e:
        logger.exception("手动触发续保扫描失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


# ============================================================
# 配置（提醒窗口天数 / 客服二维码）
# ============================================================

@renewal_bp.route("/config", methods=["GET"])
@login_required
def get_config():
    window_days = get_insurance_config(_db, _CONFIG_KEY_WINDOW_DAYS, 90)
    qrcode_url = get_insurance_config(_db, _CONFIG_KEY_QRCODE_URL, "")
    return jsonify({"code": 0, "data": {"window_days": window_days, "qrcode_url": qrcode_url}})


@renewal_bp.route("/config", methods=["POST"])
@admin_required
def save_config():
    body = request.get_json(force=True) or {}
    if "window_days" in body:
        try:
            days = int(body["window_days"])
        except (TypeError, ValueError):
            return jsonify({"code": 400, "msg": "window_days 必须是整数"}), 400
        set_insurance_config(_db, _CONFIG_KEY_WINDOW_DAYS, days)
    return jsonify({"code": 0, "msg": "已保存"})


@renewal_bp.route("/qrcode/upload", methods=["POST"])
@admin_required
def upload_qrcode():
    if not _cos_storage:
        return jsonify({"code": 503, "msg": "COS存储未初始化"}), 503
    f = request.files.get("file")
    if not f:
        return jsonify({"code": 400, "msg": "缺少文件"}), 400
    data = f.read()
    ext = (f.filename or "").rsplit(".", 1)[-1].lower() if "." in (f.filename or "") else "png"
    if ext not in ("png", "jpg", "jpeg"):
        ext = "png"
    url = _cos_storage.upload_bytes(data, f"renewal/qrcode.{ext}")
    if not url:
        return jsonify({"code": 500, "msg": "上传失败"}), 500
    set_insurance_config(_db, _CONFIG_KEY_QRCODE_URL, url)
    return jsonify({"code": 0, "data": {"url": url}})


# ============================================================
# 导出 Excel（独立轻量实现，不复用 ledger_export.py 的环回HTTP模式：
# 续保任务字段少且固定，没有台账那套用户自定义列配置/公式系统的复用价值）
# ============================================================

_EXPORT_HEADERS = [
    ("plate_no", "车牌号"), ("vin", "车架号"), ("insured", "被保险人"),
    ("applicant", "投保人"), ("company_short", "承保公司"), ("policy_type", "险种"),
    ("end_date", "到期日"), ("total_premium", "上期保费"),
    ("assignee_name", "跟进人"), ("status", "状态"), ("remark", "备注"),
]
_STATUS_LABEL = {"pending": "待跟进", "following": "跟进中", "won": "已成交", "lost": "已流失", "expired": "已过期"}


@renewal_bp.route("/export", methods=["GET"])
@login_required
def export():
    import openpyxl
    from openpyxl.utils import get_column_letter

    enterprise_id, assignee_ids = _resolve_scope()
    filters = _parse_list_filters(request)
    assignee_ids = _apply_explicit_assignee_filter(assignee_ids, filters["assignee"])

    result = rdb.list_tasks(
        _db, page=1, page_size=_EXPORT_ROW_LIMIT + 1, enterprise_id=enterprise_id,
        assignee_ids=assignee_ids, status=filters["status"],
        date_start=filters["date_start"], date_end=filters["date_end"],
        keyword=filters["keyword"],
    )
    tasks = result["tasks"]
    if len(tasks) > _EXPORT_ROW_LIMIT:
        return jsonify({"code": 400, "msg": f"数据量超过{_EXPORT_ROW_LIMIT}条，请缩小筛选范围后再导出"}), 400

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "续保任务"
    ws.append([label for _, label in _EXPORT_HEADERS])
    for t in tasks:
        row = []
        for key, _ in _EXPORT_HEADERS:
            v = t.get(key, "")
            if key == "status":
                v = _STATUS_LABEL.get(v, v)
            row.append(str(v) if v is not None else "")
        ws.append(row)
    for i in range(1, len(_EXPORT_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"续保任务_{len(tasks)}条.xlsx"
    resp = send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
    return resp
