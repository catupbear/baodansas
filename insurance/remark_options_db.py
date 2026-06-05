"""
备注快捷选择 - 选项池数据层。

- read_sheet_names: 读取 Excel 全部 sheet 名（供前端选 sheet）
- parse_sheet_rows: 按表头(经办人/投保公司/备注信息/险种名称)解析指定 sheet 的行
- import_options: 清空当前 scope 旧池 + 批量插入新池（同事务）
- query_options: 按 handler/company/policy_type/keyword + mode(exact/fuzzy) 查询候选 remark
- clear_options / get_stats
"""

import io
import logging
from datetime import datetime

import openpyxl
import pymysql
import pymysql.cursors

logger = logging.getLogger(__name__)

# Excel 表头 → 内部字段
_HEADER_MAP = {
    "经办人": "handler",
    "投保公司": "company",
    "备注信息": "remark",
    "险种名称": "policy_type",
    "险种": "policy_type",
}


def read_sheet_names(file_bytes: bytes) -> list:
    """返回 Excel 全部 sheet 名（顺序与文件一致）。"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


# 期望存在的筛选列（remark 为必需，单独校验）：内部 key → 提示用列名
_EXPECTED_FILTER_COLS = [("handler", "经办人"), ("company", "投保公司"), ("policy_type", "险种名称")]


def parse_sheet_rows(file_bytes: bytes, sheet_name: str) -> dict:
    """
    解析指定 sheet，返回 {"rows": [...], "missing": [缺失的筛选列名], "found": [存在的筛选列名]}。
    rows 每项为 {"handler","company","policy_type","remark"}。
    按表头名定位列（容忍尾部多余空列、列顺序变化）；remark 为空的行跳过。
    缺少「备注信息」列直接报错；缺少经办人/投保公司/险种名称仅记录到 missing，不报错。
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"sheet 不存在: {sheet_name}")
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return {"rows": [], "missing": [c[1] for c in _EXPECTED_FILTER_COLS], "found": []}
        # 表头列名 → 列索引
        col_idx = {}
        for i, name in enumerate(header):
            key = _HEADER_MAP.get(str(name).strip()) if name is not None else None
            if key and key not in col_idx:
                col_idx[key] = i
        if "remark" not in col_idx:
            raise ValueError("未找到「备注信息」列")

        missing = [label for key, label in _EXPECTED_FILTER_COLS if key not in col_idx]
        found = [label for key, label in _EXPECTED_FILTER_COLS if key in col_idx]

        def _cell(row, key):
            i = col_idx.get(key)
            if i is None or i >= len(row) or row[i] is None:
                return ""
            return str(row[i]).strip()

        out = []
        for row in rows_iter:
            if row is None:
                continue
            remark = _cell(row, "remark")
            if not remark:
                continue
            out.append({
                "handler": _cell(row, "handler"),
                "company": _cell(row, "company"),
                "policy_type": _cell(row, "policy_type"),
                "remark": remark,
            })
        return {"rows": out, "missing": missing, "found": found}
    finally:
        wb.close()


def _scope_cond(scope, scope_id, template_name):
    """统一 scope+模板 过滤条件（scope_id 为 None 时用 IS NULL）。"""
    if scope_id is None:
        return "scope = %s AND scope_id IS NULL AND template_name = %s", [scope, template_name]
    return "scope = %s AND scope_id = %s AND template_name = %s", [scope, str(scope_id), template_name]


def import_options(db, scope, scope_id, template_name, rows: list) -> int:
    """清空当前 scope+模板 旧池 + 批量插入新池（去重 handler+company+policy_type+remark）。返回插入条数。"""
    batch_no = datetime.now().strftime("%Y%m%d%H%M%S")
    seen = set()
    dedup = []
    for r in rows:
        k = (r["handler"], r["company"], r["policy_type"], r["remark"])
        if k in seen:
            continue
        seen.add(k)
        dedup.append(r)

    cond, params = _scope_cond(scope, scope_id, template_name)
    conn = db.pool.connection()
    try:
        cursor = conn.cursor()
        cursor.execute(f"DELETE FROM insurance_remark_options WHERE {cond}", params)
        if dedup:
            sid = None if scope_id is None else str(scope_id)
            cursor.executemany(
                "INSERT INTO insurance_remark_options "
                "(scope, scope_id, template_name, handler, company, policy_type, remark, batch_no) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                [(scope, sid, template_name, r["handler"], r["company"], r["policy_type"], r["remark"], batch_no)
                 for r in dedup]
            )
        conn.commit()
        logger.info("备注选项池导入: scope=%s scope_id=%s template=%s 插入=%d batch=%s",
                    scope, scope_id, template_name, len(dedup), batch_no)
        return len(dedup)
    finally:
        conn.close()


def query_options(db, scope, scope_id, template_name, handler="", company="", policy_type="",
                  keyword="", mode="exact", limit=500) -> list:
    """
    按键查询候选 remark。mode=exact 精确等值；mode=fuzzy 子串包含。
    任一键为空则该键不参与过滤。keyword 始终按 remark 子串过滤。
    返回去重后的 remark 字符串列表（保序）。
    """
    cond, params = _scope_cond(scope, scope_id, template_name)
    where = [cond]

    def _add(field, value):
        value = (value or "").strip()
        if not value:
            return
        if mode == "fuzzy":
            where.append(f"{field} LIKE %s")
            params.append(f"%{value}%")
        else:
            where.append(f"{field} = %s")
            params.append(value)

    _add("handler", handler)
    _add("company", company)
    _add("policy_type", policy_type)
    if keyword and keyword.strip():
        where.append("remark LIKE %s")
        params.append(f"%{keyword.strip()}%")

    sql = ("SELECT remark FROM insurance_remark_options WHERE " + " AND ".join(where) +
           " ORDER BY id LIMIT %s")
    params.append(int(limit))
    conn = db.pool.connection()
    try:
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        cursor.execute(sql, params)
        rows = cursor.fetchall()
        seen = set()
        out = []
        for row in rows:
            rm = row["remark"]
            if rm and rm not in seen:
                seen.add(rm)
                out.append(rm)
        return out
    finally:
        conn.close()


def clear_options(db, scope, scope_id, template_name) -> int:
    """清空当前 scope+模板 选项池。返回删除条数。"""
    cond, params = _scope_cond(scope, scope_id, template_name)
    conn = db.pool.connection()
    try:
        cursor = conn.cursor()
        n = cursor.execute(f"DELETE FROM insurance_remark_options WHERE {cond}", params)
        conn.commit()
        return n
    finally:
        conn.close()


def get_stats(db, scope, scope_id, template_name) -> dict:
    """返回 {"count": N, "last_import": "YYYYMMDDHHMMSS" or ""}。"""
    cond, params = _scope_cond(scope, scope_id, template_name)
    conn = db.pool.connection()
    try:
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        cursor.execute(
            f"SELECT COUNT(*) AS cnt, MAX(batch_no) AS last_batch "
            f"FROM insurance_remark_options WHERE {cond}", params)
        row = cursor.fetchone() or {}
        return {"count": int(row.get("cnt") or 0), "last_import": row.get("last_batch") or ""}
    finally:
        conn.close()
