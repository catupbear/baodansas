#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
保单识别模块 — Flask API 蓝图
提供保单识别记录查询、手动上传识别、重试/同步、配置管理、字段映射、钉钉目标管理等接口
"""

import base64
import hashlib
import io
import json
import logging
import re
import time
import uuid

import pymysql
import pymysql.cursors
from flask import Blueprint, g, jsonify, render_template, request, send_file

from .db import (
    backfill_records_by_sources,
    get_all_insurance_config,
    get_company_stats,
    get_company_detail_stats,
    get_insurance_config,
    get_insurance_record,
    get_insurance_stats,
    get_record_company_list,
    query_insurance_records,
    save_insurance_record,
    set_insurance_config,
    soft_delete_record,
    update_insurance_record,
    update_records_abnormal_flag,
    upsert_insurance_record_by_policy,
)
from .field_config_db import (
    list_templates as list_field_config_templates, list_all_templates,
    rename_template, update_template_visibility, delete_template as delete_template_db,
    get_template_config, save_full_template,
    get_active_template, set_active_template,
    get_effective_config, apply_user_config_to_fields, match_fee_formulas, _format_date,
    get_column_config, save_column_config, delete_column_config,
    get_user_fixed_values, save_user_fixed_values,
    get_remark_selector_config, save_remark_selector_config,
    get_merge_by_plate, save_merge_by_plate,
    get_plate_format_pingan, save_plate_format_pingan,
    MERGE_SHARED_FIELDS, MERGE_SPLIT_FIELDS, MERGE_PREFIXES, MERGE_EXTRA_COLUMNS,
    DEFAULT_COLUMNS,
)
from .monitor_config_db import (
    list_monitor_configs,
    get_monitor_config,
    create_monitor_config,
    update_monitor_config,
    delete_monitor_config,
)
from .dingtalk_sync import DingTalkTableService, parse_dingtalk_doc_url
from .field_mapping import (
    OUTPUT_COLUMNS,
    apply_mapping,
    get_available_ocr_fields,
    load_mapping_config,
    save_mapping_config,
)
from .policy_parser import get_extraction_rules, parse_policy_text, parse_policy_text_multi, get_policy_type_code
from .ocr_service import extract_text_from_pdf
from . import remark_options_db
from . import internal_notes_db
from auth.decorators import login_required
from auth.db import ROLE_SUPER_ADMIN, ROLE_ENTERPRISE, ROLE_EMPLOYEE, ROLE_SALES, get_sender_user_name_map, get_sender_alias_map
from core.notify import notify_error

logger = logging.getLogger(__name__)

# ------------------------------------------------------------------ #
# 轻量 TTL 内存缓存：减少每次列表请求都重复读 DB 的用户配置查询
# ------------------------------------------------------------------ #
_cfg_cache: dict = {}
_CFG_TTL = 60  # 秒


def _cache_get(key):
    entry = _cfg_cache.get(key)
    if entry and time.time() < entry[1]:
        return entry[0], True
    return None, False


def _cache_set(key, value, ttl=_CFG_TTL):
    _cfg_cache[key] = (value, time.time() + ttl)
    return value


def _norm_plate(plate) -> str:
    """车牌归一化（仅用于匹配，不改变显示值）：去掉连字符和空格。
    数据库存储的车牌无连字符，但前端对平安等保单会格式化为「粤B-CM3220」，
    匹配交强到期时间时需统一归一化，避免「粤B-CM3220」与「粤BCM3220」不匹配。"""
    return (plate or "").replace("-", "").replace("—", "").replace(" ", "")


_DATE_ONLY_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def _date_only_key(raw) -> str:
    """把日期字符串统一成 YYYY-MM-DD 用于比较是否同一天；解析失败则原样返回兜底比较。"""
    if not raw:
        return ""
    s = _format_date(str(raw), "YYYY-MM-DD", False)
    return s if _DATE_ONLY_RE.match(s) else str(raw).strip()


def _combine_start_dates(commercial_raw, compulsory_raw, fmt, no_pad) -> str:
    """"交强与交商起保时间"：商业险/交强险起保日期同一天只显示商业险那个；
    不同则显示"商业险起保日期交交强险起保月/日"（月/日不补零，如 交7/16）。"""
    if not commercial_raw and not compulsory_raw:
        return ""
    if not compulsory_raw:
        return _format_date(str(commercial_raw), fmt, no_pad) if fmt else str(commercial_raw)
    if not commercial_raw:
        return _format_date(str(compulsory_raw), fmt, no_pad) if fmt else str(compulsory_raw)
    main = _format_date(str(commercial_raw), fmt, no_pad) if fmt else str(commercial_raw)
    if _date_only_key(commercial_raw) == _date_only_key(compulsory_raw):
        return main
    sub = _format_date(str(compulsory_raw), "MM/DD", True)
    return f"{main}交{sub}"


# 创建蓝图，不设 url_prefix（路由中已写全路径）
insurance_bp = Blueprint("insurance", __name__)

# 模块级别的依赖注入变量
_db = None
_handler = None
_contacts = None


def init_insurance_api(db, handler=None, contacts=None):
    """
    初始化保单识别 API 所需的依赖。
    db      : 数据库实例（提供 db.pool.connection()）
    handler : InsuranceHandler 实例
    contacts: 联系人/群组解析服务实例
    """
    global _db, _handler, _contacts
    _db = db
    _handler = handler
    _contacts = contacts


def _try_sdk_download(record: dict) -> bytes:
    """尝试通过企业微信 SDK 重新下载媒体文件，返回 bytes 或 b""。"""
    from flask import current_app

    msg_seq = record.get("msg_seq")
    if not msg_seq:
        logger.warning("SDK 重新下载跳过: record_id=%s 无 msg_seq", record.get("id"))
        return b""

    # 从 messages 表获取 sdkfileid 和 corpid（可能有多条同 seq 不同 corpid）
    conn = _db.pool.connection()
    try:
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        cursor.execute(
            "SELECT parsed_content, corpid FROM messages WHERE seq = %s",
            (msg_seq,)
        )
        msg_rows = cursor.fetchall()
    finally:
        conn.close()

    if not msg_rows:
        logger.warning("SDK 重新下载跳过: seq=%s 在 messages 表中不存在", msg_seq)
        return b""

    filename = record.get("filename", "")
    logger.info("SDK 重新下载: seq=%s, filename=%s, 找到 %d 条消息记录, corpids=%s",
                msg_seq, filename, len(msg_rows), [r.get("corpid", "") for r in msg_rows])

    # 收集所有候选的 sdkfileid（不同 corpid 可能不同）
    wxbot_ext = current_app.extensions.get("wxbot", {})

    # 构建 corpid → SDK 映射
    sdk_map = {}
    if _handler and _handler.finance_sdk:
        # 主企业
        main_fetcher = wxbot_ext.get("fetcher")
        main_corpid = getattr(main_fetcher, "corpid", "") if main_fetcher else ""
        sdk_map[main_corpid] = ("主企业", _handler.finance_sdk, _handler.sdk_config)

    for ef in wxbot_ext.get("extra_fetchers", []):
        f = ef.get("fetcher")
        if f and f.finance_sdk:
            conf = {"proxy": f.proxy, "proxy_passwd": f.passwd, "timeout": f.timeout}
            sdk_map[f.corpid] = (ef.get("name", "额外企业"), f.finance_sdk, conf)

    logger.info("SDK 重新下载: 可用SDK=%s", list(sdk_map.keys()))

    # 对每条消息记录，用对应 corpid 的 SDK 尝试下载
    for msg_row in msg_rows:
        parsed_content = msg_row.get("parsed_content") or "{}"
        if isinstance(parsed_content, str):
            try:
                parsed_content = json.loads(parsed_content)
            except (TypeError, json.JSONDecodeError):
                continue

        sdkfileid = parsed_content.get("sdkfileid", "")
        if not sdkfileid:
            continue

        msg_corpid = msg_row.get("corpid", "")

        # 优先用匹配 corpid 的 SDK，然后尝试其他所有 SDK
        sdks_to_try = []
        if msg_corpid in sdk_map:
            sdks_to_try.append(sdk_map[msg_corpid])
        for cid, entry in sdk_map.items():
            if entry not in sdks_to_try:
                sdks_to_try.append(entry)

        for name, sdk, conf in sdks_to_try:
            proxy = conf.get("proxy", "")
            passwd = conf.get("proxy_passwd", "")
            timeout = conf.get("timeout", 30)
            try:
                ret, data = sdk.get_media_data_bytes(sdkfileid, proxy, passwd, timeout)
                if ret == 0 and data:
                    logger.info("SDK 重新下载成功: [%s], seq=%s, corpid=%s, 大小=%d",
                                name, msg_seq, msg_corpid, len(data))
                    return data
                logger.info("SDK 重新下载失败: [%s] ret=%d, seq=%s, corpid=%s",
                            name, ret, msg_seq, msg_corpid)
            except Exception as e:
                logger.warning("SDK 重新下载异常: [%s] seq=%s: %s", name, msg_seq, e)

    logger.warning("所有 SDK 均无法下载, seq=%s, filename=%s", msg_seq, filename)
    return b""


def recover_stuck_processing_records(timeout_minutes=10):
    """
    自愈：扫描长时间卡在 status='processing' 的记录（多因 gunicorn 重启时后台队列消费线程
    被杀死、内存中的任务丢失导致），通过内部环回调用 /api/insurance/reocr 自动重新识别恢复。
    以超管账号身份签发内部调用凭证，与记录归属的 user_id 无关（reocr 接口本身不做归属校验）。
    """
    import requests

    from auth.jwt_utils import generate_token

    if not _db:
        return

    conn = _db.pool.connection()
    try:
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        cursor.execute(
            "SELECT id FROM insurance_records "
            "WHERE status = 'processing' AND deleted_at IS NULL "
            "AND created_at < DATE_SUB(NOW(), INTERVAL %s MINUTE)",
            (timeout_minutes,)
        )
        stuck_ids = [row["id"] for row in cursor.fetchall()]
    finally:
        conn.close()

    if not stuck_ids:
        return

    logger.warning("自愈扫描：发现 %d 条卡在 processing 状态超过 %d 分钟的记录: %s",
                    len(stuck_ids), timeout_minutes, stuck_ids)

    conn = _db.pool.connection()
    try:
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        cursor.execute(
            "SELECT id, phone, role FROM users WHERE role = %s AND enabled = 1 LIMIT 1",
            (ROLE_SUPER_ADMIN,)
        )
        admin = cursor.fetchone()
    finally:
        conn.close()

    if not admin:
        logger.error("自愈扫描：未找到可用超管账号，无法生成内部调用凭证，跳过本次恢复")
        return

    token = generate_token(admin["id"], admin["phone"], admin["role"])
    headers = {"Authorization": f"Bearer {token}"}

    recovered, expired, failed = 0, 0, 0
    for rid in stuck_ids:
        try:
            resp = requests.post(
                f"http://127.0.0.1:8090/api/insurance/reocr/{rid}",
                headers=headers, timeout=90,
            )
            body = {}
            try:
                body = resp.json()
            except Exception:
                pass
            if resp.ok and body.get("code") == 0:
                recovered += 1
                logger.info("自愈扫描：记录 %d 恢复成功", rid)
            elif body.get("need_resend"):
                # 接口明确告知文件已彻底过期(SDK/COS都拿不到了)，不是网络抖动等可重试的情况，
                # 再怎么重试也不会成功——标记为 failed 并写清原因，停止后续每次自愈都重复失败，
                # 需要客户在群里重新发送该文件才能恢复
                update_insurance_record(_db, rid, {
                    "status": "failed",
                    "error_message": "自愈失败：文件已过期无法下载，需在群内重新发送该保单文件",
                })
                expired += 1
                logger.warning("自愈扫描：记录 %d 文件已彻底过期，标记为失败(需客户重新发送)", rid)
            else:
                failed += 1
                logger.warning("自愈扫描：记录 %d 恢复失败 status=%s body=%s",
                                rid, resp.status_code, resp.text[:300])
        except Exception as e:
            failed += 1
            logger.warning("自愈扫描：记录 %d 恢复异常: %s", rid, e)

    logger.warning("自愈扫描完成：成功恢复 %d 条，彻底过期标记失败 %d 条，其他失败 %d 条",
                    recovered, expired, failed)


# ============================================================
# 全局鉴权：所有保单识别 API 均需登录
# ============================================================

@insurance_bp.before_request
def _require_login():
    """保单识别 API 统一鉴权"""
    from auth.jwt_utils import verify_token
    from auth.db import get_user_by_id

    # 页面路由不需要鉴权（由前端 JS 处理登录跳转）
    if request.path == "/insurance/settings":
        return
    # 经营报告推送给客户查看：报告数据(format=json，凭链接签名) 和 反馈接口 不需登录
    if request.path == "/api/insurance/enterprise-report" and request.args.get("format") == "json":
        return
    if request.path == "/api/insurance/report-feedback":
        return
    # 使用手册查看不需要登录(公开给潜在客户/未注册用户看)，编辑/传图仍需登录+内部超管(各自接口内部再判断)
    if request.path == "/api/insurance/manual" and request.method == "GET":
        return

    auth_header = request.headers.get("Authorization", "")
    token = auth_header[7:] if auth_header.startswith("Bearer ") else request.args.get("token")
    if not token:
        return jsonify({"code": 401, "msg": "未登录"}), 401

    payload = verify_token(token)
    if not payload:
        return jsonify({"code": 401, "msg": "登录已过期，请重新登录"}), 401

    user = get_user_by_id(_db, payload["user_id"])
    if not user or not user.get("enabled"):
        return jsonify({"code": 401, "msg": "账号已禁用"}), 401

    from flask import g
    g.current_user = {
        "user_id": payload["user_id"],
        "phone": payload.get("phone", payload.get("username", "")),
        "role": payload["role"],
        "parent_id": user.get("parent_id"),
        "activated": int(user.get("activated", 1)),
    }

    # 未激活用户（自助注册待管理员激活）：禁止上传/识别保单等工作类接口
    if not user.get("activated", 1) and payload.get("role") != ROLE_SUPER_ADMIN:
        _ACTIVATE_GUARD_PREFIXES = (
            "/api/insurance/ocr",
            "/api/insurance/parse",
            "/api/insurance/retry",
            "/api/insurance/reupload-reocr",
            "/api/insurance/reocr",
            "/api/insurance/batch-reocr-sync",
            "/api/insurance/sync",
            "/api/insurance/export",
            "/api/insurance/batch-download",
        )
        if any(request.path.startswith(p) for p in _ACTIVATE_GUARD_PREFIXES):
            return jsonify({
                "code": 403,
                "msg": "您的账号尚未激活，请联系管理员激活后再使用",
                "not_activated": True,
            }), 403


def _get_user_ids_filter():
    """根据当前用户角色返回 user_ids 过滤列表，超管和企业管理员返回 None（不按用户过滤）"""
    from flask import g
    role = g.current_user["role"]
    uid = g.current_user["user_id"]
    if role == ROLE_SUPER_ADMIN:
        return None
    # 企业管理员：不按 user_id 过滤，由 enterprise_id 过滤
    if role == ROLE_ENTERPRISE:
        return None
    # 员工：只看自己的记录
    return [uid]


def _get_enterprise_id_filter():
    """返回当前用户的 enterprise_id，超管返回 None（不过滤），其他角色按企业过滤"""
    from flask import g
    role = g.current_user["role"]
    if role == ROLE_SUPER_ADMIN:
        return None
    # 企业管理员和员工：parent_id 指向 enterprises 表，优先使用；无 parent_id 则用自身 user_id 兜底
    parent_id = g.current_user.get("parent_id")
    return parent_id if parent_id else g.current_user["user_id"]


def _get_user_scope():
    """根据当前用户角色返回 (scope, scope_id)"""
    from flask import g
    user = g.current_user
    role = user["role"]
    if role == "super_admin":
        return "global", None
    elif role == "enterprise":
        return "enterprise", user.get("parent_id") or user["user_id"]
    else:
        return "user", user["user_id"]


# ============================================================
# 全局异常处理：捕获未处理的 500 错误并通知
# ============================================================

@insurance_bp.app_errorhandler(500)
def _handle_500(error):
    notify_error("Web服务", request.path, f"500 内部错误: {error}", f"method={request.method}")
    return jsonify({"error": "服务器内部错误"}), 500


# ============================================================
# 识别记录
# ============================================================

@insurance_bp.route("/api/insurance/records", methods=["GET"])
def list_records():
    """分页查询保单识别记录，支持按群、状态、关键词、来源筛选"""
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("page_size", 20))
    roomid = request.args.get("roomid", "")
    status = request.args.get("status", "")
    keyword = request.args.get("keyword", "")
    source = request.args.get("source", "")
    source_type = request.args.get("source_type", "")
    sender = request.args.get("sender", "")
    ocr_engine = request.args.get("ocr_engine", "")
    doc_category = request.args.get("doc_category", "")
    is_abnormal = request.args.get("is_abnormal", "")
    review_status = request.args.get("review_status", "")
    if g.current_user.get("role") != "super_admin":
        review_status = ""  # 核对状态筛选仅超管可用，防止非超管借此推断内部备注情况
        # 注：此处用字面量而非 ROLE_SUPER_ADMIN 名字，因本函数体后段有局部 `from auth.db import ROLE_SUPER_ADMIN`
        # （第373行附近，早于此处执行会导致 UnboundLocalError：Python 视其为函数局部名）
    company_short = request.args.get("company_short", "")
    policy_type = request.args.get("policy_type", "")
    date_start = request.args.get("date_start", "")
    date_end = request.args.get("date_end", "")
    updated_at_date_start = request.args.get("updated_at_date_start", "")
    updated_at_date_end = request.args.get("updated_at_date_end", "")
    # 关联表模糊搜索
    search_company = request.args.get("search_company", "")
    search_policy_no = request.args.get("search_policy_no", "")
    search_plate_no = request.args.get("search_plate_no", "")
    search_applicant = request.args.get("search_applicant", "")
    search_insured = request.args.get("search_insured", "")
    search_salesperson = request.args.get("search_salesperson", "")
    # 关联表日期筛选
    sign_date_start = request.args.get("sign_date_start", "")
    sign_date_end = request.args.get("sign_date_end", "")
    start_date_start = request.args.get("start_date_start", "")
    start_date_end = request.args.get("start_date_end", "")
    end_date_start = request.args.get("end_date_start", "")
    end_date_end = request.args.get("end_date_end", "")
    dedup = request.args.get("dedup", "") == "1"
    sort_by = request.args.get("sort_by", "")
    sort_order = request.args.get("sort_order", "desc")
    # 超管可通过 enterprise_id 参数筛选指定企业，其他角色由权限控制
    req_enterprise_id = request.args.get("enterprise_id", "").strip()

    try:
        # 超管传了 enterprise_id 参数则用参数值，否则走权限控制
        from auth.db import ROLE_SUPER_ADMIN
        if g.current_user.get("role") == ROLE_SUPER_ADMIN and req_enterprise_id:
            effective_enterprise_id = int(req_enterprise_id)
            effective_user_ids = None
        else:
            effective_enterprise_id = _get_enterprise_id_filter()
            effective_user_ids = _get_user_ids_filter()

        result = query_insurance_records(
            _db,
            page=page,
            page_size=page_size,
            roomid=roomid,
            status=status,
            keyword=keyword,
            source=source,
            source_type=source_type,
            sender=sender,
            ocr_engine=ocr_engine,
            doc_category=doc_category,
            is_abnormal=is_abnormal,
            company_short=company_short,
            policy_type=policy_type,
            date_start=date_start,
            date_end=date_end,
            updated_at_date_start=updated_at_date_start,
            updated_at_date_end=updated_at_date_end,
            user_ids=effective_user_ids,
            enterprise_id=effective_enterprise_id,
            search_company=search_company,
            search_policy_no=search_policy_no,
            search_plate_no=search_plate_no,
            search_applicant=search_applicant,
            search_insured=search_insured,
            search_salesperson=search_salesperson,
            sign_date_start=sign_date_start,
            sign_date_end=sign_date_end,
            start_date_start=start_date_start,
            start_date_end=start_date_end,
            end_date_start=end_date_start,
            end_date_end=end_date_end,
            dedup=dedup,
            review_status=review_status,
            sort_by=sort_by,
            sort_order=sort_order,
        )
        # 列表返回 display_fields（轻量映射字段）替代 parsed_fields
        # 获取用户字段配置，应用简称/日期格式/公式计算（结果 TTL 缓存 60 秒）
        user = g.current_user
        _uid, _role, _pid = user["user_id"], user["role"], user.get("parent_id")
        # 超管选了企业筛选时，列展示配置（列布局/简称/公式/固定值）必须按被查看的企业解析，
        # 不能用超管自己账号的配置——否则切换企业后列表列/字段格式还是超管自己的默认样式，
        # 跟该企业自己配置的模板对不上（与 export_excel 的同类修复一致）。
        if _role == ROLE_SUPER_ADMIN and effective_enterprise_id:
            from auth.db import get_enterprise_admin_user
            _ent_admin = get_enterprise_admin_user(_db, effective_enterprise_id)
            if _ent_admin:
                _uid, _role, _pid = _ent_admin["id"], _ent_admin["role"], _ent_admin["parent_id"]
        _eff_key = ("eff_cfg", _uid, _role, _pid)
        user_config, _hit = _cache_get(_eff_key)
        if not _hit:
            user_config = _cache_set(_eff_key, get_effective_config(_db, _uid, _role, _pid))
        # 注入用户个人的自定义字段固定值（独立于模板）
        _col_key = ("col_cfg", "list_columns", _uid, _role, _pid)
        list_col_cfg, _hit = _cache_get(_col_key)
        if not _hit:
            list_col_cfg = _cache_set(_col_key, get_column_config(_db, "list_columns", _uid, _role, _pid))
        _fv_key = ("fixed_vals", _uid)
        fixed_vals, _hit = _cache_get(_fv_key)
        if not _hit:
            fixed_vals = _cache_set(_fv_key, get_user_fixed_values(_db, _uid))
        if fixed_vals:
            if user_config is None:
                user_config = {}
            user_config = dict(user_config)
            user_config["fixed_values"] = fixed_vals
        has_config = user_config and any(user_config.get(k) for k in user_config)

        # 构建 sender → 用户姓名映射（TTL 缓存 30 秒）
        sender_name_map = {}
        try:
            ent_id = _get_enterprise_id_filter()
            _sm_key = ("sender_map", ent_id)
            sender_name_map, _hit = _cache_get(_sm_key)
            if not _hit:
                sender_name_map = _cache_set(_sm_key, get_sender_user_name_map(_db, ent_id), ttl=30)
        except Exception:
            logger.debug("获取 sender→用户姓名映射失败，跟单人字段将为空")

        # sender → 绑定别名 映射：管理员在「绑定发送人」设的别名，用于覆盖记录的发送人显示名
        sender_alias_map = {}
        try:
            _sa_key = ("sender_alias_map", ent_id)
            sender_alias_map, _hit = _cache_get(_sa_key)
            if not _hit:
                sender_alias_map = _cache_set(_sa_key, get_sender_alias_map(_db, ent_id), ttl=30)
        except Exception:
            logger.debug("获取 sender→绑定别名映射失败，发送人显示名不覆盖")

        # 批量查询无 sender 记录的上传用户姓名（避免每条都单独查 DB）
        _upload_user_ids = {r["user_id"] for r in result.get("records", [])
                            if not r.get("sender") and r.get("user_id")}
        _upload_user_name_map: dict = {}
        if _upload_user_ids:
            try:
                _conn = _db.pool.connection()
                try:
                    _cur = _conn.cursor(pymysql.cursors.DictCursor)
                    _ph = ",".join(["%s"] * len(_upload_user_ids))
                    _cur.execute(f"SELECT id, name FROM users WHERE id IN ({_ph})", list(_upload_user_ids))
                    for _row in _cur.fetchall():
                        if _row.get("name"):
                            _upload_user_name_map[_row["id"]] = _row["name"]
                finally:
                    _conn.close()
            except Exception:
                logger.debug("批量查询上传用户姓名失败")

        # 批量查询记录所属企业名称（超管列表「所属企业」列展示用，避免逐条查 DB）
        _ent_ids = {r.get("enterprise_id") for r in result.get("records", [])
                    if r.get("enterprise_id")}
        _ent_name_map: dict = {}
        # TZ-008(鹏杰华智能)专属：车牌为空时用车架号(VIN)填入车牌显示
        _tz008_ids: set = set()
        if _ent_ids:
            try:
                _conn = _db.pool.connection()
                try:
                    _cur = _conn.cursor(pymysql.cursors.DictCursor)
                    _ph = ",".join(["%s"] * len(_ent_ids))
                    _cur.execute(f"SELECT id, name, enterprise_no FROM enterprises WHERE id IN ({_ph})", list(_ent_ids))
                    for _row in _cur.fetchall():
                        _ent_name_map[_row["id"]] = _row.get("name") or ""
                        _eno = re.sub(r"[^A-Za-z0-9]", "", (_row.get("enterprise_no") or "")).upper()
                        if _eno == "TZ008":
                            _tz008_ids.add(_row["id"])
                finally:
                    _conn.close()
            except Exception:
                logger.debug("批量查询企业名称失败")

        for record in result.get("records", []):
            # 注入所属企业名称（超管列表展示）
            record["enterprise_name"] = _ent_name_map.get(record.get("enterprise_id"), "")
            # 反序列化 display_fields
            if isinstance(record.get("display_fields"), str):
                try:
                    record["display_fields"] = json.loads(record["display_fields"])
                except (TypeError, json.JSONDecodeError):
                    record["display_fields"] = {}
            # 补充车主：display_fields 缺失时从关联表补充
            if not record.get("display_fields"):
                record["display_fields"] = {}
            # DEBUG：记录从DB读取后的交强到期时间
            _df_debug = record.get("display_fields", {})
            if _df_debug.get("交强到期时间"):
                logger.debug("[DEBUG-列表] record_id=%s DB读取后 交强到期时间=%s, manual_fields=%s",
                            record.get("id"), _df_debug.get("交强到期时间"), record.get("manual_fields"))
            pf_owner = record.pop("_pf_owner", None)
            if pf_owner and not record["display_fields"].get("车主"):
                record["display_fields"]["车主"] = pf_owner
            # 历史记录兼容：车牌号→车牌（旧映射配置未包含此规则时遗留）
            if not record["display_fields"].get("车牌") and record["display_fields"].get("车牌号"):
                record["display_fields"]["车牌"] = record["display_fields"]["车牌号"]
            # TZ-008 专属：车牌仍为空时，用车架号(VIN)填入车牌（未上牌车按车架号当车牌用）
            if record.get("enterprise_id") in _tz008_ids and not record["display_fields"].get("车牌"):
                _tz_vin = record["display_fields"].get("车架号") or record["display_fields"].get("车架号VIN")
                if _tz_vin:
                    record["display_fields"]["车牌"] = _tz_vin
            # 实时关联跟单人：通过 sender 查找绑定用户姓名，无 sender 时取上传用户姓名
            sender = record.get("sender", "")
            # 发送人显示名：若管理员为该 sender 设了绑定别名，用别名覆盖识别时抓取的微信名
            if sender and sender_alias_map.get(sender):
                record["sender_name"] = sender_alias_map[sender]
            if sender and sender in sender_name_map:
                record["display_fields"]["跟单人"] = sender_name_map[sender]
            elif not sender and record.get("user_id"):
                _uname = _upload_user_name_map.get(record["user_id"], "")
                if _uname:
                    record["display_fields"]["跟单人"] = _uname
            # 时间戳转字符串（避免 Flask jsonify 将 datetime 序列化为 GMT 格式导致前端时区偏移）
            for ts_field in ("created_at", "updated_at"):
                if record.get(ts_field) and not isinstance(record[ts_field], str):
                    record[ts_field] = str(record[ts_field])

        # 内部管理备注摘要注入（仅超管可见；其他角色不返回该字段）
        try:
            from flask import g as _g
            if _g.current_user.get("role") == ROLE_SUPER_ADMIN:
                _recs = result.get("records", [])
                _note_map = internal_notes_db.get_summaries(
                    _db, [r.get("id") for r in _recs if r.get("id")])
                for r in _recs:
                    r["internal_note"] = _note_map.get(r.get("id")) or {
                        "count": 0, "latest": "", "latest_attribution": "",
                        "latest_at": "", "reviewed_today": False,
                    }
        except Exception:
            logger.debug("注入内部备注摘要失败", exc_info=True)

        # 同车牌互补：保司信息 + 车主 + 证件号码 + 签单日期 + 业务员
        # pool 结构：{车牌: {"name", "addr", "owner", "id_no", "sign_date", "salesperson"}}
        _plate_pool = {}
        _plates_need_db = set()
        _plates_need_sign_db = set()
        for record in result.get("records", []):
            df = record.get("display_fields", {})
            plate = df.get("车牌", "") or df.get("车牌号", "")
            if not plate:
                continue
            if plate not in _plate_pool:
                _plate_pool[plate] = {"name": "", "addr": "", "owner": "", "id_no": "", "sign_date": "", "salesperson": ""}
            p = _plate_pool[plate]
            if not p["name"] and df.get("保司公司名称"):
                p["name"] = df["保司公司名称"]
            if not p["addr"] and df.get("保司地址"):
                p["addr"] = df["保司地址"]
            if not p["owner"] and df.get("车主"):
                p["owner"] = df["车主"]
            if not p["id_no"]:
                p["id_no"] = (df.get("投保人身份证号码") or df.get("证件号码") or
                              df.get("被保险人身份证号码") or "")
            if not p["sign_date"] and df.get("签单日期"):
                p["sign_date"] = df["签单日期"]
            if not p["salesperson"] and df.get("业务员"):
                p["salesperson"] = df["业务员"]
            if not p["name"] or not p["addr"] or not p["owner"] or not p["id_no"]:
                _plates_need_db.add(plate)
            if not p["sign_date"] or not p["salesperson"]:
                _plates_need_sign_db.add(plate)
        # 当前页不够的，从数据库查保司/车主/证件
        _still_missing = {p for p in _plates_need_db
                          if not all([_plate_pool.get(p, {}).get(k) for k in ("name", "addr", "owner", "id_no")])}
        if _still_missing:
            try:
                from insurance.db import find_insurer_info_by_plates
                _db_info = find_insurer_info_by_plates(_db, list(_still_missing))
                for p, info in _db_info.items():
                    if p not in _plate_pool:
                        _plate_pool[p] = {"name": "", "addr": "", "owner": "", "id_no": "", "sign_date": "", "salesperson": ""}
                    ep = _plate_pool[p]
                    if not ep["name"] and info.get("insurer_name"):
                        ep["name"] = info["insurer_name"]
                    if not ep["addr"] and info.get("insurer_address"):
                        ep["addr"] = info["insurer_address"]
                    if not ep["owner"] and info.get("owner"):
                        ep["owner"] = info["owner"]
                    if not ep["id_no"] and info.get("id_number"):
                        ep["id_no"] = info["id_number"]
            except Exception:
                logger.debug("查询同车牌信息失败")
        # 当前页不够的，从数据库查签单日期/业务员
        _still_sign_missing = {p for p in _plates_need_sign_db
                                if not _plate_pool.get(p, {}).get("sign_date") or not _plate_pool.get(p, {}).get("salesperson")}
        if _still_sign_missing:
            try:
                from insurance.db import find_sign_info_by_plates
                _sign_info = find_sign_info_by_plates(_db, list(_still_sign_missing))
                for p, info in _sign_info.items():
                    if p not in _plate_pool:
                        _plate_pool[p] = {"name": "", "addr": "", "owner": "", "id_no": "", "sign_date": "", "salesperson": ""}
                    ep = _plate_pool[p]
                    if not ep["sign_date"] and info.get("sign_date"):
                        ep["sign_date"] = info["sign_date"]
                    if not ep["salesperson"] and info.get("salesperson"):
                        ep["salesperson"] = info["salesperson"]
            except Exception:
                logger.debug("查询同车牌签单信息失败")
        # 回填缺失字段
        for record in result.get("records", []):
            df = record.get("display_fields", {})
            plate = df.get("车牌", "") or df.get("车牌号", "")
            if not plate or plate not in _plate_pool:
                continue
            pool = _plate_pool[plate]
            if not df.get("保司公司名称") and pool["name"]:
                df["保司公司名称"] = pool["name"]
            _cur_addr = df.get("保司地址", "")
            if (not _cur_addr or re.match(r'^邮政编码', _cur_addr)) and pool["addr"]:
                df["保司地址"] = pool["addr"]
            if not df.get("车主") and pool["owner"]:
                df["车主"] = pool["owner"]
            if pool["id_no"]:
                # 按记录里已有的 key 回填，都没有则写入"投保人身份证号码"
                filled_id = False
                for id_key in ("投保人身份证号码", "证件号码", "被保险人身份证号码"):
                    if id_key in df and not df[id_key]:
                        df[id_key] = pool["id_no"]
                        filled_id = True
                        break
                if not filled_id and not df.get("投保人身份证号码"):
                    df["投保人身份证号码"] = pool["id_no"]
            if not df.get("签单日期") and pool["sign_date"]:
                df["签单日期"] = pool["sign_date"]
            if not df.get("业务员") and pool["salesperson"]:
                df["业务员"] = pool["salesperson"]

        # 车架号互补：车牌为空但有车架号的记录，从同车架号其他记录补车牌和证件号
        _vin_pool = {}  # {VIN: {"plate": "", "id_no": "", "owner": ""}}
        for record in result.get("records", []):
            df = record.get("display_fields", {})
            vin = df.get("车架号", "")
            if not vin:
                continue
            if vin not in _vin_pool:
                _vin_pool[vin] = {"plate": "", "id_no": "", "owner": ""}
            vp = _vin_pool[vin]
            plate = df.get("车牌", "") or df.get("车牌号", "")
            if not vp["plate"] and plate:
                vp["plate"] = plate
            if not vp["owner"] and df.get("车主"):
                vp["owner"] = df["车主"]
            if not vp["id_no"]:
                vp["id_no"] = (df.get("投保人身份证号码") or df.get("证件号码") or
                               df.get("被保险人身份证号码") or "")
        for record in result.get("records", []):
            df = record.get("display_fields", {})
            vin = df.get("车架号", "")
            plate = df.get("车牌", "") or df.get("车牌号", "")
            if plate or not vin or vin not in _vin_pool:
                continue
            vp = _vin_pool[vin]
            if not df.get("车牌") and vp["plate"]:
                df["车牌"] = vp["plate"]
            if not df.get("车主") and vp["owner"]:
                df["车主"] = vp["owner"]
            if vp["id_no"]:
                filled_id = False
                for id_key in ("投保人身份证号码", "证件号码", "被保险人身份证号码"):
                    if id_key in df and not df[id_key]:
                        df[id_key] = vp["id_no"]
                        filled_id = True
                        break
                if not filled_id and not df.get("投保人身份证号码"):
                    df["投保人身份证号码"] = vp["id_no"]

        # 同 PDF 互补：同一 file_md5 中互补车牌、签单日期、业务员
        # 适用于：意外险/驾乘险与交强险/商业险来自同一 PDF，驾乘险本身无车牌/签单日期等字段
        _md5_pool = {}  # {md5: {"plate": "", "sign_date": "", "salesperson": ""}}
        for record in result.get("records", []):
            md5 = record.get("file_md5", "")
            if not md5:
                continue
            if md5 not in _md5_pool:
                _md5_pool[md5] = {"plate": "", "sign_date": "", "salesperson": ""}
            p = _md5_pool[md5]
            df = record.get("display_fields", {})
            if not p["plate"]:
                p["plate"] = df.get("车牌", "") or df.get("车牌号", "")
            if not p["sign_date"]:
                p["sign_date"] = df.get("签单日期", "")
            if not p["salesperson"]:
                p["salesperson"] = df.get("业务员", "")
        for record in result.get("records", []):
            md5 = record.get("file_md5", "")
            if not md5 or md5 not in _md5_pool:
                continue
            df = record.get("display_fields", {})
            p = _md5_pool[md5]
            if not df.get("车牌") and not df.get("车牌号") and p["plate"]:
                df["车牌"] = p["plate"]
            if not df.get("签单日期") and p["sign_date"]:
                df["签单日期"] = p["sign_date"]
            if not df.get("业务员") and p["salesperson"]:
                df["业务员"] = p["salesperson"]

        # 文件名兜底：车牌仍为空时，从文件名中提取
        _PLATE_RE = re.compile(
            r'[京津沪渝冀豫云辽黑湘皖鲁新苏浙赣鄂桂甘晋蒙陕吉闽贵粤川青藏琼宁夏]'
            r'[A-Z]-?[A-Z0-9]{4,6}'
        )
        for record in result.get("records", []):
            df = record.get("display_fields", {})
            if df.get("车牌") or df.get("车牌号"):
                continue
            filename = record.get("filename", "") or ""
            m = _PLATE_RE.search(filename)
            if m:
                df["车牌"] = m.group(0).replace("-", "")

        # 车牌→车架号互补：有车牌无车架号的记录，从同车牌其他记录补车架号
        # 必须在文件名兜底填完车牌之后执行，否则兜底车牌的记录无法获益
        _plate_vin_pool = {}
        for record in result.get("records", []):
            df = record.get("display_fields", {})
            plate = df.get("车牌", "") or df.get("车牌号", "")
            vin = df.get("车架号", "")
            if plate and vin and plate not in _plate_vin_pool:
                _plate_vin_pool[plate] = vin
        for record in result.get("records", []):
            df = record.get("display_fields", {})
            if df.get("车架号"):
                continue
            plate = df.get("车牌", "") or df.get("车牌号", "")
            if plate and plate in _plate_vin_pool:
                df["车架号"] = _plate_vin_pool[plate]

        # 交强到期时间：必须在 apply_user_config（日期格式化）之前，用「原始」终保日期
        # 构建 车牌→交强险终保日期 映射。否则等终保日期被格式化后再读取，若其格式不含年份
        # （如 MM月DD日），年份信息已丢失，交强到期时间将无法按自身配置的格式正确显示。
        list_visible_keys = {c["key"] for c in list_col_cfg.get("columns", []) if c.get("visible")}
        _need_compulsory_end = "交强到期时间" in list_visible_keys
        _plate_compulsory_end = {}
        if _need_compulsory_end:
            for record in result.get("records", []):
                df = record.get("display_fields", {})
                plate = df.get("车牌", "") or df.get("车牌号", "")
                policy_type = df.get("险种", "") or df.get("险种类型", "")
                if plate and ("交强" in policy_type or "交通事故责任强制" in policy_type):
                    end_date = df.get("终保日期", "")
                    # 仅采用「含 4 位年份」的终保日期；无年份的（已被格式化为 MM月DD日 等）
                    # 留给数据库兜底查询，避免格式化时丢失年份导致格式不统一
                    if end_date and re.search(r"\d{4}", str(end_date)):
                        _plate_compulsory_end[_norm_plate(plate)] = end_date

        # 交强与交商起保时间：同上，需在格式化前用原始起保日期按车牌分险种收集
        _need_start_combo = "交强与交商起保时间" in list_visible_keys
        _plate_start_dates = {}  # {车牌: {"commercial": 起保日期, "compulsory": 起保日期}}
        if _need_start_combo:
            for record in result.get("records", []):
                df = record.get("display_fields", {})
                plate = df.get("车牌", "") or df.get("车牌号", "")
                policy_type = df.get("险种", "") or df.get("险种类型", "")
                start_date = df.get("起保日期", "")
                if not (plate and start_date and re.search(r"\d{4}", str(start_date))):
                    continue
                is_compulsory = "交强" in policy_type or "交通事故责任强制" in policy_type
                bucket = _plate_start_dates.setdefault(_norm_plate(plate), {})
                bucket.setdefault("compulsory" if is_compulsory else "commercial", start_date)

        # 互补后：计算"保司（带地区）" + 应用用户配置
        for record in result.get("records", []):
            df = record.get("display_fields", {})
            if not df.get("保司（带地区）"):
                _addr = df.get("保司地址", "")
                _comp = df.get("承保公司", "")
                if _comp:
                    if _addr:
                        from insurance.policy_parser import extract_city_from_address
                        _city = extract_city_from_address(_addr)
                        df["保司（带地区）"] = (_city + _comp) if _city else _comp
                    else:
                        df["保司（带地区）"] = _comp
            # 将创建时间/识别时间注入 display_fields，使日期格式配置能生效
            if record.get("display_fields") is not None:
                if record.get("created_at") and "创建时间" not in record["display_fields"]:
                    record["display_fields"]["创建时间"] = str(record["created_at"])
                # 识别时间：优先用首次识别成功完成时间(first_recognized_at)，重新识别不会覆盖它；
                # 老记录/异常情况兜底回退到updated_at
                _recog_time = record.get("first_recognized_at") or record.get("updated_at")
                if _recog_time and "识别时间" not in record["display_fields"]:
                    record["display_fields"]["识别时间"] = str(_recog_time)
            if has_config and record.get("display_fields"):
                # 传入手动修改字段列表，避免公式覆盖用户手动编辑的值
                _mf = record.get("manual_fields")
                if _mf:
                    if isinstance(_mf, str):
                        try:
                            _mf = json.loads(_mf)
                        except (TypeError, json.JSONDecodeError):
                            _mf = []
                    user_config["_manual_fields"] = set(_mf) if isinstance(_mf, list) else set()
                else:
                    user_config.pop("_manual_fields", None)
                _before_val = record["display_fields"].get("交强到期时间", "<不存在>")
                record["display_fields"] = apply_user_config_to_fields(user_config, record["display_fields"])
                _after_val = record["display_fields"].get("交强到期时间", "<不存在>")
                if _before_val != _after_val:
                    logger.debug("[DEBUG-列表] record_id=%s apply_user_config前后交强到期时间变化: %s → %s",
                                record.get("id"), _before_val, _after_val)

        # 注入交强到期时间（使用格式化前构建的原始终保日期映射，并按其配置格式格式化）
        if _need_compulsory_end:
            # 如果当前页没有交强险记录，从数据库查同车牌的交强险终保日期（原始值）
            _plates_need_lookup = set()
            for record in result.get("records", []):
                df = record.get("display_fields", {})
                plate = df.get("车牌", "") or df.get("车牌号", "")
                np = _norm_plate(plate)
                if np and np not in _plate_compulsory_end:
                    _plates_need_lookup.add(np)
            if _plates_need_lookup:
                try:
                    from insurance.db import find_compulsory_end_dates
                    _db_end_dates = find_compulsory_end_dates(_db, list(_plates_need_lookup))
                    for _p, _v in _db_end_dates.items():
                        _plate_compulsory_end.setdefault(_norm_plate(_p), _v)
                except Exception:
                    logger.debug("查询交强到期时间失败")
            # 交强到期时间的显示格式（从用户配置读取）
            _compulsory_fmt = ""
            _compulsory_no_pad = False
            for _item in (user_config or {}).get("date_format", []):
                if _item.get("key") == "交强到期时间":
                    _compulsory_fmt = _item.get("value", "")
                elif _item.get("key") == "__no_pad__" and _item.get("value"):
                    _compulsory_no_pad = True
            for record in result.get("records", []):
                df = record.get("display_fields", {})
                # 手动填写过的交强到期时间已在 apply_user_config 中按配置格式化，保持不变
                _mf_raw = record.get("manual_fields")
                if isinstance(_mf_raw, str):
                    try:
                        _mf_raw = json.loads(_mf_raw)
                    except (TypeError, json.JSONDecodeError):
                        _mf_raw = []
                if isinstance(_mf_raw, list) and "交强到期时间" in _mf_raw:
                    continue
                # 未手动填写过，从同车牌其他保单注入交强到期时间（原始终保日期）并格式化
                plate = df.get("车牌", "") or df.get("车牌号", "")
                if not plate:
                    continue
                _inject_val = _plate_compulsory_end.get(_norm_plate(plate), "")
                if _inject_val and _compulsory_fmt:
                    _inject_val = _format_date(str(_inject_val), _compulsory_fmt, _compulsory_no_pad)
                df["交强到期时间"] = _inject_val

        # 注入交强与交商起保时间（同上，缺哪个险种的记录当前页没有就查数据库补）
        if _need_start_combo:
            _plates_need_lookup2 = set()
            for record in result.get("records", []):
                df = record.get("display_fields", {})
                np = _norm_plate(df.get("车牌", "") or df.get("车牌号", ""))
                if not np:
                    continue
                _bucket = _plate_start_dates.get(np, {})
                if "commercial" not in _bucket or "compulsory" not in _bucket:
                    _plates_need_lookup2.add(np)
            if _plates_need_lookup2:
                try:
                    from insurance.db import find_commercial_compulsory_start_dates
                    _db_start_dates = find_commercial_compulsory_start_dates(_db, list(_plates_need_lookup2))
                    for _p, _v in _db_start_dates.items():
                        _bucket = _plate_start_dates.setdefault(_norm_plate(_p), {})
                        for _k, _dv in _v.items():
                            _bucket.setdefault(_k, _dv)
                except Exception:
                    logger.debug("查询交强与交商起保时间失败")
            _start_combo_fmt = ""
            _start_combo_no_pad = False
            for _item in (user_config or {}).get("date_format", []):
                if _item.get("key") == "交强与交商起保时间":
                    _start_combo_fmt = _item.get("value", "")
                elif _item.get("key") == "__no_pad__" and _item.get("value"):
                    _start_combo_no_pad = True
            for record in result.get("records", []):
                df = record.get("display_fields", {})
                _mf_raw2 = record.get("manual_fields")
                if isinstance(_mf_raw2, str):
                    try:
                        _mf_raw2 = json.loads(_mf_raw2)
                    except (TypeError, json.JSONDecodeError):
                        _mf_raw2 = []
                if isinstance(_mf_raw2, list) and "交强与交商起保时间" in _mf_raw2:
                    continue
                plate = df.get("车牌", "") or df.get("车牌号", "")
                if not plate:
                    continue
                _bucket = _plate_start_dates.get(_norm_plate(plate), {})
                df["交强与交商起保时间"] = _combine_start_dates(
                    _bucket.get("commercial"), _bucket.get("compulsory"),
                    _start_combo_fmt, _start_combo_no_pad)

        # 注：is_abnormal（需人工补充）统一由 _compute_abnormal 在入库/字段更新/启动回填
        # 时按固定 10 列口径计算并写库。此处不再做"浏览时惰性重判+写库"，避免浏览行为
        # 持续改写历史记录的 is_abnormal、导致同一历史区间统计数字反复漂移。
        # 列表展示与统计均直接信任 is_abnormal 列，两者口径一致。

        # 复用已加载的页面列配置
        result["column_config"] = list_col_cfg

        # 用「与列表完全相同的筛选」内联返回 Tab 统计，确保列表与统计同源同时刻、
        # 永不打架（彻底杜绝"记数与实际识别数量对不上"）。statu/doc_category/is_abnormal
        # 不传，由统计自身按全部状态分类计算。
        try:
            _sf = {
                "roomid": roomid, "source_type": source_type, "source": source,
                # sender 仅在私聊筛选(source_type=user)时生效——与列表查询口径一致
                # （列表的 sender 也只在 source_type=user 时使用）；否则"全部来源"下
                # 残留的 sender 会让统计多筛一层、与列表对不上。
                "sender": (sender if source_type == "user" else ""),
                "keyword": keyword, "company_short": company_short,
                "ocr_engine": ocr_engine, "policy_type": policy_type,
                "date_start": date_start, "date_end": date_end,
                "updated_at_date_start": updated_at_date_start, "updated_at_date_end": updated_at_date_end,
                "search_company": search_company, "search_policy_no": search_policy_no,
                "search_plate_no": search_plate_no, "search_applicant": search_applicant,
                "search_insured": search_insured, "search_salesperson": search_salesperson,
                "sign_date_start": sign_date_start, "sign_date_end": sign_date_end,
                "start_date_start": start_date_start, "start_date_end": start_date_end,
                "end_date_start": end_date_start, "end_date_end": end_date_end,
            }
            _sf = {k: v for k, v in _sf.items() if v}
            if dedup:
                _sf["dedup"] = True
            if effective_user_ids is not None:
                _sf["user_ids"] = effective_user_ids
            if effective_enterprise_id is not None:
                _sf["enterprise_id"] = effective_enterprise_id
            result["stats"] = get_insurance_stats(_db, _sf or None)
        except Exception:
            logger.debug("列表内联统计计算失败", exc_info=True)

        # 企业可在「服务企业管理」里开启「默认按车牌合并显示」，前端据此自动开启合并视图
        _merge_default = False
        if effective_enterprise_id:
            try:
                from auth.db import get_enterprise_by_id
                _ent = get_enterprise_by_id(_db, effective_enterprise_id)
                _merge_default = bool(_ent and _ent.get("merge_by_plate"))
            except Exception:
                pass
        result["default_merge_by_plate"] = _merge_default

        return jsonify({"code": 0, "data": result})
    except Exception as e:
        err_msg = str(e)
        # MySQL 查询超时（MAX_EXECUTION_TIME）返回友好提示
        if "3024" in err_msg or "MAX_EXECUTION_TIME" in err_msg:
            logger.warning("查询保单记录超时: %s", err_msg)
            return jsonify({"code": 504, "msg": "查询超时，请缩小筛选范围后重试"}), 504
        logger.exception("查询保单记录失败")
        return jsonify({"code": 500, "msg": err_msg}), 500


@insurance_bp.route("/api/insurance/records/<int:record_id>", methods=["GET"])
def get_record(record_id):
    """获取单条保单识别记录详情"""
    try:
        record = get_insurance_record(_db, record_id)
        if not record:
            return jsonify({"code": 404, "msg": "记录不存在"}), 404

        # JSON 字段自动反序列化
        for field in ("parsed_fields", "mapped_fields", "manual_fields"):
            if isinstance(record.get(field), str):
                try:
                    record[field] = json.loads(record[field])
                except (TypeError, json.JSONDecodeError):
                    pass
        # 时间戳转字符串
        for ts_field in ("created_at", "updated_at", "first_recognized_at"):
            if record.get(ts_field) and not isinstance(record[ts_field], str):
                record[ts_field] = str(record[ts_field])

        # 历史记录兼容：车牌号→车牌
        if isinstance(record.get("display_fields"), dict):
            if not record["display_fields"].get("车牌") and record["display_fields"].get("车牌号"):
                record["display_fields"]["车牌"] = record["display_fields"]["车牌号"]

        # 实时关联跟单人：通过 sender 查找绑定用户姓名，无 sender 时取上传用户姓名
        tracker_name = ""
        sender = record.get("sender", "")
        if sender:
            try:
                from auth.db import get_binding_by_sender, get_user_by_id
                bound = get_binding_by_sender(_db, sender)
                if bound:
                    user_info = get_user_by_id(_db, bound["user_id"])
                    if user_info and user_info.get("name"):
                        tracker_name = user_info["name"]
            except Exception:
                pass
        elif record.get("user_id"):
            try:
                from auth.db import get_user_by_id
                upload_user = get_user_by_id(_db, record["user_id"])
                if upload_user and upload_user.get("name"):
                    tracker_name = upload_user["name"]
            except Exception:
                pass
        if tracker_name:
            if isinstance(record.get("parsed_fields"), dict):
                record["parsed_fields"]["跟单人"] = tracker_name
            if isinstance(record.get("display_fields"), str):
                try:
                    df = json.loads(record["display_fields"])
                    df["跟单人"] = tracker_name
                    record["display_fields"] = df
                except (TypeError, json.JSONDecodeError):
                    pass
            elif isinstance(record.get("display_fields"), dict):
                record["display_fields"]["跟单人"] = tracker_name

        # 应用用户配置（简称/日期格式等），与列表端点保持一致
        if isinstance(record.get("display_fields"), str):
            try:
                record["display_fields"] = json.loads(record["display_fields"])
            except (TypeError, json.JSONDecodeError):
                record["display_fields"] = {}
        if isinstance(record.get("display_fields"), dict):
            user = g.current_user
            user_config = get_effective_config(_db, user["user_id"], user["role"], user.get("parent_id"))
            has_config = user_config and any(user_config.get(k) for k in user_config)
            if has_config:
                _mf = record.get("manual_fields")
                if _mf:
                    if isinstance(_mf, str):
                        try:
                            _mf = json.loads(_mf)
                        except (TypeError, json.JSONDecodeError):
                            _mf = []
                    user_config["_manual_fields"] = set(_mf) if isinstance(_mf, list) else set()
                else:
                    user_config.pop("_manual_fields", None)
                record["display_fields"] = apply_user_config_to_fields(user_config, record["display_fields"])

        # 同一 PDF 的兄弟记录摘要（互补前先拿到，兄弟记录也用于互补）
        siblings = []
        if record.get("file_md5"):
            siblings = _get_sibling_records(_db, record["file_md5"], record_id)

        # 详情页互补：与列表端保持一致，覆盖所有运行时互补字段
        df = record.get("display_fields")
        if not isinstance(df, dict):
            df = {}
            record["display_fields"] = df

        # 1. 同 PDF 兄弟记录互补：车牌、签单日期、业务员
        _need_from_sib = (not df.get("车牌") and not df.get("车牌号"),
                          not df.get("签单日期"), not df.get("业务员"))
        if any(_need_from_sib):
            for _sib in siblings:
                _sib_df = _sib.get("display_fields") or {}
                if isinstance(_sib_df, str):
                    try:
                        _sib_df = json.loads(_sib_df)
                    except Exception:
                        _sib_df = {}
                if not df.get("车牌") and not df.get("车牌号"):
                    _sp = _sib_df.get("车牌") or _sib_df.get("车牌号")
                    if _sp:
                        df["车牌"] = _sp
                if not df.get("签单日期") and _sib_df.get("签单日期"):
                    df["签单日期"] = _sib_df["签单日期"]
                if not df.get("业务员") and _sib_df.get("业务员"):
                    df["业务员"] = _sib_df["业务员"]
                if df.get("车牌") and df.get("签单日期") and df.get("业务员"):
                    break

        # 2. 文件名兜底：车牌仍为空时从文件名提取
        if not df.get("车牌") and not df.get("车牌号"):
            _fn = record.get("filename", "") or ""
            _pm = re.search(
                r'[京津沪渝冀豫云辽黑湘皖鲁新苏浙赣鄂桂甘晋蒙陕吉闽贵粤川青藏琼宁夏]'
                r'[A-Z]-?[A-Z0-9]{4,6}', _fn
            )
            if _pm:
                df["车牌"] = _pm.group(0).replace("-", "")

        # 3. 同车牌互补：车主、车架号、保司公司名称、保司地址、证件号码、签单日期、业务员
        _plate = df.get("车牌") or df.get("车牌号")
        if _plate:
            _need = (not df.get("车主") or not df.get("车架号") or
                     not df.get("保司公司名称") or not df.get("保司地址") or
                     not df.get("投保人身份证号码"))
            if _need:
                try:
                    from insurance.db import find_insurer_info_by_plates
                    _info_map = find_insurer_info_by_plates(_db, [_plate])
                    _info = _info_map.get(_plate, {})
                    if not df.get("车主") and _info.get("owner"):
                        df["车主"] = _info["owner"]
                    if not df.get("保司公司名称") and _info.get("insurer_name"):
                        df["保司公司名称"] = _info["insurer_name"]
                    if not df.get("保司地址") and _info.get("insurer_address"):
                        df["保司地址"] = _info["insurer_address"]
                    if _info.get("id_number"):
                        for _id_key in ("投保人身份证号码", "证件号码", "被保险人身份证号码"):
                            if _id_key in df and not df[_id_key]:
                                df[_id_key] = _info["id_number"]
                                break
                        if not df.get("投保人身份证号码"):
                            df["投保人身份证号码"] = _info["id_number"]
                except Exception:
                    pass
            # 补车架号（关联表里没有，从同车牌记录查）
            if not df.get("车架号"):
                try:
                    from insurance.db import find_records_by_plate
                    for _pr in find_records_by_plate(_db, _plate, exclude_id=record_id):
                        _pf = _pr.get("parsed_fields") or {}
                        if isinstance(_pf, str):
                            try:
                                _pf = json.loads(_pf)
                            except Exception:
                                _pf = {}
                        _vin = _pf.get("车架号VIN") or _pf.get("车架号")
                        if _vin:
                            df["车架号"] = _vin
                            break
                except Exception:
                    pass
            # 补签单日期/业务员（从同车牌其他记录查）
            if not df.get("签单日期") or not df.get("业务员"):
                try:
                    from insurance.db import find_sign_info_by_plates
                    _sign_map = find_sign_info_by_plates(_db, [_plate])
                    _sign = _sign_map.get(_plate, {})
                    if not df.get("签单日期") and _sign.get("sign_date"):
                        df["签单日期"] = _sign["sign_date"]
                    if not df.get("业务员") and _sign.get("salesperson"):
                        df["业务员"] = _sign["salesperson"]
                except Exception:
                    pass

        return jsonify({"code": 0, "data": record, "siblings": siblings})
    except Exception as e:
        logger.exception("获取保单记录 %d 失败", record_id)
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/records/<int:record_id>/fields", methods=["PUT"])
def update_record_fields(record_id):
    """手动修改保单识别字段值，并记录哪些字段被手动修改过"""
    _require_login()
    try:
        record = get_insurance_record(_db, record_id)
        if not record:
            return jsonify({"code": 404, "msg": "记录不存在"}), 404

        data = request.get_json(force=True)
        updated_fields = data.get("fields", {})  # {字段名: 新值}
        if not updated_fields:
            return jsonify({"code": 400, "msg": "未提供修改字段"}), 400

        # 读取现有 parsed_fields
        pf = record.get("parsed_fields") or "{}"
        if isinstance(pf, str):
            try:
                pf = json.loads(pf)
            except (TypeError, json.JSONDecodeError):
                pf = {}

        # 读取现有 manual_fields（手动修改过的字段名列表）
        mf = record.get("manual_fields") or "[]"
        if isinstance(mf, str):
            try:
                mf = json.loads(mf)
            except (TypeError, json.JSONDecodeError):
                mf = []
        if not isinstance(mf, list):
            mf = []

        # 合并修改
        for field_name, new_value in updated_fields.items():
            pf[field_name] = new_value
            if field_name not in mf:
                mf.append(field_name)

        # 读取现有 display_fields 作为基础（包含跟单人、车主等补充字段）
        df = record.get("display_fields") or "{}"
        if isinstance(df, str):
            try:
                df = json.loads(df)
            except (TypeError, json.JSONDecodeError):
                df = {}
        # 只更新本次修改的字段到 display_fields，保留其他已有的补充字段
        # 注意：field_name 可能是 OCR 原始字段名（如"车牌号"），列表/详情实际渲染读取的
        # 是映射后的展示列名（如"车牌"）——必须经 get_mapping_for_company 转换，否则会
        # 写出一个没人读取的影子字段，展示列上的旧值永远不刷新（表现为列表和详情不一致）
        from insurance.field_mapping import get_mapping_for_company
        company_short = pf.get("保险公司简称", "")
        field_mapping_table = get_mapping_for_company(company_short)
        display_fields = dict(df)
        for field_name, new_value in updated_fields.items():
            display_key = field_mapping_table.get(field_name) or field_name
            display_fields[display_key] = new_value

        # 更新数据库（同步更新 display_fields）
        logger.info("[DEBUG-保存] record_id=%s, 更新字段=%s, manual_fields=%s", record_id, updated_fields, mf)
        logger.info("[DEBUG-保存] 写入DB前 display_fields中交强到期时间=%s", display_fields.get("交强到期时间", "<不存在>"))
        update_insurance_record(_db, record_id, {
            "parsed_fields": pf,
            "manual_fields": mf,
            "display_fields": display_fields,
        })

        # 验证写入是否成功：重新从DB读取
        _verify_record = get_insurance_record(_db, record_id)
        _verify_df = _verify_record.get("display_fields", {}) if _verify_record else {}
        if isinstance(_verify_df, str):
            try:
                _verify_df = json.loads(_verify_df)
            except Exception:
                _verify_df = {}
        _verify_mf = _verify_record.get("manual_fields", []) if _verify_record else []
        if isinstance(_verify_mf, str):
            try:
                _verify_mf = json.loads(_verify_mf)
            except Exception:
                _verify_mf = []
        logger.info("[DEBUG-保存] 写入DB后重读 display_fields中交强到期时间=%s, manual_fields=%s",
                     _verify_df.get("交强到期时间", "<不存在>"), _verify_mf)

        # 重新应用用户配置（公式计算等），手动修改的字段不被公式覆盖
        user = g.current_user
        user_config = get_effective_config(_db, user["user_id"], user["role"], user.get("parent_id"))
        if user_config and any(user_config.get(k) for k in user_config):
            user_config["_manual_fields"] = set(mf)
            display_fields = apply_user_config_to_fields(user_config, display_fields)

        logger.info("[DEBUG-保存] apply_user_config后 display_fields中交强到期时间=%s", display_fields.get("交强到期时间", "<不存在>"))

        return jsonify({"code": 0, "msg": "保存成功", "data": {
            "parsed_fields": pf,
            "manual_fields": mf,
            "display_fields": display_fields,
        }})
    except Exception as e:
        logger.exception("更新保单字段 %d 失败", record_id)
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/records/<int:record_id>/mark-normal", methods=["PUT"])
def mark_record_normal(record_id):
    """手动标记需人工补充记录为正常，需提供原因"""
    _require_login()
    try:
        record = get_insurance_record(_db, record_id)
        if not record:
            return jsonify({"code": 404, "msg": "记录不存在"}), 404

        data = request.get_json(force=True)
        reason = (data.get("reason") or "").strip()
        if not reason:
            return jsonify({"code": 400, "msg": "请提供标记原因"}), 400

        # 更新 abnormal_override_reason，触发 is_abnormal 重算
        update_insurance_record(_db, record_id, {
            "abnormal_override_reason": reason,
        })

        return jsonify({"code": 0, "msg": "已标记为正常"})
    except Exception as e:
        logger.exception("标记正常失败 record_id=%d", record_id)
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/records/<int:record_id>/undo-mark-normal", methods=["PUT"])
def undo_mark_record_normal(record_id):
    """撤销手动标记正常"""
    _require_login()
    try:
        record = get_insurance_record(_db, record_id)
        if not record:
            return jsonify({"code": 404, "msg": "记录不存在"}), 404

        # 清除 override，重算异常状态
        update_insurance_record(_db, record_id, {
            "abnormal_override_reason": None,
        })

        return jsonify({"code": 0, "msg": "已撤销标记"})
    except Exception as e:
        logger.exception("撤销标记失败 record_id=%d", record_id)
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/records/<int:record_id>/internal-notes", methods=["GET"])
def list_internal_notes(record_id):
    """获取某保单的内部管理备注（每日追加日志，最新在前）。仅超级管理员。"""
    _require_login()
    from flask import g
    if g.current_user.get("role") != ROLE_SUPER_ADMIN:
        return jsonify({"code": 403, "msg": "无权限"}), 403
    try:
        notes = internal_notes_db.list_notes(_db, record_id)
        return jsonify({"code": 0, "data": {"notes": notes}})
    except Exception as e:
        logger.exception("获取内部备注失败 record_id=%d", record_id)
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/records/<int:record_id>/internal-notes", methods=["POST"])
def add_internal_note(record_id):
    """追加一条内部管理备注。仅超级管理员。body: {content, attribution}"""
    _require_login()
    from flask import g
    if g.current_user.get("role") != ROLE_SUPER_ADMIN:
        return jsonify({"code": 403, "msg": "无权限"}), 403
    try:
        record = get_insurance_record(_db, record_id)
        if not record:
            return jsonify({"code": 404, "msg": "记录不存在"}), 404
        data = request.get_json(force=True) or {}
        content = (data.get("content") or "").strip()
        attribution = (data.get("attribution") or "").strip()
        if not content:
            return jsonify({"code": 400, "msg": "请填写备注内容"}), 400
        uid = g.current_user.get("user_id")
        # 操作人姓名：服务端按 user_id 查 users 表（同库），避免前端伪造
        author_name = ""
        try:
            from auth.db import get_user_by_id
            _u = get_user_by_id(_db, uid)
            author_name = (_u or {}).get("name") or ""
        except Exception:
            pass
        note_id = internal_notes_db.add_note(
            _db, record_id, content, attribution, author_id=uid, author_name=author_name)
        notes = internal_notes_db.list_notes(_db, record_id)
        summary = internal_notes_db.get_summaries(_db, [record_id]).get(record_id)
        return jsonify({"code": 0, "msg": "已添加", "data": {
            "note_id": note_id, "notes": notes, "summary": summary}})
    except Exception as e:
        logger.exception("添加内部备注失败 record_id=%d", record_id)
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/records/<int:record_id>", methods=["DELETE"])
def delete_record(record_id):
    """软删除保单识别记录（仅超级管理员）。记录不会物理删除，仅打 deleted_at 标记。"""
    from flask import g
    if g.current_user.get("role") != ROLE_SUPER_ADMIN:
        return jsonify({"code": 403, "msg": "无权限，仅超级管理员可删除记录"}), 403
    try:
        record = get_insurance_record(_db, record_id)
        if not record:
            return jsonify({"code": 404, "msg": "记录不存在"}), 404
        ok = soft_delete_record(_db, record_id)
        if not ok:
            return jsonify({"code": 400, "msg": "记录已被删除"}), 400
        logger.info("软删除保单记录 id=%d by user_id=%s, filename=%s",
                    record_id, g.current_user.get("user_id"), record.get("filename"))
        return jsonify({"code": 0, "msg": "已删除"})
    except Exception as e:
        logger.exception("软删除保单记录失败 record_id=%d", record_id)
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/records/<int:record_id>/preview", methods=["GET"])
def preview_record_file(record_id):
    """代理预览保单文件（避免 COS 直接下载）"""
    import requests as http_requests

    try:
        record = get_insurance_record(_db, record_id)
        if not record:
            return jsonify({"code": 404, "msg": "记录不存在"}), 404

        cos_url = record.get("cos_url", "")
        if not cos_url:
            return jsonify({"code": 404, "msg": "文件 URL 不存在"}), 404

        # 请求 COS 文件
        resp = http_requests.get(cos_url, timeout=30, stream=True)
        if resp.status_code != 200:
            return jsonify({"code": 502, "msg": "文件下载失败"}), 502

        filename = record.get("filename", "file.pdf")
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "pdf"
        mime_map = {
            "pdf": "application/pdf",
            "jpg": "image/jpeg", "jpeg": "image/jpeg",
            "png": "image/png", "bmp": "image/bmp",
        }
        mimetype = mime_map.get(ext, "application/octet-stream")

        return send_file(
            io.BytesIO(resp.content),
            mimetype=mimetype,
            as_attachment=False,
            download_name=filename,
        )
    except Exception as e:
        logger.exception("预览保单文件 %d 失败", record_id)
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/records/<int:record_id>/raw-text", methods=["GET"])
def get_record_raw_text(record_id):
    """获取保单记录的 OCR 原文（优先从数据库缓存读取，否则从 COS 提取并缓存）"""
    import requests as http_requests

    try:
        record = get_insurance_record(_db, record_id)
        if not record:
            return jsonify({"code": 404, "msg": "记录不存在"}), 404

        # 优先从数据库缓存读取
        cached_text = record.get("raw_text", "")
        ocr_text = record.get("ocr_text", "") or ""
        if cached_text:
            return jsonify({"code": 0, "data": {"raw_text": cached_text, "ocr_text": ocr_text}})

        # 数据库无缓存，从 COS 下载提取
        cos_url = record.get("cos_url", "")
        if not cos_url:
            return jsonify({"code": 404, "msg": "文件 URL 不存在"}), 404

        resp = http_requests.get(cos_url, timeout=30)
        if resp.status_code != 200:
            return jsonify({"code": 502, "msg": "文件下载失败"}), 502

        text = ""
        filename = record.get("filename", "")
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "pdf"

        if ext == "pdf":
            try:
                from insurance.ocr_service import extract_text_from_pdf_bytes
                result = extract_text_from_pdf_bytes(resp.content)
                if result.get("success"):
                    text = result.get("text", "")
            except Exception as e:
                logger.warning("pdfplumber 提取原文失败: %s", e)

        # 写回数据库缓存
        if text:
            try:
                update_insurance_record(_db, record_id, {"raw_text": text})
            except Exception:
                pass

        return jsonify({"code": 0, "data": {"raw_text": text, "ocr_text": ocr_text}})
    except Exception as e:
        logger.exception("获取保单原文 %d 失败", record_id)
        return jsonify({"code": 500, "msg": str(e)}), 500


_stats_cache = {}
_STATS_TTL = 30  # stats 缓存 30 秒

@insurance_bp.route("/api/insurance/stats", methods=["GET"])
def get_stats():
    """获取保单识别统计信息，支持与列表相同的筛选参数"""
    import time as _t
    try:
        # 简单缓存：以完整 query string 为 key
        cache_key = request.query_string.decode()
        now = _t.time()
        if cache_key in _stats_cache:
            exp, data = _stats_cache[cache_key]
            if now < exp:
                return jsonify({"code": 0, "data": data})
        filters = {}
        for key in ("roomid", "source_type", "source", "sender", "keyword",
                     "company_short", "policy_type", "ocr_engine", "date_start", "date_end",
                     "updated_at_date_start", "updated_at_date_end",
                     "search_company", "search_policy_no", "search_plate_no",
                     "search_applicant", "search_insured", "search_salesperson",
                     "sign_date_start", "sign_date_end",
                     "start_date_start", "start_date_end",
                     "end_date_start", "end_date_end"):
            val = request.args.get(key, "").strip()
            if val:
                filters[key] = val
        if request.args.get("dedup") == "1":
            filters["dedup"] = True
        # 按账号权限过滤（超管可通过参数指定企业）
        req_eid = request.args.get("enterprise_id", "").strip()
        from auth.db import ROLE_SUPER_ADMIN
        if g.current_user.get("role") == ROLE_SUPER_ADMIN and req_eid:
            filters["enterprise_id"] = int(req_eid)
        else:
            user_ids = _get_user_ids_filter()
            if user_ids is not None:
                filters["user_ids"] = user_ids
            eid = _get_enterprise_id_filter()
            if eid is not None:
                filters["enterprise_id"] = eid
        stats = get_insurance_stats(_db, filters=filters if filters else None)
        _stats_cache[cache_key] = (_t.time() + _STATS_TTL, stats)
        # 防止缓存无限增长
        if len(_stats_cache) > 200:
            oldest = sorted(_stats_cache, key=lambda k: _stats_cache[k][0])[:50]
            for k in oldest: _stats_cache.pop(k, None)
        return jsonify({"code": 0, "data": stats})
    except Exception as e:
        logger.exception("获取保单统计失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


_dashboard_cache = {}   # key → (expire_ts, data)
_DASHBOARD_TTL = 120    # 缓存 2 分钟

@insurance_bp.route("/api/insurance/dashboard", methods=["GET"])
@login_required
def get_dashboard():
    """数据看板（仅超管），返回综合统计数据"""
    import time as _time_mod
    from auth.db import ROLE_SUPER_ADMIN
    if g.current_user.get("role") != ROLE_SUPER_ADMIN:
        return jsonify({"code": 403, "msg": "无权限"}), 403
    try:
        period        = request.args.get("period", "30").strip()
        enterprise_id = request.args.get("enterprise_id", "").strip()
        company_short = request.args.get("company_short", "").strip()

        # 命中缓存直接返回
        cache_key = (period, enterprise_id, company_short)
        now_ts = _time_mod.time()
        if cache_key in _dashboard_cache:
            expire_ts, cached = _dashboard_cache[cache_key]
            if now_ts < expire_ts:
                return jsonify({"code": 0, "data": cached, "cached": True})

        # 根据 period 构建时间过滤片段
        if period == "today":
            time_filter      = "AND DATE({col}) = CURDATE()"
            time_filter_p    = []
            trend_group_hint = 1   # 只有1天，趋势退化为单点
        elif period == "yesterday":
            time_filter      = "AND DATE({col}) = DATE_SUB(CURDATE(), INTERVAL 1 DAY)"
            time_filter_p    = []
            trend_group_hint = 1
        elif period == "all":
            time_filter      = ""
            time_filter_p    = []
            trend_group_hint = 365
        else:
            days = int(period) if period.isdigit() else 30
            time_filter      = "AND {col} >= DATE_SUB(CURDATE(), INTERVAL %s DAY)"
            time_filter_p    = [days]
            trend_group_hint = days

        def _time(col="created_at"):
            """返回 (sql片段, params)，col 为字段名（含表前缀）"""
            return time_filter.format(col=col), list(time_filter_p)

        conn = _db.pool.connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)

        def _where(prefix="", with_enterprise=True, with_company=True):
            p = prefix
            parts = [f"{p}deleted_at IS NULL", f"{p}status != 'duplicate'"]
            params = []
            if with_enterprise and enterprise_id:
                parts.append(f"{p}enterprise_id = %s")
                params.append(int(enterprise_id))
            if with_company and company_short:
                parts.append(f"{p}company_short = %s")
                params.append(company_short)
            return " AND ".join(parts), params

        # 1. 总览指标
        wh, wp = _where()
        tf, tp = _time()
        cursor.execute(f"""
            SELECT COUNT(*) AS total,
              SUM(status='done') AS done, SUM(status='failed') AS failed,
              SUM(status='abnormal') AS abnormal, SUM(status='pending') AS pending,
              SUM(DATE(created_at) = CURDATE()) AS today_count,
              SUM(DATE(created_at) >= DATE_SUB(CURDATE(), INTERVAL 7 DAY)) AS week7
            FROM insurance_records WHERE {wh} {tf}
        """, wp + tp or None)
        overview = cursor.fetchone() or {}
        # period_* 直接用 total/done/failed（已按时间过滤）
        overview["period_total"]    = overview.get("total", 0)
        overview["period_done"]     = overview.get("done", 0)
        overview["period_failed"]   = overview.get("failed", 0)
        overview["period_abnormal"] = overview.get("abnormal", 0)
        overview["today"]           = overview.pop("today_count", 0)

        # 2. 趋势（今日/昨日按小时，其余按天）
        wh, wp = _where()
        tf, tp = _time()
        trend_type = "hourly" if period in ("today", "yesterday") else "daily"
        if trend_type == "hourly":
            cursor.execute(f"""
                SELECT HOUR(created_at) AS hour, COUNT(*) AS total,
                       SUM(status='done') AS done, SUM(status='failed') AS failed
                FROM insurance_records
                WHERE {wh} {tf}
                GROUP BY hour ORDER BY hour ASC
            """, (wp + tp) if (wp or tp) else None)
            raw_trend = cursor.fetchall() or []
            # 补全 0-23 小时（没数据的小时补 0）
            hour_map = {r["hour"]: r for r in raw_trend}
            trend = []
            for h in range(24):
                r = hour_map.get(h, {"hour": h, "total": 0, "done": 0, "failed": 0})
                r["label"] = f"{h:02d}:00"
                trend.append(r)
        else:
            cursor.execute(f"""
                SELECT DATE(created_at) AS day, COUNT(*) AS total,
                       SUM(status='done') AS done, SUM(status='failed') AS failed
                FROM insurance_records
                WHERE {wh} {tf}
                GROUP BY day ORDER BY day ASC
            """, (wp + tp) if (wp or tp) else None)
            trend = cursor.fetchall() or []
            for r in trend:
                if hasattr(r.get("day"), "strftime"):
                    r["day"] = r["day"].strftime("%Y-%m-%d")
                r["label"] = r.get("day", "")

        # 3. 保司排行 Top10
        wh, wp = _where(with_company=False)
        tf, tp = _time()
        cursor.execute(f"""
            SELECT company_short AS company, COUNT(*) AS total,
                   SUM(status='done') AS done, SUM(status='failed') AS failed
            FROM insurance_records
            WHERE {wh} {tf} AND company_short IS NOT NULL AND company_short != ''
            GROUP BY company_short ORDER BY total DESC LIMIT 10
        """, (wp + tp) if (wp or tp) else None)
        company_rank = cursor.fetchall() or []

        # 4. 险种分布
        wh, wp = _where(prefix="r.")
        tf, tp = _time("r.created_at")
        cursor.execute(f"""
            SELECT pf.policy_type AS policy_type, COUNT(*) AS total
            FROM insurance_records r
            JOIN insurance_policy_fields pf ON pf.record_id = r.id
            WHERE {wh} {tf} AND pf.policy_type IS NOT NULL AND pf.policy_type != ''
            GROUP BY pf.policy_type ORDER BY total DESC LIMIT 10
        """, (wp + tp) if (wp or tp) else None)
        policy_type_dist = cursor.fetchall() or []

        # 5. OCR 引擎分布
        wh, wp = _where()
        tf, tp = _time()
        cursor.execute(f"""
            SELECT ocr_engine, COUNT(*) AS total
            FROM insurance_records
            WHERE {wh} {tf} AND ocr_engine IS NOT NULL
            GROUP BY ocr_engine ORDER BY total DESC
        """, (wp + tp) if (wp or tp) else None)
        ocr_dist = cursor.fetchall() or []

        # 6. 来源分布
        wh, wp = _where()
        tf, tp = _time()
        cursor.execute(f"""
            SELECT
              SUM(source='group' OR (roomid IS NOT NULL AND roomid != '')) AS from_group,
              SUM(source='user') AS from_user,
              SUM(source='manual' OR source='') AS from_manual
            FROM insurance_records WHERE {wh} {tf}
        """, (wp + tp) if (wp or tp) else None)
        source_dist = cursor.fetchone() or {}

        # 7. 各企业使用量 Top10
        wh, wp = _where(prefix="r.", with_enterprise=False)
        tf, tp = _time("r.created_at")
        cursor.execute(f"""
            SELECT e.name AS enterprise, COUNT(*) AS total, SUM(r.status='done') AS done
            FROM insurance_records r
            JOIN enterprises e ON e.id = r.enterprise_id
            WHERE {wh} {tf}
            GROUP BY r.enterprise_id ORDER BY total DESC LIMIT 10
        """, (wp + tp) if (wp or tp) else None)
        enterprise_rank = cursor.fetchall() or []

        conn.close()
        result = {
            "overview": overview,
            "trend": trend,
            "trend_type": trend_type,
            "company_rank": company_rank,
            "policy_type_dist": policy_type_dist,
            "ocr_dist": ocr_dist,
            "source_dist": source_dist,
            "enterprise_rank": enterprise_rank,
            "period": period,
        }
        # 写入缓存（today/yesterday 缓存 30 秒，其他 2 分钟）
        ttl = 30 if period in ("today", "yesterday") else _DASHBOARD_TTL
        _dashboard_cache[cache_key] = (_time_mod.time() + ttl, result)
        return jsonify({"code": 0, "data": result})
    except Exception as e:
        logger.exception("获取看板数据失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


_record_companies_cache = {}
_RECORD_COMPANIES_TTL = 120  # 2分钟

@insurance_bp.route("/api/insurance/record-companies", methods=["GET"])
def list_record_companies():
    """
    获取列表「保司」筛选下拉的全库保司简称列表（权限范围内、去重排序）。
    与前端「从当前页记录现算」不同，这里返回全库数据，不受分页影响。
    返回: {code:0, data: ["人民财产", "国寿财产", ...]}
    """
    import time as _t
    try:
        user_ids = _get_user_ids_filter()
        eid = _get_enterprise_id_filter()
        ck = (str(user_ids), str(eid))
        now = _t.time()
        if ck in _record_companies_cache:
            exp, data = _record_companies_cache[ck]
            if now < exp:
                return jsonify({"code": 0, "data": data})
        companies = get_record_company_list(_db, user_ids=user_ids, enterprise_id=eid)
        _record_companies_cache[ck] = (now + _RECORD_COMPANIES_TTL, companies)
        return jsonify({"code": 0, "data": companies})
    except Exception as e:
        logger.exception("获取保司列表失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


# 支持保司统计缓存（后端5分钟缓存，避免频繁查库）
_company_cache = {"data": None, "time": 0}
_COMPANY_CACHE_TTL = 300  # 5分钟


_DEFAULT_MANUAL_SECTIONS = [
    {
        "id": "welcome",
        "title": "欢迎使用",
        "content": "# 欢迎使用午虎AI保单台账\n\n这是使用手册，内部超级管理员可以点击右上角「编辑」按钮修改这里的内容。\n\n左侧是章节目录，点击可以切换查看。",
    },
]


@insurance_bp.route("/api/insurance/manual", methods=["GET"])
def get_manual():
    """获取使用手册章节内容（登录即可查看，编辑需内部超管权限）"""
    sections = get_insurance_config(_db, "help_manual_sections", None)
    if not sections:
        sections = _DEFAULT_MANUAL_SECTIONS
    return jsonify({"code": 0, "data": {"sections": sections}})


@insurance_bp.route("/api/insurance/manual", methods=["POST"])
def save_manual():
    """保存使用手册章节内容（仅内部超级管理员）"""
    user = g.current_user
    if not user or user.get("role") != "super_admin":
        return jsonify({"code": 403, "msg": "仅内部超级管理员可编辑使用手册"}), 403
    data = request.get_json(force=True) or {}
    sections = data.get("sections")
    if not isinstance(sections, list):
        return jsonify({"code": 400, "msg": "参数格式错误"}), 400
    for s in sections:
        if not isinstance(s, dict) or "title" not in s or "content" not in s:
            return jsonify({"code": 400, "msg": "章节格式错误，需包含 title 和 content"}), 400
    set_insurance_config(_db, "help_manual_sections", sections)
    return jsonify({"code": 0, "msg": "保存成功"})


@insurance_bp.route("/api/insurance/manual/upload-image", methods=["POST"])
def upload_manual_image():
    """使用手册编辑：上传插图，返回可直接用于 Markdown 的图片URL（仅内部超级管理员）"""
    user = g.current_user
    if not user or user.get("role") != "super_admin":
        return jsonify({"code": 403, "msg": "仅内部超级管理员可上传图片"}), 403
    if not _handler or not _handler.cos_storage:
        return jsonify({"code": 503, "msg": "云存储未初始化"}), 503
    f = request.files.get("file")
    if not f:
        return jsonify({"code": 400, "msg": "缺少文件"}), 400
    ext = (f.filename.rsplit(".", 1)[-1] if "." in f.filename else "png").lower()
    if ext not in ("png", "jpg", "jpeg", "gif", "webp"):
        return jsonify({"code": 400, "msg": "仅支持 png/jpg/jpeg/gif/webp 格式图片"}), 400
    try:
        data = f.read()
        cos_key = f"manual/upload_{uuid.uuid4().hex}.{ext}"
        url = _handler.cos_storage.upload_bytes(data, cos_key)
        if not url:
            return jsonify({"code": 500, "msg": "上传失败"}), 500
        return jsonify({"code": 0, "data": {"url": url}})
    except Exception as e:
        logger.exception("使用手册图片上传失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/supported-companies", methods=["GET"])
def get_supported_companies():
    """
    获取支持的保司列表（从已识别记录中统计，5分钟缓存）。
    返回: {companies: [{company, count, supported}], threshold, aliases: {原名: 简称}}
    """
    import time as _time

    # 检查缓存
    if _company_cache["data"] and (_time.time() - _company_cache["time"] < _COMPANY_CACHE_TTL):
        return jsonify({"code": 0, "data": _company_cache["data"]})

    try:
        # 从配置读取阈值和别名映射
        threshold = get_insurance_config(_db, "company_threshold", 10)
        if isinstance(threshold, str):
            try:
                threshold = int(threshold)
            except ValueError:
                threshold = 10
        aliases = get_insurance_config(_db, "company_aliases", {})
        if not isinstance(aliases, dict):
            aliases = {}

        # 从数据库统计各保司明细数据（总量/本月量/险种分布/最近识别时间）
        stats = get_company_detail_stats(_db)

        # 应用别名合并统计：多个原名 → 同一简称，各项数值累加、最近时间取较大者
        _num_keys = ("count", "month_count", "compulsory", "commercial", "accident", "non_vehicle")
        merged = {}
        for item in stats:
            company = item["company"]
            display_name = aliases.get(company, company)
            if display_name in merged:
                m = merged[display_name]
                for k in _num_keys:
                    m[k] += item.get(k, 0)
                if item.get("last_time", "") > m.get("last_time", ""):
                    m["last_time"] = item["last_time"]
                _fa, _fb = m.get("first_time", ""), item.get("first_time", "")
                if _fb and (not _fa or _fb < _fa):
                    m["first_time"] = _fb
            else:
                merged[display_name] = {**item, "company": display_name}

        companies = sorted(merged.values(), key=lambda x: x["count"], reverse=True)
        for c in companies:
            c["supported"] = c["count"] >= threshold

        result_data = {
            "companies": companies,
            "threshold": threshold,
            "aliases": aliases,
        }
        _company_cache["data"] = result_data
        _company_cache["time"] = _time.time()

        return jsonify({"code": 0, "data": result_data})
    except Exception as e:
        logger.exception("获取支持保司列表失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/supported-companies/config", methods=["PUT"])
def update_supported_companies_config():
    """
    更新保司支持配置。
    请求体: {threshold?: int, aliases?: {原名: 简称, ...}}
    """
    try:
        body = request.get_json(force=True) or {}

        if "threshold" in body:
            set_insurance_config(_db, "company_threshold", int(body["threshold"]))
        if "aliases" in body:
            if not isinstance(body["aliases"], dict):
                return jsonify({"code": 400, "msg": "aliases 必须是对象格式"}), 400
            set_insurance_config(_db, "company_aliases", body["aliases"])

        # 清除缓存
        _company_cache["data"] = None
        _company_cache["time"] = 0

        return jsonify({"code": 0, "msg": "配置已更新"})
    except Exception as e:
        logger.exception("更新保司配置失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


def _cross_fill_from_siblings(db, record_id: int, file_md5: str, parsed_fields: dict):
    """从同一 PDF（相同 file_md5）的兄弟记录中互补缺失的基础字段"""
    if not file_md5:  # 空 md5 不能作分组键，否则会把所有空md5记录当成兄弟互相污染
        return
    fill_fields = ["车牌号", "投保人", "被保险人", "车主", "证件号码",
                   "车架号VIN", "发动机号", "厂牌型号",
                   "保司公司名称", "保司地址"]
    # 检查当前记录是否有缺失字段
    missing = [f for f in fill_fields if not parsed_fields.get(f)]
    if not missing:
        return

    conn = db.pool.connection()
    try:
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        cursor.execute(
            "SELECT parsed_fields FROM insurance_records "
            "WHERE file_md5 = %s AND id != %s AND status = 'done'",
            (file_md5, record_id)
        )
        rows = cursor.fetchall()
    finally:
        conn.close()

    filled = []
    for row in rows:
        pf = row.get("parsed_fields") or {}
        if isinstance(pf, str):
            try:
                pf = json.loads(pf) or {}
            except (TypeError, json.JSONDecodeError):
                continue
        if not isinstance(pf, dict):
            continue
        for f in list(missing):
            val = pf.get(f, "")
            if val:
                parsed_fields[f] = val
                filled.append(f"{f}={val}")
                missing.remove(f)
        if not missing:
            break

    if filled:
        logger.info("兄弟记录互补: record_id=%d, 补充=%s", record_id, ", ".join(filled))


def _get_sibling_records(db, file_md5: str, exclude_id: int) -> list:
    """查询同一 PDF（相同 file_md5）的其他保单记录摘要，每个 policy_index 只保留最新一条"""
    if not file_md5:  # 空 md5 不能作分组键
        return []
    conn = db.pool.connection()
    try:
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        # 按 policy_index 分组，取最新（id 最大）的记录
        cursor.execute(
            "SELECT r.id, r.policy_index, r.policy_count, r.page_range, "
            "       r.doc_category, r.status, r.parsed_fields "
            "FROM insurance_records r "
            "INNER JOIN ("
            "  SELECT MAX(id) AS max_id FROM insurance_records "
            "  WHERE file_md5 = %s AND id != %s "
            "  GROUP BY COALESCE(NULLIF(policy_index, 0), id)"
            ") latest ON r.id = latest.max_id "
            "ORDER BY r.policy_index",
            (file_md5, exclude_id)
        )
        rows = cursor.fetchall()
        result = []
        for row in rows:
            parsed = {}
            if row.get("parsed_fields"):
                try:
                    parsed = json.loads(row["parsed_fields"]) if isinstance(row["parsed_fields"], str) else row["parsed_fields"]
                except (TypeError, json.JSONDecodeError):
                    pass
            row["summary"] = parsed.get("险种类型", "") or parsed.get("保单号", "") or row.get("doc_category", "")
            row.pop("parsed_fields", None)
            result.append(dict(row))
        return result
    finally:
        conn.close()


# ============================================================
# 手动上传识别
# ============================================================

def _save_manual_record(file_name, policy, ocr_engine, file_data_b64=None, upload_cos=False,
                        policy_count=1, policy_index=1):
    """
    保存手动识别记录到数据库（按保单号去重：存在则更新，否则新建）。
    upload_cos=True 时上传文件到 COS。
    返回 (record_id, is_update)。
    """
    parsed_fields = policy.get("fields", {})
    doc_category = policy.get("doc_category", "")
    confidence = policy.get("confidence", 0.0)

    # 字段映射
    company_short = parsed_fields.get("保险公司简称", "")
    mapped_fields = apply_mapping(parsed_fields, company_short)

    cos_url = ""
    # 上传 COS（仅在开关打开且有文件数据时）
    if upload_cos and file_data_b64 and _handler and getattr(_handler, "cos_storage", None):
        try:
            import os
            import time
            pdf_bytes = base64.b64decode(file_data_b64)
            name, ext = os.path.splitext(file_name)
            key_suffix = f"insurance/manual/{file_name}"
            if _handler.cos_storage.file_exists(key_suffix):
                ts = int(time.time() * 1000)
                key_suffix = f"insurance/manual/{name}_{ts}{ext}"
            cos_url = _handler.cos_storage.upload_bytes(pdf_bytes, key_suffix) or ""
        except Exception as e:
            logger.warning("手动上传 COS 失败: %s", e)

    # 计算文件 MD5（用于去重）
    file_md5 = None
    if file_data_b64:
        try:
            raw_bytes = base64.b64decode(file_data_b64)
            file_md5 = hashlib.md5(raw_bytes).hexdigest()
        except Exception:
            pass

    # 关联当前登录用户和企业
    from flask import g
    current_user_id = g.current_user["user_id"] if hasattr(g, "current_user") else None
    if hasattr(g, "current_user"):
        _cu = g.current_user
        current_enterprise_id = _cu.get("parent_id") or _cu["user_id"]
    else:
        current_enterprise_id = None

    page_range = policy.get("page_range", "")
    record = {
        "filename": file_name,
        "cos_url": cos_url,
        "ocr_engine": ocr_engine,
        "doc_category": doc_category,
        "confidence": confidence,
        "parsed_fields": parsed_fields,
        "mapped_fields": mapped_fields,
        "status": "done",
        "source": "manual",
        "user_id": current_user_id,
        "enterprise_id": current_enterprise_id,
        "policy_count": policy_count,
        "policy_index": policy_index,
        "page_range": page_range,
    }
    if file_md5:
        record["file_md5"] = file_md5

    try:
        record_id, is_update = upsert_insurance_record_by_policy(_db, record)
        logger.info("手动识别记录已%s, id=%d, file=%s",
                     "更新" if is_update else "新建", record_id, file_name)
        return record_id, is_update
    except Exception as e:
        logger.warning("手动识别记录保存失败: %s", e)
        return None, False


@insurance_bp.route("/api/insurance/parse", methods=["POST"])
def parse_text_only():
    """
    纯文本解析（不做 OCR），接收前端提取的文本直接解析保单字段。
    请求体: {text, file_name, upload_cos, file_data(可选，upload_cos=true时用于上传COS)}
    """
    try:
        body = request.get_json(force=True) or {}
        text = body.get("text", "")
        file_name = body.get("file_name", "upload")
        upload_cos = body.get("upload_cos", False)
        file_data_b64 = body.get("file_data", "")

        if not text:
            return jsonify({
                "success": False, "file_name": file_name,
                "error": "文本为空", "invoices": [],
            })

        from .ocr_service import clean_text
        text = clean_text(text)
        policies = parse_policy_text_multi(text)

        # 过滤无效保单
        valid_policies = [p for p in policies if p.get("fields") and p.get("confidence", 0) > 0]
        if not valid_policies:
            return jsonify({
                "success": False, "file_name": file_name,
                "error": "need_ocr", "invoices": [],
            })

        # 同一 PDF 多保单互补基础字段
        if len(valid_policies) > 1 and _handler:
            _handler._cross_fill_same_pdf(valid_policies)

        # 逐条保存识别记录到数据库
        total = len(valid_policies)
        record_ids = []
        is_updates = []
        for idx, policy in enumerate(valid_policies):
            rid, is_upd = _save_manual_record(
                file_name, policy, "pdfjs",
                file_data_b64=file_data_b64, upload_cos=upload_cos,
                policy_count=total, policy_index=idx + 1)
            record_ids.append(rid)
            is_updates.append(is_upd)

        return jsonify({
            "success": True,
            "file_name": file_name,
            "invoices": valid_policies,
            "char_count": len(text),
            "ocr_engine": "pdfjs",
            "record_id": record_ids[0] if record_ids else None,
            "record_ids": record_ids,
            "is_update": is_updates[0] if is_updates else False,
        })
    except Exception as e:
        logger.exception("文本解析失败")
        return jsonify({
            "success": False, "file_name": file_name,
            "error": str(e), "invoices": [],
        })


@insurance_bp.route("/api/insurance/ocr", methods=["POST"])
def manual_ocr():
    """
    手动上传文件进行保单识别，识别结果保存到数据库（按保单号去重）。
    upload_cos=true 时同时上传文件到 COS。
    请求体: {file_data(base64), file_type, file_name, pdf_page, upload_cos}
    """
    try:
        _cu = g.current_user if hasattr(g, "current_user") else None
        if _cu and _cu.get("role") != ROLE_SUPER_ADMIN:
            # 未绑定企业：不允许上传
            if not _cu.get("parent_id"):
                return jsonify({
                    "code": 403,
                    "msg": "您的账号尚未绑定企业，请联系管理员分配企业后再上传",
                    "no_enterprise": True,
                }), 403

            # 检查企业月度配额（标准版限 3000 次/月）
            from auth.db import check_enterprise_quota, get_enterprise_by_id
            _eid = _cu.get("parent_id")
            # 企业未设置套餐（理论上不应出现，但做兜底）
            _ent = get_enterprise_by_id(_db, _eid)
            if not _ent or not _ent.get("plan_type"):
                return jsonify({
                    "code": 403,
                    "msg": "您所属的企业尚未开通套餐，请联系管理员配置套餐后再上传",
                    "no_plan": True,
                }), 403

            _quota = check_enterprise_quota(_db, _eid)
            if _quota.get("expired"):
                return jsonify({
                    "code": 403,
                    "msg": "套餐已到期，请联系管理员续费后再上传",
                    "plan_expired": True,
                }), 403
            if _quota.get("exceeded"):
                return jsonify({
                    "code": 429,
                    "msg": f"本月PDF识别额度已用完（{_quota['usage']}/{_quota['limit']}次），请联系管理员升级套餐",
                    "quota_exceeded": True,
                }), 429

        body = request.get_json(force=True) or {}
        file_data_b64 = body.get("file_data", "")
        file_type = body.get("file_type", "pdf")
        file_name = body.get("file_name", "upload")
        pdf_page = body.get("pdf_page", 0) or 0  # 0=提取所有页
        upload_cos = body.get("upload_cos", False)

        if not file_data_b64:
            return jsonify({"code": 400, "msg": "缺少 file_data 参数"}), 400

        # 初始化百度 OCR（从 handler 或配置中获取）
        baidu_ocr_client = getattr(_handler, "_baidu_ocr", None) if _handler else None

        extract_result = None
        ocr_engine = "pdfplumber"

        # 图片文件直接走百度 OCR
        if file_type in ("jpg", "jpeg", "png", "bmp"):
            if not baidu_ocr_client:
                return jsonify({
                    "success": False, "file_name": file_name,
                    "error": "图片识别需要配置百度OCR", "invoices": [],
                })
            extract_result = baidu_ocr_client.recognize_image(file_data_b64)
            ocr_engine = "baidu"
        else:
            # PDF：先用 pdfplumber 提取文本层
            extract_result = extract_text_from_pdf(
                file_data_base64=file_data_b64, pdf_page=pdf_page)

            # pdfplumber 失败（无文本层）→ 降级到百度 OCR（多页识别）
            if not extract_result["success"] and baidu_ocr_client:
                logger.info("[%s] pdfplumber 无文本层，降级使用百度OCR（多页）", file_name)
                pdf_bytes = base64.b64decode(file_data_b64)
                extract_result = baidu_ocr_client.recognize_pdf_multi_pages(
                    pdf_bytes, max_pages=6)
                ocr_engine = "baidu"

        if not extract_result["success"]:
            error_msg = extract_result.get("error", "OCR 识别失败")
            if "无文本层" in error_msg and not baidu_ocr_client:
                error_msg += "（提示：配置百度OCR可自动识别扫描件）"
            return jsonify({
                "success": False, "file_name": file_name,
                "error": error_msg, "invoices": [],
            })

        policies = parse_policy_text_multi(extract_result["text"])

        # 质量不佳时降级百度 OCR（用第一条保单判断质量）
        first_policy = policies[0] if policies else {}
        need_fallback = (
            not first_policy.get("fields")
            or first_policy.get("doc_category") == "其他"
            or first_policy.get("confidence", 0) == 0
        )
        if need_fallback and ocr_engine == "pdfplumber" and baidu_ocr_client:
            logger.info("[%s] pdfplumber 效果不佳(类型=%s, 置信度=%s), 降级百度OCR（多页）",
                        file_name, first_policy.get("doc_category"), first_policy.get("confidence"))
            pdf_bytes = base64.b64decode(file_data_b64)
            extract_result = baidu_ocr_client.recognize_pdf_multi_pages(
                pdf_bytes, max_pages=6)
            ocr_engine = "baidu"
            if extract_result["success"]:
                policies = parse_policy_text_multi(extract_result["text"])

        # 过滤无效保单
        valid_policies = [p for p in policies if p.get("fields") and p.get("confidence", 0) > 0]
        if not valid_policies:
            return jsonify({
                "success": False, "file_name": file_name,
                "error": "未识别到任何保单关键字段", "invoices": [],
            })

        # 同一 PDF 多保单互补基础字段
        if len(valid_policies) > 1 and _handler:
            _handler._cross_fill_same_pdf(valid_policies)

        # 逐条处理：同车牌互补 + 保存记录
        total = len(valid_policies)
        record_ids = []
        is_updates = []
        for idx, policy in enumerate(valid_policies):
            if _handler and policy.get("fields"):
                _handler._cross_fill_by_plate(policy["fields"])
                _handler._cross_fill_by_person(policy["fields"])
            rid, is_upd = _save_manual_record(
                file_name, policy, ocr_engine,
                file_data_b64=file_data_b64, upload_cos=upload_cos,
                policy_count=total, policy_index=idx + 1)
            record_ids.append(rid)
            is_updates.append(is_upd)

        return jsonify({
            "success": True,
            "file_name": file_name,
            "invoices": valid_policies,
            "char_count": extract_result.get("char_count", 0),
            "ocr_engine": ocr_engine,
            "record_id": record_ids[0] if record_ids else None,
            "record_ids": record_ids,
            "is_update": is_updates[0] if is_updates else False,
        })
    except Exception as e:
        logger.exception("手动上传识别失败")
        return jsonify({
            "success": False, "file_name": file_name,
            "error": str(e), "invoices": [],
        })


# ============================================================
# 重试 / 同步
# ============================================================

@insurance_bp.route("/api/insurance/retry/<int:record_id>", methods=["POST"])
def retry_record(record_id):
    """
    重试失败的保单识别记录。
    逻辑：从 messages 表取原始 sdkfileid → 删除旧记录 → 重新入队
    """
    if not _handler:
        return jsonify({"code": 503, "msg": "识别服务未初始化"}), 503

    try:
        # 1. 获取原始记录
        record = get_insurance_record(_db, record_id)
        if not record:
            return jsonify({"code": 404, "msg": "记录不存在"}), 404

        # 2. 仅允许重试失败状态的记录
        if record.get("status") != "failed":
            return jsonify({"code": 400, "msg": f"当前状态 {record.get('status')} 不允许重试，只有 failed 状态可重试"}), 400

        # 3. 手动上传的记录没有 msg_seq，无法重试
        msg_seq = record.get("msg_seq")
        if not msg_seq:
            return jsonify({"code": 400, "msg": "手动上传的记录无法重试（缺少消息序号）"}), 400

        # 4. 从 messages 表获取原始 sdkfileid
        conn = _db.pool.connection()
        try:
            cursor = conn.cursor(pymysql.cursors.DictCursor)
            cursor.execute(
                "SELECT parsed_content FROM messages WHERE seq = %s",
                (msg_seq,)
            )
            msg_row = cursor.fetchone()
        finally:
            conn.close()

        if not msg_row:
            return jsonify({"code": 404, "msg": f"未找到消息记录 seq={msg_seq}"}), 404

        # 解析 parsed_content 获取 sdkfileid
        parsed_content = msg_row.get("parsed_content") or "{}"
        if isinstance(parsed_content, str):
            try:
                parsed_content = json.loads(parsed_content)
            except (TypeError, json.JSONDecodeError):
                parsed_content = {}

        sdkfileid = parsed_content.get("sdkfileid", "")
        if not sdkfileid:
            return jsonify({"code": 400, "msg": "消息中未找到 sdkfileid，无法重试"}), 400

        # 5. 删除旧的识别记录
        conn2 = _db.pool.connection()
        try:
            cursor2 = conn2.cursor(pymysql.cursors.DictCursor)
            cursor2.execute("DELETE FROM insurance_records WHERE id = %s", (record_id,))
            conn2.commit()
        finally:
            conn2.close()

        # 6. 重新入队
        _handler.enqueue({
            "seq": msg_seq,
            "roomid": record.get("roomid", ""),
            "sender": record.get("sender", ""),
            "filename": record.get("filename", ""),
            "sdkfileid": sdkfileid,
            "filesize": record.get("filesize", 0),
        })

        return jsonify({"code": 0, "msg": "已重新加入识别队列"})
    except Exception as e:
        logger.exception("重试保单记录 %d 失败", record_id)
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/reupload-reocr/<int:record_id>", methods=["POST"])
def reupload_reocr(record_id):
    """
    重新上传 PDF 并识别，更新已有记录。
    请求体: {file_data(base64), file_name}
    """
    import base64 as b64mod

    if not _handler:
        return jsonify({"code": 503, "msg": "识别服务未初始化"}), 503

    try:
        record = get_insurance_record(_db, record_id)
        if not record:
            return jsonify({"code": 404, "msg": "记录不存在"}), 404

        body = request.get_json(force=True) or {}
        file_data_b64 = body.get("file_data", "")
        file_name = body.get("file_name", record.get("filename", "upload.pdf"))

        if not file_data_b64:
            return jsonify({"code": 400, "msg": "缺少 file_data 参数"}), 400

        pdf_bytes = b64mod.b64decode(file_data_b64)

        # 上传到 COS
        cos_url = ""
        if _handler.cos_storage:
            try:
                roomid = record.get("roomid", "manual")
                sender = record.get("sender", "manual")
                cos_key = f"insurance/{roomid}/{sender}/{file_name}"
                cos_url = _handler.cos_storage.upload_bytes(pdf_bytes, cos_key)
            except Exception as e:
                logger.warning("重新上传 COS 失败: %s", e)

        # 更新 COS URL 和文件名
        update_fields = {"filename": file_name}
        if cos_url:
            update_fields["cos_url"] = cos_url
        file_md5 = hashlib.md5(pdf_bytes).hexdigest()
        update_fields["file_md5"] = file_md5
        update_insurance_record(_db, record_id, update_fields)

        # OCR 识别
        ocr_result = _handler._do_ocr(pdf_bytes, file_name)
        if not ocr_result.get("success"):
            return jsonify({"code": 500, "msg": f"OCR 识别失败: {ocr_result.get('error', '未知错误')}"}), 500

        policies = ocr_result.get("policies", [])
        ocr_engine = ocr_result.get("ocr_engine", "pdfplumber")

        if not policies:
            return jsonify({"code": 500, "msg": "OCR 未解析到任何保单"}), 500

        best_policy = policies[0]
        parsed_fields = best_policy.get("fields") or {}

        if len(policies) > 1:
            _handler._cross_fill_same_pdf(policies)
            parsed_fields = best_policy.get("fields") or {}

        _handler._cross_fill_by_plate(parsed_fields, record_id)
        _handler._cross_fill_by_person(parsed_fields, record_id)

        from insurance.field_mapping import apply_mapping
        company_short = parsed_fields.get("保险公司简称", "")
        mapped_fields = apply_mapping(parsed_fields, company_short)

        raw_text = best_policy.get("raw_text", "")
        update_insurance_record(_db, record_id, {
            "status": "done",
            "ocr_engine": ocr_engine,
            "doc_category": best_policy.get("doc_category", ""),
            "confidence": best_policy.get("confidence", 0.0),
            "parsed_fields": parsed_fields,
            "mapped_fields": mapped_fields,
            "raw_text": raw_text,
            "error_message": "",
            "policy_count": len(policies),
            "policy_index": 1,
        })

        updated = get_insurance_record(_db, record_id)
        # 重新识别后也检测新保司（与自动识别共用同一函数，避免手动重新识别漏报）
        from insurance.handler import maybe_notify_new_company
        maybe_notify_new_company(_db, parsed_fields, best_policy.get("doc_category", ""),
                                 company_short, (updated or {}).get("filename", ""), record_id)
        return jsonify({"code": 0, "msg": "重新上传识别成功", "data": updated})

    except Exception as e:
        logger.exception("重新上传识别记录 %d 失败", record_id)
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/reocr/<int:record_id>", methods=["POST"])
def reocr_record(record_id):
    """
    重新识别：优先从 COS 下载，无 COS URL 时尝试通过 SDK 重新下载。
    不会自动同步钉钉，需要用户预览确认后手动触发同步。
    """
    import requests as http_requests

    if not _handler:
        return jsonify({"code": 503, "msg": "识别服务未初始化"}), 503

    try:
        record = get_insurance_record(_db, record_id)
        if not record:
            return jsonify({"code": 404, "msg": "记录不存在"}), 404

        cos_url = record.get("cos_url", "")
        pdf_bytes = b""

        if cos_url:
            # 优先从 COS 下载
            resp = http_requests.get(cos_url, timeout=60)
            if resp.status_code == 200:
                pdf_bytes = resp.content

        if not pdf_bytes:
            # 无 COS URL 或 COS 下载失败，尝试通过 SDK 重新下载
            pdf_bytes = _try_sdk_download(record)
            if pdf_bytes and _handler and _handler.cos_storage:
                # SDK 下载成功，上传到 COS 以便后续使用
                try:
                    roomid = record.get("roomid", "unknown")
                    sender = record.get("sender", "unknown")
                    filename = record.get("filename", "file.pdf")
                    cos_key = f"insurance/{roomid}/{sender}/{filename}"
                    cos_url = _handler.cos_storage.upload_bytes(pdf_bytes, cos_key)
                    if cos_url:
                        update_insurance_record(_db, record_id, {"cos_url": cos_url})
                except Exception as e:
                    logger.warning("重新识别时上传 COS 失败: %s", e)

        if not pdf_bytes:
            return jsonify({
                "code": 400,
                "msg": "文件已过期无法下载，请在群内重新发送该保单文件",
                "need_resend": True,
            }), 400

        # 2. 重新 OCR
        filename = record.get("filename", "file.pdf")
        ocr_result = _handler._do_ocr(pdf_bytes, filename)
        if not ocr_result.get("success"):
            return jsonify({"code": 500, "msg": f"OCR 识别失败: {ocr_result.get('error', '未知错误')}"}), 500

        policies = ocr_result.get("policies", [])
        ocr_engine = ocr_result.get("ocr_engine", "pdfplumber")
        reocr_ocr_text = ocr_result.get("ocr_text", "")

        if not policies:
            return jsonify({"code": 500, "msg": "OCR 未解析到任何保单"}), 500

        from insurance.field_mapping import apply_mapping

        # 同一 PDF 多保单互补
        if len(policies) > 1:
            _handler._cross_fill_same_pdf(policies)

        total_policies = len(policies)
        file_md5 = record.get("file_md5", "")

        # 收集同一 PDF 的兄弟记录（按 policy_index 去重，每个 index 只保留最新）
        sibling_records = []
        if file_md5:
            conn = _db.pool.connection()
            try:
                cursor = conn.cursor(pymysql.cursors.DictCursor)
                cursor.execute(
                    "SELECT * FROM insurance_records WHERE file_md5 = %s AND id != %s",
                    (file_md5, record_id)
                )
                for sib in cursor.fetchall():
                    for f in ("parsed_fields", "mapped_fields"):
                        if isinstance(sib.get(f), str):
                            try: sib[f] = json.loads(sib[f])
                            except: sib[f] = {}
                    sibling_records.append(dict(sib))
            finally:
                conn.close()
            # 按 policy_index 去重，每个 index 只保留最新
            by_index = {}
            for sib in sibling_records:
                idx = sib.get("policy_index", 0) or 0
                if idx not in by_index or sib["id"] > by_index[idx]["id"]:
                    by_index[idx] = sib
            sibling_records = list(by_index.values())

        # 为当前记录匹配最佳 policy（通过保单号匹配）
        original_policy_no = ""
        try:
            orig_fields = record.get("parsed_fields") or {}
            if isinstance(orig_fields, str):
                orig_fields = json.loads(orig_fields) if orig_fields else {}
            original_policy_no = orig_fields.get("保单号", "")
        except (TypeError, json.JSONDecodeError):
            pass

        best_idx = 0
        if original_policy_no and len(policies) > 1:
            for i, p in enumerate(policies):
                if p.get("fields", {}).get("保单号", "") == original_policy_no:
                    best_idx = i
                    break
        used_policies = {best_idx}

        # 更新当前记录
        def _apply_policy_to_record(rec_id, policy, policy_idx):
            pf = policy.get("fields") or {}
            _handler._cross_fill_by_plate(pf, rec_id)
            _handler._cross_fill_by_person(pf, rec_id)
            # TZ-008 专属：车牌为空时用车架号填入车牌号（持久化）
            from insurance.db import apply_tz008_plate_fallback
            apply_tz008_plate_fallback(_db, pf, record.get("enterprise_id"))
            cs = pf.get("保险公司简称", "")
            mf = apply_mapping(pf, cs)
            doc_cat = policy.get("doc_category", "")
            # 文件名带"投保"时文档类型标记为投保单
            if filename and "投保" in filename:
                doc_cat = "投保单"
                pf["文档类型"] = "投保单"
            updates = {
                "status": "done",
                "ocr_engine": ocr_engine,
                "doc_category": doc_cat,
                "confidence": policy.get("confidence", 0.0),
                "parsed_fields": pf,
                "mapped_fields": mf,
                "raw_text": policy.get("raw_text", ""),
                "error_message": "",
                "cos_url": cos_url,
                "policy_count": total_policies,
                "policy_index": policy_idx + 1,
                "page_range": policy.get("page_range", ""),
            }
            if reocr_ocr_text:
                updates["ocr_text"] = reocr_ocr_text
            update_insurance_record(_db, rec_id, updates)
            # 重新识别也检测新保司（与自动识别共用同一函数；函数自带去重，多保单不会重复通知）
            from insurance.handler import maybe_notify_new_company
            maybe_notify_new_company(_db, pf, doc_cat, cs, filename, rec_id)

        _apply_policy_to_record(record_id, policies[best_idx], best_idx)
        updated_count = 1

        # 为每条兄弟记录匹配 policy 并更新
        # 同步主记录的归属信息到兄弟记录（企业ID、用户ID）
        owner_sync = {}
        if record.get("enterprise_id") is not None:
            owner_sync["enterprise_id"] = record["enterprise_id"]
        if record.get("user_id") is not None:
            owner_sync["user_id"] = record["user_id"]

        for sib in sibling_records:
            sib_pf = sib.get("parsed_fields", {})
            sib_no = sib_pf.get("保单号", "") if isinstance(sib_pf, dict) else ""

            matched_idx = None
            # 通过保单号匹配
            if sib_no:
                for i, p in enumerate(policies):
                    if i not in used_policies and p.get("fields", {}).get("保单号", "") == sib_no:
                        matched_idx = i
                        break
            # 未匹配则取第一个未使用的
            if matched_idx is None:
                for i in range(len(policies)):
                    if i not in used_policies:
                        matched_idx = i
                        break
            if matched_idx is None:
                continue

            used_policies.add(matched_idx)
            _apply_policy_to_record(sib["id"], policies[matched_idx], matched_idx)
            # 同步归属信息
            if owner_sync:
                update_insurance_record(_db, sib["id"], owner_sync)
            updated_count += 1

        # 为多出的保单新建记录（沿用主记录的创建时间）
        original_created_at = record.get("created_at", "")
        new_sibling_ids = []
        logger.info(
            "重新识别多保单: total_policies=%d, used_policies=%s, sibling_count=%d, file_md5=%s",
            total_policies, used_policies, len(sibling_records), file_md5,
        )
        for i in range(len(policies)):
            if i in used_policies:
                continue
            policy = policies[i]
            pf = policy.get("fields", {})
            policy_type = pf.get("险种类型", "未知")
            logger.info(
                "重新识别需新建第%d/%d份保单: 险种=%s, file_md5=%s",
                i + 1, total_policies, policy_type, file_md5,
            )
            _handler._cross_fill_by_plate(pf, record_id)
            _handler._cross_fill_by_person(pf, record_id)
            from insurance.db import apply_tz008_plate_fallback
            apply_tz008_plate_fallback(_db, pf, record.get("enterprise_id"))
            cs = pf.get("保险公司简称", "")
            mf = apply_mapping(pf, cs)
            extra_record = {
                "msg_seq": record.get("msg_seq", 0),
                "roomid": record.get("roomid", ""),
                "room_name": record.get("room_name", ""),
                "sender": record.get("sender", ""),
                "sender_name": record.get("sender_name", ""),
                "filename": record.get("filename", ""),
                "filesize": record.get("filesize", 0),
                "file_md5": file_md5,
                "cos_url": cos_url,
                "status": "done",
                "source": record.get("source", "auto"),
                "user_id": record.get("user_id", ""),
                "enterprise_id": record.get("enterprise_id", ""),
                "ocr_engine": ocr_engine,
                "doc_category": policy.get("doc_category", ""),
                "confidence": policy.get("confidence", 0.0),
                "parsed_fields": pf,
                "mapped_fields": mf,
                "raw_text": policy.get("raw_text", ""),
                "policy_count": total_policies,
                "policy_index": i + 1,
                "page_range": policy.get("page_range", ""),
                "created_at": original_created_at,
            }
            try:
                new_id = save_insurance_record(_db, extra_record)
                new_sibling_ids.append(new_id)
                updated_count += 1
                logger.info("重新识别新增保单记录成功 record_id=%d, policy_index=%d/%d, 险种=%s", new_id, i + 1, total_policies, policy_type)
            except Exception as e:
                logger.error(
                    "重新识别新增第%d/%d条记录失败(险种=%s, file_md5=%s): %s",
                    i + 1, total_policies, policy_type, file_md5, e,
                    exc_info=True,
                )

        # 最终处理：统一更新所有兄弟记录的 policy_count + 字段互补
        if file_md5:
            all_ids = [record_id] + [s["id"] for s in sibling_records] + new_sibling_ids
            for rid in all_ids:
                rec = get_insurance_record(_db, rid)
                if not rec:
                    continue
                updates = {}
                # 统一修正 policy_count（以本次 OCR 结果为准）
                if rec.get("policy_count") != total_policies:
                    updates["policy_count"] = total_policies
                # 兄弟字段互补
                if rec.get("status") == "done":
                    pf = rec.get("parsed_fields", {})
                    if isinstance(pf, str):
                        try: pf = json.loads(pf) if pf else {}
                        except: pf = {}
                    before = {f: pf.get(f, "") for f in _handler.SAME_PDF_FILL_FIELDS}
                    _cross_fill_from_siblings(_db, rid, file_md5, pf)
                    after = {f: pf.get(f, "") for f in _handler.SAME_PDF_FILL_FIELDS}
                    if before != after:
                        cs = pf.get("保险公司简称", "")
                        updates["parsed_fields"] = pf
                        updates["mapped_fields"] = apply_mapping(pf, cs)
                if updates:
                    update_insurance_record(_db, rid, updates)

        # 返回更新后的主记录
        updated = get_insurance_record(_db, record_id)
        for field in ("parsed_fields", "mapped_fields"):
            if isinstance(updated.get(field), str):
                try:
                    updated[field] = json.loads(updated[field])
                except (TypeError, json.JSONDecodeError):
                    pass
        for ts_field in ("created_at", "updated_at"):
            if updated.get(ts_field) and not isinstance(updated[ts_field], str):
                updated[ts_field] = str(updated[ts_field])

        msg = "重新识别完成，请预览确认后同步钉钉"
        if updated_count > 1:
            msg = f"重新识别完成，已同步更新 {updated_count} 条记录"

        return jsonify({
            "code": 0,
            "data": updated,
            "msg": msg,
        })
    except Exception as e:
        logger.exception("重新识别保单记录 %d 失败", record_id)
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/sync/<int:record_id>", methods=["POST"])
def sync_record_to_dingtalk(record_id):
    """手动触发单条记录同步到钉钉（支持覆盖已同步的记录）"""
    if not _handler:
        return jsonify({"code": 503, "msg": "识别服务未初始化"}), 503

    try:
        record = get_insurance_record(_db, record_id)
        if not record:
            return jsonify({"code": 404, "msg": "记录不存在"}), 404

        # 重置同步状态，允许重新同步
        if record.get("dingtalk_synced"):
            update_insurance_record(_db, record_id, {"dingtalk_synced": 0})

        # 解析字段
        parsed_fields = record.get("parsed_fields") or {}
        if isinstance(parsed_fields, str):
            try:
                parsed_fields = json.loads(parsed_fields)
            except (TypeError, json.JSONDecodeError):
                parsed_fields = {}
        mapped_fields = record.get("mapped_fields") or {}
        if isinstance(mapped_fields, str):
            try:
                mapped_fields = json.loads(mapped_fields)
            except (TypeError, json.JSONDecodeError):
                mapped_fields = {}

        # 如果没有 mapped_fields，重新生成
        if not mapped_fields and parsed_fields:
            company_short = parsed_fields.get("保险公司简称", "")
            mapped_fields = apply_mapping(parsed_fields, company_short)

        roomid = record.get("roomid", "")
        sender = record.get("sender", "")
        monitors = _handler._get_matched_monitors(roomid, sender)
        if not monitors:
            return jsonify({"code": 400, "msg": "该记录无匹配的监控配置，无法确定同步目标"}), 400

        synced = False
        for monitor in monitors:
            if not (monitor.get("dingtalk_base_id") and monitor.get("dingtalk_sheet_id")):
                continue
            monitor_mapping = monitor.get("field_mapping") or {}
            if monitor_mapping:
                from insurance.field_mapping import DEFAULT_MAPPING
                export_to_ocr = {v: k for k, v in DEFAULT_MAPPING.items() if v}
                sync_fields = {}
                for export_col, target_col in monitor_mapping.items():
                    if not target_col:
                        continue
                    ocr_field = export_to_ocr.get(export_col, export_col)
                    if ocr_field in parsed_fields:
                        sync_fields[target_col] = parsed_fields[ocr_field]
                    elif export_col in parsed_fields:
                        sync_fields[target_col] = parsed_fields[export_col]
            else:
                sync_fields = mapped_fields
            _handler._sync_to_dingtalk_v2(
                record_id, sync_fields,
                sender_name=record.get("sender_name", ""),
                cos_url=record.get("cos_url", ""),
                doc_category=record.get("doc_category", ""),
                monitor=monitor,
            )
            synced = True

        if synced:
            update_insurance_record(_db, record_id, {"dingtalk_synced": 1})
            return jsonify({"code": 0, "msg": "同步成功"})
        else:
            return jsonify({"code": 400, "msg": "监控配置未绑定钉钉文档"}), 400
    except Exception as e:
        logger.exception("同步记录 %d 到钉钉失败", record_id)
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/batch-reocr-sync", methods=["POST"])
def batch_reocr_sync():
    """
    批量重新识别并同步钉钉。
    对选中的记录逐条执行：重新OCR → 同步到钉钉。
    请求体: {record_ids: [int, ...]}
    """
    import requests as http_requests

    if not _handler:
        return jsonify({"code": 503, "msg": "识别服务未初始化"}), 503

    try:
        body = request.get_json(force=True) or {}
        record_ids = body.get("record_ids", [])

        if not record_ids:
            return jsonify({"code": 400, "msg": "缺少 record_ids 参数"}), 400

        results = {"reocr_success": 0, "reocr_failed": 0, "sync_success": 0, "sync_failed": 0, "errors": []}

        for rid in record_ids:
            # 1. 获取记录
            record = get_insurance_record(_db, rid)
            if not record:
                results["errors"].append({"id": rid, "error": "记录不存在"})
                results["reocr_failed"] += 1
                continue

            cos_url = record.get("cos_url", "")

            # 2. 重新识别
            try:
                pdf_bytes = b""
                if cos_url:
                    resp = http_requests.get(cos_url, timeout=60)
                    if resp.status_code == 200:
                        pdf_bytes = resp.content

                if not pdf_bytes:
                    # 无 COS URL 或 COS 下载失败，尝试 SDK 重新下载
                    pdf_bytes = _try_sdk_download(record)
                    if pdf_bytes and _handler and _handler.cos_storage:
                        try:
                            roomid = record.get("roomid", "unknown")
                            sender = record.get("sender", "unknown")
                            fname = record.get("filename", "file.pdf")
                            cos_key = f"insurance/{roomid}/{sender}/{fname}"
                            new_cos_url = _handler.cos_storage.upload_bytes(pdf_bytes, cos_key)
                            if new_cos_url:
                                update_insurance_record(_db, rid, {"cos_url": new_cos_url})
                        except Exception:
                            pass

                if not pdf_bytes:
                    results["errors"].append({"id": rid, "error": "文件已过期无法下载，请在群内重新发送"})
                    results["reocr_failed"] += 1
                    continue
                filename = record.get("filename", "file.pdf")
                ocr_result = _handler._do_ocr(pdf_bytes, filename)
                if not ocr_result.get("success"):
                    results["errors"].append({"id": rid, "error": f"OCR失败: {ocr_result.get('error', '')}"})
                    results["reocr_failed"] += 1
                    continue

                policies = ocr_result.get("policies", [])
                ocr_engine = ocr_result.get("ocr_engine", "pdfplumber")

                if not policies:
                    results["errors"].append({"id": rid, "error": "OCR 未解析到任何保单"})
                    results["reocr_failed"] += 1
                    continue

                # 匹配原记录对应的保单
                original_policy_no = ""
                try:
                    orig_fields = record.get("parsed_fields") or {}
                    if isinstance(orig_fields, str):
                        orig_fields = json.loads(orig_fields) or {}
                    if isinstance(orig_fields, dict):
                        original_policy_no = orig_fields.get("保单号", "")
                except Exception:
                    pass

                best_policy = policies[0]
                if original_policy_no and len(policies) > 1:
                    for p in policies:
                        if p.get("fields", {}).get("保单号", "") == original_policy_no:
                            best_policy = p
                            break

                parsed_fields = best_policy.get("fields") or {}
                doc_category = best_policy.get("doc_category", "")
                confidence = best_policy.get("confidence", 0.0)

                # 同一 PDF 多保单互补
                if len(policies) > 1:
                    _handler._cross_fill_same_pdf(policies)
                    parsed_fields = best_policy.get("fields") or {}

                # 同 file_md5 兄弟记录互补
                rec_md5 = record.get("file_md5", "")
                if rec_md5:
                    _cross_fill_from_siblings(_db, rid, rec_md5, parsed_fields)

                # 同车牌互补 + 按人名互补
                _handler._cross_fill_by_plate(parsed_fields, rid)
                _handler._cross_fill_by_person(parsed_fields, rid)

                # 字段映射
                company_short = parsed_fields.get("保险公司简称", "")
                mapped_fields = apply_mapping(parsed_fields, company_short)

                # 更新记录
                raw_text = best_policy.get("raw_text", "")
                page_range = best_policy.get("page_range", "")
                update_insurance_record(_db, rid, {
                    "status": "done",
                    "ocr_engine": ocr_engine,
                    "doc_category": doc_category,
                    "confidence": confidence,
                    "parsed_fields": parsed_fields,
                    "mapped_fields": mapped_fields,
                    "raw_text": raw_text,
                    "error_message": "",
                    "policy_count": len(policies),
                    "policy_index": (policies.index(best_policy) + 1) if best_policy in policies else 1,
                    "page_range": page_range,
                })
                results["reocr_success"] += 1

            except Exception as e:
                logger.exception("批量重新识别 record_id=%d 异常", rid)
                results["errors"].append({"id": rid, "error": f"识别异常: {str(e)[:200]}"})
                results["reocr_failed"] += 1
                continue

            # 3. 同步钉钉（使用监控配置绑定关系，与自动同步一致）
            try:
                roomid = record.get("roomid", "")
                sender = record.get("sender", "")
                sender_name = record.get("sender_name", "")
                cos_url = record.get("cos_url", "")
                monitors = _handler._get_matched_monitors(roomid, sender)
                if not monitors:
                    logger.info("record_id=%d 无匹配监控配置，跳过钉钉同步", rid)
                    results["sync_failed"] += 1
                    results["errors"].append({"id": rid, "error": "无匹配的监控配置"})
                else:
                    synced = False
                    for monitor in monitors:
                        if not (monitor.get("dingtalk_base_id") and monitor.get("dingtalk_sheet_id")):
                            continue
                        # 使用监控配置自身的字段映射（如果有）
                        monitor_mapping = monitor.get("field_mapping") or {}
                        if monitor_mapping:
                            from insurance.field_mapping import DEFAULT_MAPPING
                            export_to_ocr = {v: k for k, v in DEFAULT_MAPPING.items() if v}
                            sync_fields = {}
                            for export_col, target_col in monitor_mapping.items():
                                if not target_col:
                                    continue
                                ocr_field = export_to_ocr.get(export_col, export_col)
                                if ocr_field in parsed_fields:
                                    sync_fields[target_col] = parsed_fields[ocr_field]
                                elif export_col in parsed_fields:
                                    sync_fields[target_col] = parsed_fields[export_col]
                        else:
                            sync_fields = mapped_fields
                        _handler._sync_to_dingtalk_v2(
                            rid, sync_fields,
                            sender_name=sender_name,
                            cos_url=cos_url,
                            doc_category=doc_category,
                            monitor=monitor,
                        )
                        synced = True
                    if synced:
                        update_insurance_record(_db, rid, {"dingtalk_synced": 1})
                        results["sync_success"] += 1
                    else:
                        results["sync_failed"] += 1
                        results["errors"].append({"id": rid, "error": "监控配置未绑定钉钉文档"})
            except Exception as e:
                results["errors"].append({"id": rid, "error": f"同步失败: {str(e)[:200]}"})
                results["sync_failed"] += 1

        msg_parts = []
        if results["reocr_success"]:
            msg_parts.append(f"识别成功 {results['reocr_success']} 条")
        if results["sync_success"]:
            msg_parts.append(f"同步成功 {results['sync_success']} 条")
        if results["reocr_failed"]:
            msg_parts.append(f"识别失败 {results['reocr_failed']} 条")
        if results["sync_failed"]:
            msg_parts.append(f"同步失败 {results['sync_failed']} 条")

        return jsonify({
            "code": 0,
            "data": results,
            "msg": "、".join(msg_parts) if msg_parts else "无操作",
        })
    except Exception as e:
        logger.exception("批量重新识别同步失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


# ============================================================
# 配置管理
# ============================================================

@insurance_bp.route("/api/insurance/config", methods=["GET"])
def get_config():
    """获取全部保单识别配置"""
    try:
        config = get_all_insurance_config(_db, max_value_len=10000)
        return jsonify({"code": 0, "data": config})
    except Exception as e:
        logger.exception("获取保单配置失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/config", methods=["PUT"])
def update_config():
    """
    更新保单识别配置。
    请求体: {key: value, ...}
    更新 baidu_ocr 相关配置时，自动重载 OCR 客户端。
    """
    try:
        body = request.get_json(force=True) or {}
        if not body:
            return jsonify({"code": 400, "msg": "请求体不能为空"}), 400

        # 判断是否涉及 baidu_ocr 配置变更
        baidu_ocr_keys = {"baidu_ocr_app_id", "baidu_ocr_api_key", "baidu_ocr_secret_key", "baidu_ocr_enabled"}
        need_reload_ocr = bool(set(body.keys()) & baidu_ocr_keys)

        # 逐个写入配置
        for key, value in body.items():
            set_insurance_config(_db, key, value)

        # 重载百度 OCR 客户端
        if need_reload_ocr and _handler:
            try:
                _handler.reload_baidu_ocr()
                logger.info("百度 OCR 配置已重载")
            except Exception as reload_err:
                logger.warning("百度 OCR 重载失败: %s", reload_err)

        return jsonify({"code": 0, "msg": "配置已更新"})
    except Exception as e:
        logger.exception("更新保单配置失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


# ============================================================
# 备注快捷选择（选项池 + 配置）—— 按模板存储
# ============================================================

def _resolve_remark_settings_target(extra=None):
    """
    设置态：解析 (scope, scope_id, template_name)。
    - 超管带 target_scope → 代管写到对方企业/员工；否则按当前登录人 scope。
    - template_name 取自请求（args/form/body），默认「默认模板」。
    extra: 可传 JSON body（PUT 用）。
    """
    from flask import g
    user = g.current_user

    def _pick(key):
        v = request.values.get(key)
        if v is None and isinstance(extra, dict):
            v = extra.get(key)
        return v

    template_name = _pick("template_name") or _pick("template") or "默认模板"
    target_scope = _pick("target_scope")
    target_scope_id = _pick("target_scope_id")
    if user["role"] == ROLE_SUPER_ADMIN and target_scope:
        scope = target_scope
        scope_id = int(target_scope_id) if target_scope_id else None
    else:
        scope, scope_id = _get_user_scope()
    return scope, scope_id, template_name


def _resolve_remark_runtime_target():
    """
    运行态（台账列表）：按用户当前启用模板解析 (scope, scope_id, template_name)。
    逻辑与 get_effective_config 一致。
    """
    from flask import g
    user = g.current_user
    active = get_active_template(_db, user["user_id"])
    source = active.get("source", "own")
    template_name = active.get("template_name", "默认模板")
    role = user["role"]
    parent_id = user.get("parent_id")
    if source == "system":
        scope, scope_id = "global", None
    elif source == "enterprise":
        scope, scope_id = "enterprise", parent_id
    else:
        if role == ROLE_SUPER_ADMIN:
            scope, scope_id = "global", None
        elif role == ROLE_ENTERPRISE:
            scope, scope_id = "enterprise", parent_id
        else:
            scope, scope_id = "user", user["user_id"]
    return scope, scope_id, template_name


@insurance_bp.route("/api/insurance/remark-options/config", methods=["GET"])
@login_required
def get_remark_selector_config_api():
    """
    读取备注快捷选择配置 + 选项池统计。
    - 带 template 参数 → 设置态（读指定模板，支持代管）。
    - 不带 → 运行态（读当前启用模板）。
    """
    try:
        if request.args.get("template") or request.args.get("template_name"):
            scope, scope_id, template_name = _resolve_remark_settings_target()
        else:
            scope, scope_id, template_name = _resolve_remark_runtime_target()
        cfg = get_remark_selector_config(_db, scope, scope_id, template_name)
        stats = remark_options_db.get_stats(_db, scope, scope_id, template_name)
        return jsonify({"code": 0, "data": {"config": cfg, "stats": stats, "template_name": template_name}})
    except Exception as e:
        logger.exception("读取备注快捷选择配置失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/remark-options/config", methods=["PUT"])
@login_required
def save_remark_selector_config_api():
    """保存指定模板的备注快捷选择配置（支持代管）。"""
    try:
        body = request.get_json(force=True) or {}
        scope, scope_id, template_name = _resolve_remark_settings_target(extra=body)
        save_remark_selector_config(_db, scope, scope_id, body.get("config", body), template_name)
        return jsonify({"code": 0, "msg": "已保存"})
    except Exception as e:
        logger.exception("保存备注快捷选择配置失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/remark-options/sheets", methods=["POST"])
@login_required
def remark_options_sheets():
    """上传 Excel，返回 sheet 名列表（不落库）。"""
    try:
        f = request.files.get("file")
        if not f:
            return jsonify({"code": 400, "msg": "缺少文件"}), 400
        names = remark_options_db.read_sheet_names(f.read())
        return jsonify({"code": 0, "data": {"sheets": names}})
    except Exception as e:
        logger.exception("读取 sheet 列表失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/remark-options/import", methods=["POST"])
@login_required
def remark_options_import():
    """上传 Excel + 选定 sheet → 清空指定模板旧池 + 导入新池（支持代管）。"""
    try:
        f = request.files.get("file")
        sheet = request.form.get("sheet", "")
        if not f or not sheet:
            return jsonify({"code": 400, "msg": "缺少文件或 sheet"}), 400
        parsed = remark_options_db.parse_sheet_rows(f.read(), sheet)
        rows = parsed["rows"]
        missing = parsed["missing"]
        scope, scope_id, template_name = _resolve_remark_settings_target()
        n = remark_options_db.import_options(_db, scope, scope_id, template_name, rows)
        msg = f"导入成功 {n} 条"
        if missing:
            msg += f"；注意：该表缺少 [{'、'.join(missing)}] 列，对应筛选条件将无法生效"
        return jsonify({"code": 0, "msg": msg, "data": {"count": n, "missing": missing}})
    except Exception as e:
        logger.exception("导入备注选项池失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/remark-options/query", methods=["GET"])
@login_required
def remark_options_query():
    """按 handler/company/policy_type/keyword + mode 查询候选 remark（运行态，按启用模板）。"""
    try:
        scope, scope_id, template_name = _resolve_remark_runtime_target()
        opts = remark_options_db.query_options(
            _db, scope, scope_id, template_name,
            handler=request.args.get("handler", ""),
            company=request.args.get("company", ""),
            policy_type=request.args.get("policy_type", ""),
            keyword=request.args.get("keyword", ""),
            mode=request.args.get("mode", "exact"),
        )
        return jsonify({"code": 0, "data": {"options": opts}})
    except Exception as e:
        logger.exception("查询备注候选失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/remark-options", methods=["DELETE"])
@login_required
def remark_options_clear():
    """清空指定模板选项池（支持代管）。"""
    try:
        scope, scope_id, template_name = _resolve_remark_settings_target()
        n = remark_options_db.clear_options(_db, scope, scope_id, template_name)
        return jsonify({"code": 0, "msg": f"已清空 {n} 条"})
    except Exception as e:
        logger.exception("清空备注选项池失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/reload-ocr", methods=["POST"])
def reload_ocr():
    """重新加载所有 OCR 引擎配置（支持 config.yaml 热更新）"""
    try:
        if not _handler:
            return jsonify({"code": 500, "msg": "handler 未初始化"}), 500
        _handler.reload_all_ocr()
        # 返回各引擎状态
        engines = []
        if _handler._volc_ocr:
            engines.append("火山引擎: 已启用")
        if _handler._baidu_ocr:
            acc = {"accurate": "高精度版", "standard": "标准版"}.get(
                getattr(_handler._baidu_ocr, "accuracy", ""), "未知"
            )
            engines.append(f"百度OCR: {acc}")
        if not engines:
            engines.append("无可用引擎")
        return jsonify({"code": 0, "msg": "OCR 引擎已重载", "engines": engines})
    except Exception as e:
        logger.exception("重载 OCR 引擎失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/rooms", methods=["GET"])
def get_rooms():
    """
    获取可监控的对象列表（群聊 + 私聊联系人）。
    群名以统一群信息表 wecom_groups 为权威来源（task_group_sync 独立任务维护），
    过渡期回退 contacts_cache 旧缓存；不再现场调企微 API（接口毫秒级返回）。
    返回: {"rooms": [...], "users": [...]}
    """
    try:
        conn = _db.pool.connection()
        try:
            cursor = conn.cursor(pymysql.cursors.DictCursor)

            # 群聊列表：wecom_groups 优先，contacts_cache 兜底
            cursor.execute("""
                SELECT m.roomid AS id,
                       COALESCE(NULLIF(g.name, ''), NULLIF(NULLIF(c.name, '__unresolvable__'), '')) AS cached_name
                FROM (
                    SELECT DISTINCT roomid FROM messages
                    WHERE roomid IS NOT NULL AND roomid != ''
                ) m
                LEFT JOIN wecom_groups g ON g.roomid = m.roomid
                LEFT JOIN contacts_cache c ON c.id = m.roomid
                ORDER BY m.roomid
            """)
            room_rows = cursor.fetchall()

            # 私聊联系人
            cursor.execute("""
                SELECT m.sender AS id, c.name AS cached_name
                FROM (
                    SELECT DISTINCT sender FROM messages
                    WHERE (roomid IS NULL OR roomid = '') AND sender IS NOT NULL AND sender != ''
                ) m
                LEFT JOIN contacts_cache c ON c.id = m.sender
                ORDER BY m.sender
            """)
            user_rows = cursor.fetchall()
        finally:
            conn.close()

        # 群名已在 SQL 里从 wecom_groups/contacts_cache 取好，都没有的显示 roomid
        # （由 task_group_sync 独立任务负责解析，这里不再现场调企微 API）
        rooms = [
            {"id": r["id"], "name": r["cached_name"] or r["id"], "type": "room"}
            for r in room_rows
        ]

        users = []
        for u in user_rows:
            name = u["cached_name"]
            if name == "__unresolvable__":
                name = None
            if not name and _contacts:
                try:
                    name = _contacts.get_name(u["id"])
                except Exception:
                    pass
            users.append({"id": u["id"], "name": name or u["id"], "type": "user"})

        # 按名称排序
        rooms.sort(key=lambda x: x["name"])
        users.sort(key=lambda x: x["name"])

        return jsonify({"rooms": rooms, "users": users})
    except Exception as e:
        logger.exception("获取监控对象列表失败")
        return jsonify({"rooms": [], "users": []})


@insurance_bp.route("/api/insurance/monitored-rooms", methods=["GET"])
@login_required
def get_monitored_rooms():
    """返回当前账号"正在监控"的群与私聊（来自启用的监控配置），用于筛选下拉。
    与 /rooms（messages 里全部历史群）不同：这里只列当前监控配置中的对象，
    并按账号权限过滤——超管看全部配置，企业/员工只看本企业的配置。
    """
    try:
        eid = _get_enterprise_id_filter()  # 超管 None=看全部；其他角色=本企业 id
        configs = list_monitor_configs(_db)
        rooms, users = {}, {}
        room_ent = {}  # roomid -> 所属企业 id（供超管按企业筛选群下拉）
        for cfg in configs:
            if not cfg.get("enabled"):
                continue
            if eid is not None and cfg.get("enterprise_id") != eid:
                continue
            # rooms/users 可能是字符串数组（["wrxxx",...]）或对象数组（[{"id","name"},...]），兼容两种
            for room in (cfg.get("rooms") or []):
                rid = room.get("id") if isinstance(room, dict) else room
                rname = room.get("name") if isinstance(room, dict) else None
                if rid:
                    rooms[rid] = rname or rooms.get(rid) or rid
                    room_ent.setdefault(rid, cfg.get("enterprise_id"))
            for u in (cfg.get("users") or []):
                uid = u.get("id") if isinstance(u, dict) else u
                uname = u.get("name") if isinstance(u, dict) else None
                if uid:
                    users[uid] = uname or users.get(uid) or uid
        # 群名以统一群信息表 wecom_groups 为准（task_group_sync 维护，群改名会自动更新）；
        # 配置里固化的旧 name 仅在 wecom_groups 没有时作兜底
        if rooms:
            try:
                from storage.group_db import get_group_name_map
                gm = get_group_name_map(_db, list(rooms.keys()))
                for rid, gname in gm.items():
                    rooms[rid] = gname
            except Exception:
                logger.warning("查询统一群信息表失败，群名使用配置内固化值", exc_info=True)
        # 人名（及 wecom_groups 还没解析到的群）继续用 contacts_cache 补
        ids = [rid for rid in rooms if rooms[rid] == rid] + list(users.keys())
        if ids:
            conn = _db.pool.connection()
            try:
                cur = conn.cursor(pymysql.cursors.DictCursor)
                ph = ",".join(["%s"] * len(ids))
                cur.execute(f"SELECT id, name FROM contacts_cache WHERE id IN ({ph})", ids)
                nm = {r["id"]: r["name"] for r in cur.fetchall()
                      if r["name"] and r["name"] != "__unresolvable__"}
            finally:
                conn.close()
            for rid in rooms:
                if rooms[rid] == rid and nm.get(rid):
                    rooms[rid] = nm[rid]
            for uid in users:
                if users[uid] == uid and nm.get(uid):
                    users[uid] = nm[uid]
        room_list = sorted([{"id": k, "name": v, "enterprise_id": room_ent.get(k)} for k, v in rooms.items()], key=lambda x: x["name"])
        user_list = sorted([{"id": k, "name": v} for k, v in users.items()], key=lambda x: x["name"])
        return jsonify({"code": 0, "data": {"rooms": room_list, "users": user_list}})
    except Exception:
        logger.exception("获取监控群列表失败")
        return jsonify({"code": 0, "data": {"rooms": [], "users": []}})


@insurance_bp.route("/api/insurance/sync-contacts", methods=["POST"])
def sync_contacts():
    """
    手动同步通讯录：强制刷新所有群和联系人的名称缓存。
    遍历 messages 表中所有 roomid 和 sender，逐个调企微 API 更新缓存。
    """
    if not _contacts:
        return jsonify({"error": "通讯录模块未初始化"}), 503

    conn = _db.pool.connection()
    try:
        cursor = conn.cursor(pymysql.cursors.DictCursor)

        # 所有群 ID
        cursor.execute(
            "SELECT DISTINCT roomid FROM messages WHERE roomid IS NOT NULL AND roomid != ''"
        )
        room_ids = [r["roomid"] for r in cursor.fetchall()]

        # 所有用户 ID
        cursor.execute(
            "SELECT DISTINCT sender FROM messages WHERE sender IS NOT NULL AND sender != ''"
        )
        user_ids = [r["sender"] for r in cursor.fetchall()]
    finally:
        conn.close()

    synced_rooms = 0
    synced_users = 0

    # 逐个解析群名（contacts 内部会自动写缓存）
    for rid in room_ids:
        try:
            name = _contacts.get_room_name(rid)
            if name and name != rid:
                synced_rooms += 1
        except Exception:
            pass

    # 逐个解析用户名
    for uid in user_ids:
        try:
            name = _contacts.get_name(uid)
            if name and name != uid:
                synced_users += 1
        except Exception:
            pass

    total = len(room_ids) + len(user_ids)
    synced = synced_rooms + synced_users
    return jsonify({
        "success": True,
        "message": f"同步完成：共 {total} 项，成功解析 {synced} 项（群 {synced_rooms}/{len(room_ids)}，联系人 {synced_users}/{len(user_ids)}）",
        "synced_rooms": synced_rooms,
        "synced_users": synced_users,
        "total_rooms": len(room_ids),
        "total_users": len(user_ids),
    })


# 配置状态缓存（60秒）
_config_status_cache = {"data": None, "time": 0}


@insurance_bp.route("/api/insurance/config/status", methods=["GET"])
def config_status():
    """检查配置状态（OCR、钉钉等关键配置是否就绪），60秒缓存"""
    import time as _time

    if _config_status_cache["data"] and (_time.time() - _config_status_cache["time"] < 60):
        return jsonify({"code": 0, "data": _config_status_cache["data"]})

    try:
        # 只查 config_key 是否存在 + value 长度，完全不读 LONGTEXT 内容
        conn = _db.pool.connection()
        try:
            cursor = conn.cursor(pymysql.cursors.DictCursor)
            cursor.execute(
                "SELECT config_key, LENGTH(config_value) as val_len "
                "FROM insurance_config "
                "WHERE config_key IN ("
                "  'baidu_ocr_app_id','baidu_ocr_api_key','baidu_ocr_secret_key','baidu_ocr_enabled',"
                "  'dingtalk_app_key','dingtalk_app_secret','dingtalk_operator_id'"
                ")"
            )
            existing = {row["config_key"]: row["val_len"] for row in cursor.fetchall()}
        finally:
            conn.close()

        baidu_ok = all(existing.get(k, 0) for k in ["baidu_ocr_app_id", "baidu_ocr_api_key", "baidu_ocr_secret_key"])
        dingtalk_ok = all(existing.get(k, 0) for k in ["dingtalk_app_key", "dingtalk_app_secret", "dingtalk_operator_id"])
        baidu_enabled = "baidu_ocr_enabled" not in existing or existing.get("baidu_ocr_enabled", 0) > 0

        # 钉钉目标数量：用监控配置表计数，不读大 JSON
        targets_count = 0
        try:
            from .monitor_config_db import list_monitor_configs
            monitors = list_monitor_configs(_db)
            targets_count = len([m for m in monitors if m.get("dingtalk_base_id")])
        except Exception:
            pass

        status = {
            "baidu_ocr": {
                "configured": baidu_ok,
                "enabled": baidu_enabled,
            },
            "dingtalk": {
                "configured": dingtalk_ok,
                "targets_count": targets_count,
            },
            "handler_ready": _handler is not None,
        }

        _config_status_cache["data"] = status
        _config_status_cache["time"] = _time.time()

        return jsonify({"code": 0, "data": status})
    except Exception as e:
        logger.exception("获取配置状态失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


# ============================================================
# 字段映射
# ============================================================

@insurance_bp.route("/api/insurance/mapping/config", methods=["GET"])
def get_mapping_config():
    """获取字段映射配置（包含默认映射和各保司映射）"""
    try:
        config = load_mapping_config()
        # 同时返回可用 OCR 字段列表，方便前端构建下拉
        config["available_ocr_fields"] = get_available_ocr_fields()
        return jsonify({"code": 0, "data": config})
    except Exception as e:
        logger.exception("获取字段映射配置失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/mapping/save", methods=["POST"])
def save_mapping():
    """保存字段映射配置"""
    try:
        body = request.get_json(force=True) or {}
        if not body:
            return jsonify({"code": 400, "msg": "请求体不能为空"}), 400
        save_mapping_config(body)
        return jsonify({"code": 0, "msg": "映射配置已保存"})
    except Exception as e:
        logger.exception("保存字段映射配置失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/mapping/company/<company>", methods=["DELETE"])
def delete_company_mapping(company):
    """删除指定保司的字段映射配置"""
    try:
        config = load_mapping_config()
        company_mappings = config.get("company_mappings", {})
        if company not in company_mappings:
            return jsonify({"code": 404, "msg": f"保司 {company} 的映射配置不存在"}), 404
        del company_mappings[company]
        config["company_mappings"] = company_mappings
        save_mapping_config(config)
        return jsonify({"code": 0, "msg": f"已删除保司 {company} 的映射配置"})
    except Exception as e:
        logger.exception("删除保司映射配置失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/mapping/preview", methods=["POST"])
def preview_mapping():
    """
    预览字段映射效果。
    请求体: {fields: {ocr字段名: 值, ...}, company_short: 保司简称}
    """
    try:
        body = request.get_json(force=True) or {}
        fields = body.get("fields", {})
        company_short = body.get("company_short", "")

        if not fields:
            return jsonify({"code": 400, "msg": "缺少 fields 参数"}), 400

        mapped = apply_mapping(fields, company_short)
        return jsonify({"code": 0, "data": {"mapped_fields": mapped, "output_columns": OUTPUT_COLUMNS}})
    except Exception as e:
        logger.exception("预览字段映射失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


# ============================================================
# 钉钉目标管理
# ============================================================

def _get_dingtalk_targets() -> list:
    """从数据库配置中读取钉钉目标列表"""
    raw = get_insurance_config(_db, "dingtalk_targets", [])
    if isinstance(raw, str):
        try:
            raw = json.loads(raw)
        except Exception:
            raw = []
    return raw if isinstance(raw, list) else []


def _save_dingtalk_targets(targets: list):
    """将钉钉目标列表保存到数据库配置"""
    set_insurance_config(_db, "dingtalk_targets", targets)


@insurance_bp.route("/api/insurance/dingtalk/targets", methods=["GET"])
def list_dingtalk_targets():
    """获取钉钉同步目标列表"""
    try:
        targets = _get_dingtalk_targets()
        return jsonify({"code": 0, "data": targets})
    except Exception as e:
        logger.exception("获取钉钉目标列表失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/dingtalk/targets", methods=["POST"])
def add_dingtalk_target():
    """
    添加钉钉同步目标。
    请求体: {doc_url: 钉钉文档URL, name: 目标名称}
    自动解析 base_id 和 sheet_id。
    """
    try:
        body = request.get_json(force=True) or {}
        doc_url = body.get("doc_url", "").strip()
        name = body.get("name", "").strip()

        if not doc_url:
            return jsonify({"code": 400, "msg": "缺少 doc_url 参数"}), 400

        # 解析钉钉文档 URL 获取 base_id 和 sheet_id
        base_id, sheet_id = parse_dingtalk_doc_url(doc_url)
        if not base_id:
            return jsonify({"code": 400, "msg": "无法从 URL 中解析 base_id，请检查链接格式"}), 400

        # 生成稳定的目标 ID（基于 URL hash）
        target_id = hashlib.md5(doc_url.encode("utf-8")).hexdigest()[:12]

        targets = _get_dingtalk_targets()
        # 检查是否已存在
        for t in targets:
            if t.get("id") == target_id:
                return jsonify({"code": 409, "msg": "该目标已存在"}), 409

        new_target = {
            "id": target_id,
            "name": name or f"目标-{target_id}",
            "doc_url": doc_url,
            "base_id": base_id,
            "sheet_id": sheet_id,
        }
        targets.append(new_target)
        _save_dingtalk_targets(targets)

        return jsonify({"code": 0, "data": new_target, "msg": "目标已添加"})
    except Exception as e:
        logger.exception("添加钉钉目标失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/dingtalk/targets/<target_id>", methods=["PUT"])
def update_dingtalk_target(target_id):
    """更新指定钉钉同步目标"""
    try:
        body = request.get_json(force=True) or {}
        targets = _get_dingtalk_targets()

        found = False
        for t in targets:
            if t.get("id") == target_id:
                # 允许更新名称和 doc_url，更新 URL 时重新解析 base_id/sheet_id
                if "name" in body:
                    t["name"] = body["name"]
                if "doc_url" in body:
                    new_url = body["doc_url"].strip()
                    t["doc_url"] = new_url
                    base_id, sheet_id = parse_dingtalk_doc_url(new_url)
                    t["base_id"] = base_id
                    t["sheet_id"] = sheet_id
                found = True
                break

        if not found:
            return jsonify({"code": 404, "msg": f"目标 {target_id} 不存在"}), 404

        _save_dingtalk_targets(targets)
        return jsonify({"code": 0, "msg": "目标已更新"})
    except Exception as e:
        logger.exception("更新钉钉目标 %s 失败", target_id)
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/dingtalk/targets/<target_id>", methods=["DELETE"])
def delete_dingtalk_target(target_id):
    """删除指定钉钉同步目标"""
    try:
        targets = _get_dingtalk_targets()
        new_targets = [t for t in targets if t.get("id") != target_id]

        if len(new_targets) == len(targets):
            return jsonify({"code": 404, "msg": f"目标 {target_id} 不存在"}), 404

        _save_dingtalk_targets(new_targets)
        return jsonify({"code": 0, "msg": "目标已删除"})
    except Exception as e:
        logger.exception("删除钉钉目标 %s 失败", target_id)
        return jsonify({"code": 500, "msg": str(e)}), 500


# ============================================================
# 同步 & 导出
# ============================================================

@insurance_bp.route("/api/insurance/sync/check-duplicates", methods=["POST"])
def check_duplicates():
    """
    检查保单号是否在钉钉目标表中已存在（去重）。
    请求体: {target_id: str, invoices: [{fields: {...}}, ...], key_field: str}
    """
    try:
        body = request.get_json(force=True) or {}
        target_id = body.get("target_id", "")
        invoices = body.get("invoices", [])
        key_field = body.get("key_field", "保单号")

        if not target_id:
            return jsonify({"code": 400, "msg": "缺少 target_id 参数"}), 400

        # 查找目标配置
        targets = _get_dingtalk_targets()
        target = next((t for t in targets if t.get("id") == target_id), None)
        if not target:
            return jsonify({"code": 404, "msg": f"目标 {target_id} 不存在"}), 404

        # 读取钉钉应用配置
        config = get_all_insurance_config(_db)
        app_key = config.get("dingtalk_app_key", "")
        app_secret = config.get("dingtalk_app_secret", "")
        operator_id = config.get("dingtalk_operator_id", "")

        if not all([app_key, app_secret, operator_id]):
            return jsonify({"code": 400, "msg": "钉钉应用未配置，请先完成钉钉配置"}), 400

        service = DingTalkTableService(
            app_key=app_key,
            app_secret=app_secret,
            base_id=target["base_id"],
            sheet_id=target.get("sheet_id", ""),
            operator_id=operator_id,
        )

        duplicates = service.find_duplicates(invoices, key_field=key_field)
        return jsonify({"code": 0, "data": {"duplicates": duplicates, "count": len(duplicates)}})
    except Exception as e:
        logger.exception("去重检查失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/sync/dingtalk", methods=["POST"])
def batch_sync_dingtalk():
    """
    手动批量同步保单记录到钉钉。
    请求体: {target_id: str, record_ids: [int, ...], field_names: [str, ...]}
    """
    try:
        body = request.get_json(force=True) or {}
        target_id = body.get("target_id", "")
        record_ids = body.get("record_ids", [])
        field_names = body.get("field_names")
        if not field_names:
            user = g.current_user
            export_col_config = get_column_config(_db, "export_columns", user["user_id"], user["role"], user.get("parent_id"))
            visible_cols = sorted([c for c in export_col_config.get("columns", []) if c.get("visible", True)], key=lambda c: c.get("order", 0))
            field_names = [c["key"] for c in visible_cols] or list(OUTPUT_COLUMNS)

        if not target_id:
            return jsonify({"code": 400, "msg": "缺少 target_id 参数"}), 400
        if not record_ids:
            return jsonify({"code": 400, "msg": "缺少 record_ids 参数"}), 400

        # 查找目标配置
        targets = _get_dingtalk_targets()
        target = next((t for t in targets if t.get("id") == target_id), None)
        if not target:
            return jsonify({"code": 404, "msg": f"目标 {target_id} 不存在"}), 404

        # 读取钉钉应用配置
        config = get_all_insurance_config(_db)
        app_key = config.get("dingtalk_app_key", "")
        app_secret = config.get("dingtalk_app_secret", "")
        operator_id = config.get("dingtalk_operator_id", "")

        if not all([app_key, app_secret, operator_id]):
            return jsonify({"code": 400, "msg": "钉钉应用未配置，请先完成钉钉配置"}), 400

        service = DingTalkTableService(
            app_key=app_key,
            app_secret=app_secret,
            base_id=target["base_id"],
            sheet_id=target.get("sheet_id", ""),
            operator_id=operator_id,
        )

        # 构建待同步的发票数据
        invoices = []
        success_ids = []
        failed_ids = []
        for rid in record_ids:
            record = get_insurance_record(_db, rid)
            if not record:
                failed_ids.append({"id": rid, "reason": "记录不存在"})
                continue

            mapped = record.get("mapped_fields") or {}
            if isinstance(mapped, str):
                try:
                    mapped = json.loads(mapped)
                except Exception:
                    mapped = {}

            if mapped:
                invoices.append({"fields": mapped, "_record_id": rid})
                success_ids.append(rid)
            else:
                failed_ids.append({"id": rid, "reason": "无映射字段"})

        if not invoices:
            return jsonify({"code": 400, "msg": "没有可同步的记录"}), 400

        # 执行批量插入
        result = service.append_invoices(field_names, invoices)

        # 更新同步状态
        for rid in success_ids:
            try:
                update_insurance_record(_db, rid, {
                    "dingtalk_synced": 1,
                    "dingtalk_target_id": target_id,
                })
            except Exception as upd_err:
                logger.warning("更新记录 %d 同步状态失败: %s", rid, upd_err)

        inserted = result.get("inserted_count", 0)
        updated = result.get("updated_count", 0)
        msg_parts = []
        if inserted:
            msg_parts.append(f"新增 {inserted} 条")
        if updated:
            msg_parts.append(f"更新 {updated} 条")
        msg = "、".join(msg_parts) if msg_parts else f"同步 {len(success_ids)} 条记录"

        return jsonify({
            "code": 0,
            "data": {
                "synced": len(success_ids),
                "failed": failed_ids,
                "detail": result,
            },
            "msg": msg,
        })
    except Exception as e:
        logger.exception("批量同步钉钉失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


# ------------------------------------------------------------------ #
# Excel 导出：活公式 / 数字单元格 辅助
# ------------------------------------------------------------------ #

def _formula_to_excel(formula_display, col_letter_map, row_num):
    """
    将「按字段名引用」的公式转换为 Excel 活公式（按单元格引用）。

    formula_display: 如 "保费 × 应付费率"
    col_letter_map:  {字段key: 列字母}，如 {"保费": "K", "应付费率": "M"}
    row_num:         当前数据所在的 Excel 行号

    返回如 "=K2*M2"。若公式引用了「未导出的列」（替换后仍残留中文），返回 None，
    由调用方回退为静态值。
    """
    expr = str(formula_display).replace("×", "*").replace("÷", "/")
    # 按字段名长度降序替换，避免短字段名替换掉长字段名的一部分
    for key in sorted(col_letter_map.keys(), key=len, reverse=True):
        if key and key in expr:
            expr = expr.replace(key, f"{col_letter_map[key]}{row_num}")
    expr = expr.strip()
    if not expr:
        return None
    # 仍含中文 → 引用了未导出列，放弃（回退静态值）
    if re.search(r"[一-鿿]", expr):
        return None
    # 仅允许：字母(单元格列)、数字、运算符、括号、小数点、空格
    if not re.match(r"^[A-Za-z0-9\s\+\-\*\/\.\(\)]+$", expr):
        return None
    return "=" + expr


def _percent_text_to_decimal(s):
    """ "20%" / "20.0%" → 0.2（小数）。非纯百分比数字返回 None。"""
    if not isinstance(s, str):
        return None
    t = s.strip()
    if not t.endswith("%"):
        return None
    body = t[:-1].strip().replace(",", "")
    if re.fullmatch(r"-?\d+(\.\d+)?", body):
        return float(body) / 100
    return None


def _pure_number(s):
    """纯数字字符串 → int/float，否则 None（用于把公式引用的金额列转为数字单元格）。"""
    if not isinstance(s, str):
        return None
    t = s.strip().replace(",", "")
    if re.fullmatch(r"-?\d+(\.\d+)?", t):
        f = float(t)
        return int(f) if f == int(f) else f
    return None


_REPORT_SIGN_SALT = "wxbot-report-2026-a7f3k9x2q5"

def _report_sign(eid, ds, de):
    """报告链接签名：防止随意改 enterprise_id 枚举查看别家报告。"""
    import hashlib
    raw = f"{eid}|{ds}|{de}|{_REPORT_SIGN_SALT}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]


@insurance_bp.route("/api/insurance/enterprise-report", methods=["GET"])
def export_enterprise_report():
    """导出企业经营报告 Excel（仅超管）。
    含：识别量汇总、每日趋势(折线图)、保司/险种分布、上传人排名。
    参数：enterprise_id, date_start(YYYY-MM-DD), date_end, enterprise_name(可选)。"""
    from auth.db import ROLE_SUPER_ADMIN
    try:
        enterprise_id = request.args.get("enterprise_id", "").strip()
        date_start = request.args.get("date_start", "").strip()
        date_end = request.args.get("date_end", "").strip()
        ent_name = request.args.get("enterprise_name", "").strip()
        if not enterprise_id or not date_start or not date_end:
            return jsonify({"code": 400, "msg": "缺少参数"}), 400

        from insurance.db import get_enterprise_report_data

        # 网页报告(format=json)：推送给客户查看，凭链接签名访问、不需登录
        if request.args.get("format") == "json":
            if request.args.get("sign", "") != _report_sign(enterprise_id, date_start, date_end):
                return jsonify({"code": 403, "msg": "链接无效或已失效"}), 403
            data = get_enterprise_report_data(_db, int(enterprise_id), date_start, date_end)
            _nm = ent_name
            if not _nm:
                try:
                    from auth.db import get_enterprise_by_id
                    _ent = get_enterprise_by_id(_db, int(enterprise_id))
                    _nm = (_ent.get("name") if _ent else "") or f"企业{enterprise_id}"
                except Exception:
                    _nm = f"企业{enterprise_id}"
            return jsonify({"code": 0, "data": {
                "enterprise_name": _nm,
                "date_start": date_start, "date_end": date_end,
                "summary": data.get("summary", {}),
                "daily": data.get("daily", []),
                "failed_records": data.get("failed_records", []),
            }})

        # 以下：生成报告链接 / 导出Excel —— 仅超管（before_request 已验 token 设置 g.current_user）
        if g.current_user.get("role") != ROLE_SUPER_ADMIN:
            return jsonify({"code": 403, "msg": "无权限"}), 403
        # format=link：返回该报告的签名，供前端拼成带签名的分享链接
        if request.args.get("format") == "link":
            return jsonify({"code": 0, "data": {"sign": _report_sign(enterprise_id, date_start, date_end)}})

        import openpyxl
        from openpyxl.chart import LineChart, Reference
        from openpyxl.styles import Font, PatternFill, Alignment
        data = get_enterprise_report_data(_db, int(enterprise_id), date_start, date_end)

        if not ent_name:
            try:
                from auth.db import get_enterprise_by_id
                ent = get_enterprise_by_id(_db, int(enterprise_id))
                if ent:
                    ent_name = ent.get("name") or ent.get("phone") or f"企业{enterprise_id}"
            except Exception:
                ent_name = f"企业{enterprise_id}"

        wb = openpyxl.Workbook()
        title_font = Font(bold=True, size=14)
        h_font = Font(bold=True, size=11, color="FFFFFF")
        h_fill = PatternFill("solid", fgColor="ED7D31")
        sec_font = Font(bold=True, size=12, color="C55A11")
        center = Alignment(horizontal="center")

        ws = wb.active
        ws.title = "企业报告"
        ws["A1"] = f"{ent_name} 经营报告"
        ws["A1"].font = title_font
        ws["A2"] = f"报告周期：{date_start} ~ {date_end}"
        from datetime import datetime as _dtn
        ws["A3"] = f"生成时间：{_dtn.now().strftime('%Y-%m-%d %H:%M')}"
        row = [5]  # 用列表以便闭包内修改

        def section(t):
            ws.cell(row=row[0], column=1, value=t).font = sec_font
            row[0] += 1

        def table(headers, rows_data):
            for j, h in enumerate(headers, 1):
                c = ws.cell(row=row[0], column=j, value=h)
                c.font = h_font; c.fill = h_fill; c.alignment = center
            row[0] += 1
            for rd in rows_data:
                for j, v in enumerate(rd, 1):
                    ws.cell(row=row[0], column=j, value=v)
                row[0] += 1
            row[0] += 1

        s = data["summary"]
        section("① 识别量汇总")
        table(["指标", "数量"], [
            ["总识别量", s["total"]],
            ["成功(保单)", s["success"]],
            ["需人工补充", s["abnormal"]],
            ["非保单", s["nonpolicy"]],
            ["失败", s["failed"]],
            ["成功率", f"{s['success_rate']}%"],
        ])
        section("② 保司分布（成功保单）")
        table(["保险公司", "保单数", "保费合计(元)"],
              [[c["name"], c["count"], c["premium"]] for c in data["companies"]] or [["(无数据)", "", ""]])
        section("③ 险种分布（成功保单）")
        table(["险种", "保单数"],
              [[p["name"], p["count"]] for p in data["policy_types"]] or [["(无数据)", ""]])
        section("④ 上传人排名")
        table(["上传人", "识别量", "成功量"],
              [[u["sender_name"], u["total"], u["success"]] for u in data["uploaders"]] or [["(无数据)", "", ""]])
        for col, w in {"A": 22, "B": 14, "C": 16}.items():
            ws.column_dimensions[col].width = w

        # 每日趋势 + 折线图
        ws2 = wb.create_sheet("每日趋势")
        ws2.append(["日期", "总量", "成功", "失败"])
        for c in ws2[1]:
            c.font = h_font; c.fill = h_fill; c.alignment = center
        for d in data["daily"]:
            ws2.append([d["day"], d["total"], d["success"], d["failed"]])
        ws2.column_dimensions["A"].width = 14
        n = len(data["daily"])
        if n >= 1:
            chart = LineChart()
            chart.title = f"{ent_name} 识别趋势"
            chart.style = 12
            chart.y_axis.title = "数量"
            chart.x_axis.title = "日期"
            chart.height = 9; chart.width = 26
            cats = Reference(ws2, min_col=1, min_row=2, max_row=n + 1)
            dref = Reference(ws2, min_col=2, max_col=4, min_row=1, max_row=n + 1)
            chart.add_data(dref, titles_from_data=True)
            chart.set_categories(cats)
            ws2.add_chart(chart, "F2")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        fname = f"{ent_name}_经营报告_{date_start}_{date_end}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=fname,
        )
    except Exception as e:
        logger.exception("导出企业报告失败")
        return jsonify({"code": 500, "msg": f"导出失败: {e}"}), 500


def _sanitize_filename_part(s: str, max_len: int = 40) -> str:
    """去除文件名非法字符，限制长度，避免企业名/标签里混入特殊符号导致下载失败。"""
    s = re.sub(r'[\\/:*?"<>|]', "_", (s or "").strip())
    return s[:max_len] or "未命名"


def _resolve_export_enterprise_name(body: dict) -> str:
    """
    解析导出文件名用的企业名称：
    - 超管：按其筛选的 enterprise_id 取企业名，未筛选（数据混合多家企业）则用"全部企业"
    - 其他角色（企业/员工/销售）：固定为自己所属企业，不信任前端传参
    """
    role = g.current_user.get("role")
    if role == ROLE_SUPER_ADMIN:
        eid = body.get("enterprise_id")
        if not eid:
            return "全部企业"
        try:
            from auth.db import get_enterprise_by_id
            ent = get_enterprise_by_id(_db, int(eid))
            return ent.get("name") or ent.get("phone") or f"企业{eid}" if ent else f"企业{eid}"
        except Exception:
            return "全部企业"
    from auth.db import get_user_by_id, get_enterprise_by_id
    user = get_user_by_id(_db, g.current_user["user_id"])
    eid = user.get("parent_id") if user else None
    if not eid:
        return "企业"
    ent = get_enterprise_by_id(_db, eid)
    return (ent.get("name") if ent else "") or "企业"


def _build_export_filename(body: dict, export_label: str) -> str:
    """按「企业名称_导出类型_时间周期.xlsx」规则生成导出文件名。"""
    from datetime import datetime
    ent_name = _sanitize_filename_part(_resolve_export_enterprise_name(body))
    label = _sanitize_filename_part(export_label)
    date_start = (body.get("date_start") or "").strip()[:10]
    date_end = (body.get("date_end") or "").strip()[:10]
    if date_start or date_end:
        period = f"{date_start or date_end}_{date_end or date_start}"
    else:
        period = datetime.now().strftime("%Y%m%d_%H%M")
    return f"{ent_name}_{label}_{period}.xlsx"


@insurance_bp.route("/api/insurance/export/excel", methods=["POST"])
def export_excel():
    """
    导出保单数据为 Excel 文件。
    请求体: {invoices: [...], field_names: [...], export_type: "success"|"failed", sheet_name: str}
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"code": 500, "msg": "缺少 openpyxl 依赖，请安装: pip install openpyxl"}), 500

    try:
        body = request.get_json(force=True) or {}
        invoices = body.get("invoices", [])

        # 导出配置（列布局/公式/固定值）应解析到的账号：默认是当前登录用户；
        # 超管选了企业筛选（body.enterprise_id）时改用该企业自己的管理员账号，
        # 否则不管是列布局兜底还是费率公式/固定值都会错用超管自己账号的配置，
        # 跟被导出企业的真实数据/模板对不上（"数据对不上"问题的根因）。
        user = g.current_user
        config_user_id, config_role, config_parent_id = user["user_id"], user["role"], user.get("parent_id")
        if user["role"] == ROLE_SUPER_ADMIN and body.get("enterprise_id"):
            try:
                _eid = int(body["enterprise_id"])
            except (TypeError, ValueError):
                _eid = None
            if _eid:
                from auth.db import get_enterprise_admin_user
                ent_admin = get_enterprise_admin_user(_db, _eid)
                if ent_admin:
                    config_user_id, config_role, config_parent_id = ent_admin["id"], ent_admin["role"], ent_admin["parent_id"]

        # 优先使用请求中传入的 field_names（兼容旧前端），否则读取用户的导出列配置
        field_names = body.get("field_names")
        field_display_names = None
        if not field_names:
            export_col_config = get_column_config(_db, "export_columns", config_user_id, config_role, config_parent_id)
            export_columns = export_col_config.get("columns", [])
            # 按 order 排序，只取 visible 的列
            visible_cols = sorted([c for c in export_columns if c.get("visible", True)], key=lambda c: c.get("order", 0))
            merge_enabled = body.get("merge_by_plate", False)
            if merge_enabled:
                # 合并模式：去掉基础差异字段和险种（由前缀列替代，用户已在配置中勾选）
                merge_hidden = set(MERGE_SPLIT_FIELDS) | {"险种"}
                visible_cols = [c for c in visible_cols if c["key"] not in merge_hidden]
            field_names = [c["key"] for c in visible_cols]
            field_display_names = {c["key"]: c.get("display_name", c["key"]) for c in visible_cols}
        if not field_names:
            field_names = list(OUTPUT_COLUMNS)
        export_type = body.get("export_type", "success")  # "success" 或 "failed"
        sheet_name = body.get("sheet_name", "保单数据")
        export_label = (body.get("export_label") or sheet_name or "保单导出").strip()

        if not invoices:
            return jsonify({"code": 400, "msg": "没有可导出的数据"}), 400

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel sheet 名最长 31 字符

        if export_type == "failed":
            # 失败类型：文件名 + 失败原因 两列
            headers = ["文件名", "失败原因"]
            ws.append(headers)
            for inv in invoices:
                fields = inv.get("fields", inv) if isinstance(inv, dict) else {}
                ws.append([
                    inv.get("file_name", inv.get("filename", fields.get("文件名", ""))),
                    inv.get("error", inv.get("error_message", fields.get("失败原因", ""))),
                ])
        elif export_type == "nonpolicy":
            # 非保单类型：文件名 + 文档类型
            headers = ["文件名", "文档类型"]
            ws.append(headers)
            for inv in invoices:
                ws.append([
                    inv.get("file_name", ""),
                    inv.get("doc_category", "非保单"),
                ])
        else:
            # 成功类型：完全按用户勾选的列导出
            # 获取用户字段配置，导出时应用简称/日期格式/公式（config_user_id 等已在函数开头解析）
            user_config = get_effective_config(_db, config_user_id, config_role, config_parent_id)
            # 注入该企业（而非超管本人）的自定义字段固定值（独立于模板）
            export_fixed_vals = get_user_fixed_values(_db, config_user_id)
            if export_fixed_vals:
                if user_config is None:
                    user_config = {}
                user_config["fixed_values"] = export_fixed_vals
            has_config = user_config and any(user_config.get(k) for k in user_config)

            # ---- 活公式导出准备 ----
            # 字段key → 列字母（用于把公式中的字段名替换为单元格引用）
            col_letter_map = {f: get_column_letter(i + 1) for i, f in enumerate(field_names)}
            # 公式引用到的「金额列」集合：导出时需写成数字单元格，公式才能计算
            formula_referenced_cols = set()
            for _item in (user_config or {}).get("fee_formula", []):
                _val = _item.get("value", "")
                _disp = _val.get("display", "") if isinstance(_val, dict) else str(_val)
                _norm = _disp.replace("×", "*").replace("÷", "/")
                for _c in field_names:
                    if _c and _c in _norm:
                        formula_referenced_cols.add(_c)

            def _postprocess_row(row_idx, src_fields):
                """把刚写入的一行做三类单元格修正：公式列→活公式；费率列→小数+百分比格式；
                公式引用的金额列→数字。src_fields 为该行的字段字典（用于按公司/险种匹配公式）。"""
                if not has_config:
                    return
                # 该行匹配到的「计算公式」：目标字段 → 公式表达式
                formula_map = {tf: f for tf, f, is_comp in match_fee_formulas(user_config, src_fields) if is_comp}
                for cidx, col in enumerate(field_names, start=1):
                    cell = ws.cell(row=row_idx, column=cidx)
                    cur = cell.value
                    # 1) 公式目标列 → Excel 活公式（失败则回退静态值）
                    if col in formula_map:
                        excel_f = _formula_to_excel(formula_map[col], col_letter_map, row_idx)
                        if excel_f:
                            # 用静态计算值做 IFERROR 兜底：公式引用到空单元格（如未填的"补点"）时，
                            # Excel 会 #VALUE! 并级联污染"保险公司合计/公司毛利"，兜底后回退到算好的值
                            _sv = _pure_number(src_fields.get(col, ""))
                            cell.value = "=IFERROR(" + excel_f[1:] + "," + (str(_sv) if _sv is not None else "0") + ")"
                            cell.number_format = "0.00"
                            continue
                    # 2) 费率列（值形如 "20%"）→ 转小数 + 百分比格式，使其能参与公式
                    dec = _percent_text_to_decimal(cur)
                    if dec is not None:
                        cell.value = dec
                        cell.number_format = "0.0%"
                        continue
                    # 3) 公式引用到的金额列 → 转数字
                    if col in formula_referenced_cols:
                        num = _pure_number(cur)
                        if num is not None:
                            cell.value = num

            # 前端标记数据是否已经过 apply_user_config 处理（识别记录导出时为 True）
            # 避免重复应用险种简称等别名导致值被二次转换
            skip_config = body.get("config_applied", False)

            # 检查是否启用按车牌合并
            merge_enabled = body.get("merge_by_plate", False)

            # 交强到期时间：用「原始」终保日期（apply_user_config 之前）构建 车牌→交强险终保日期
            # 映射，合并/不合并两种导出模式共用。注入后再按交强到期时间自身配置的格式格式化，
            # 避免读取已格式化的终保日期导致年份丢失。
            need_compulsory_end = "交强到期时间" in field_names
            plate_compulsory_end = {}
            export_compulsory_fmt = ""
            export_compulsory_no_pad = False
            if need_compulsory_end:
                # 交强到期时间统一以「数据库原始数据」(insurance_policy_fields.end_date) 为准。
                # 导出的 invoices 来自前端请求体，config_applied=True 时其中的车牌/终保日期可能
                # 已被规则转换（车牌加连字符、日期格式化丢年份），不可作为匹配依据。
                # 故仅取前端车牌做归一化 key，终保日期值一律走数据库原始查询。
                all_plates = set()
                for inv in invoices:
                    f = inv.get("fields", {}) if isinstance(inv, dict) else {}
                    np = _norm_plate(f.get("车牌", "") or f.get("车牌号", ""))
                    if np:
                        all_plates.add(np)
                if all_plates and _db:
                    try:
                        from insurance.db import find_compulsory_end_dates
                        db_end_dates = find_compulsory_end_dates(_db, list(all_plates))
                        for _p, _v in db_end_dates.items():
                            plate_compulsory_end[_norm_plate(_p)] = _v
                    except Exception:
                        logger.debug("导出时查询交强到期时间失败")
                for _item in (user_config or {}).get("date_format", []):
                    if _item.get("key") == "交强到期时间":
                        export_compulsory_fmt = _item.get("value", "")
                    elif _item.get("key") == "__no_pad__" and _item.get("value"):
                        export_compulsory_no_pad = True

            def _inject_compulsory_end(row_fields):
                """向一行数据注入并格式化交强到期时间（手动填写过的不覆盖）。"""
                if not need_compulsory_end:
                    return
                if row_fields.get("交强到期时间"):
                    # 已有值（如手动填写，已被 apply_user_config 格式化）则不覆盖
                    return
                plate = row_fields.get("车牌", "") or row_fields.get("车牌号", "") or ""
                if not plate:
                    return
                _cv = plate_compulsory_end.get(_norm_plate(plate), "")
                if _cv and export_compulsory_fmt:
                    _cv = _format_date(str(_cv), export_compulsory_fmt, export_compulsory_no_pad)
                if _cv:
                    row_fields["交强到期时间"] = _cv

            need_start_combo = "交强与交商起保时间" in field_names
            plate_start_dates = {}
            export_start_combo_fmt = ""
            export_start_combo_no_pad = False
            if need_start_combo:
                # 同交强到期时间：一律以数据库原始 start_date 为准，不信任前端传来的字段
                all_plates2 = set()
                for inv in invoices:
                    f = inv.get("fields", {}) if isinstance(inv, dict) else {}
                    np = _norm_plate(f.get("车牌", "") or f.get("车牌号", ""))
                    if np:
                        all_plates2.add(np)
                if all_plates2 and _db:
                    try:
                        from insurance.db import find_commercial_compulsory_start_dates
                        db_start_dates = find_commercial_compulsory_start_dates(_db, list(all_plates2))
                        for _p, _v in db_start_dates.items():
                            plate_start_dates[_norm_plate(_p)] = _v
                    except Exception:
                        logger.debug("导出时查询交强与交商起保时间失败")
                for _item in (user_config or {}).get("date_format", []):
                    if _item.get("key") == "交强与交商起保时间":
                        export_start_combo_fmt = _item.get("value", "")
                    elif _item.get("key") == "__no_pad__" and _item.get("value"):
                        export_start_combo_no_pad = True

            def _inject_start_combo(row_fields):
                """向一行数据注入并格式化交强与交商起保时间（手动填写过的不覆盖）。"""
                if not need_start_combo:
                    return
                if row_fields.get("交强与交商起保时间"):
                    return
                plate = row_fields.get("车牌", "") or row_fields.get("车牌号", "") or ""
                if not plate:
                    return
                _bucket = plate_start_dates.get(_norm_plate(plate), {})
                _cv = _combine_start_dates(
                    _bucket.get("commercial"), _bucket.get("compulsory"),
                    export_start_combo_fmt, export_start_combo_no_pad)
                if _cv:
                    row_fields["交强与交商起保时间"] = _cv

            if merge_enabled:
                # ---- 按车牌合并导出 ----
                # 判断车牌是否可合并（空、新车、*-* 格式的不合并）
                def _plate_mergeable(plate: str) -> bool:
                    if not plate:
                        return False
                    p = plate.strip()
                    if not p or p == "新车":
                        return False
                    if p in ("*-*", "*—*"):
                        return False
                    return True

                # 1. 按车牌分组，不可合并的记录单独作为独立行
                #    车牌为空/新车但有车架号时，用车架号(VIN)作分组键兜底——与去重/前端合并键一致，
                #    避免"未上牌新车"的商业险+驾乘险因空车牌各自单独成行、显示成多组重复。
                plate_groups = {}
                standalone_rows = []
                for inv in invoices:
                    fields = inv.get("fields", {}) if isinstance(inv, dict) else {}
                    if has_config and not skip_config:
                        fields = apply_user_config_to_fields(user_config, fields)
                    # 记录来源 id，供 return_json 场景下前端精确匹配合并行对应的原始记录
                    # （车牌/车架号/保单号都缺失时前端无法靠字段值再算出同一个分组键）
                    fields["_src_id"] = inv.get("id") if isinstance(inv, dict) else None
                    plate = fields.get("车牌", "") or fields.get("车牌号", "") or ""
                    if _plate_mergeable(plate):
                        group_key = plate.strip()
                    else:
                        _vin = (fields.get("车架号", "") or fields.get("车架号VIN", "") or "").strip()
                        group_key = "__VIN__" + _vin.upper() if len(_vin) >= 10 else None
                    if group_key is None:
                        standalone_rows.append(fields)
                    else:
                        plate_groups.setdefault(group_key, []).append(fields)

                # 2. 合并每组
                merged_rows = []
                for plate, group in plate_groups.items():
                    merged = {}
                    # 共享字段：互相补充
                    for f in MERGE_SHARED_FIELDS:
                        for record in group:
                            val = record.get(f, "")
                            if val:
                                merged[f] = val
                                break

                    # 差异字段：按险种分配
                    for record in group:
                        policy_type = record.get("险种", "") or record.get("险种类型", "")
                        type_code, _ = get_policy_type_code(policy_type)
                        prefix = MERGE_PREFIXES.get(type_code, "非车险")
                        for f in MERGE_SPLIT_FIELDS:
                            col_name = f"{prefix}{f}"
                            val = record.get(f, "")
                            if val and not merged.get(col_name):
                                merged[col_name] = val

                    # 车船税（不拆分）
                    for record in group:
                        val = record.get("车船税", "")
                        if val:
                            merged["车船税"] = val
                            break

                    # 保留其他自定义字段（非共享非差异的字段互相补充）
                    all_known = set(MERGE_SHARED_FIELDS) | set(MERGE_SPLIT_FIELDS) | {"车船税", "险种", "险种类型", "_src_id"}
                    for record in group:
                        for k, v in record.items():
                            if k not in all_known and k not in merged and v:
                                merged[k] = v
                    # 该合并行汇总了哪些原始记录（前端按此精确匹配，不再靠字段值反推分组键）
                    merged["_src_ids"] = [r.get("_src_id") for r in group if r.get("_src_id") is not None]

                    merged_rows.append(merged)

                # 不可合并的记录也按险种拆分字段（保持列结构一致）
                for fields in standalone_rows:
                    row = {}
                    for f in MERGE_SHARED_FIELDS:
                        val = fields.get(f, "")
                        if val:
                            row[f] = val
                    policy_type = fields.get("险种", "") or fields.get("险种类型", "")
                    type_code, _ = get_policy_type_code(policy_type)
                    prefix = MERGE_PREFIXES.get(type_code, "非车险")
                    for f in MERGE_SPLIT_FIELDS:
                        col_name = f"{prefix}{f}"
                        val = fields.get(f, "")
                        if val:
                            row[col_name] = val
                    val = fields.get("车船税", "")
                    if val:
                        row["车船税"] = val
                    all_known = set(MERGE_SHARED_FIELDS) | set(MERGE_SPLIT_FIELDS) | {"车船税", "险种", "险种类型", "_src_id"}
                    for k, v in fields.items():
                        if k not in all_known and k not in row and v:
                            row[k] = v
                    _sid = fields.get("_src_id")
                    row["_src_ids"] = [_sid] if _sid is not None else []
                    merged_rows.append(row)

                # 注入交强到期时间（同车牌交强险终保日期，按配置格式格式化）
                # 注入交强与交商起保时间（同车牌商业险/交强险起保日期，一致显示一个/不同组合显示）
                for row in merged_rows:
                    _inject_compulsory_end(row)
                    _inject_start_combo(row)

                # 3. 排序（按车牌；新车无牌按车架号兜底排序）
                merged_rows.sort(key=lambda r: r.get("车牌", "") or r.get("车架号", ""))

                # 合并行重新计算公式：合并行才同时有 商业险保费/交强险保费/非车险保费 拆分列，
                # 各险种金额、保险公司合计等汇总公式只有在合并行上才能算对。
                # 注意 apply_user_config_to_fields 返回新 dict、不改入参，必须把返回值写回。
                merged_rows = [apply_user_config_to_fields(user_config, _mrow, formula_only=True)
                               for _mrow in merged_rows]

                # 页面"按车牌合并"显示：直接返回合并行 JSON（与导出完全同一套合并逻辑），不生成 Excel
                if body.get("return_json"):
                    return jsonify({"code": 0, "data": {"rows": merged_rows}})

                # 4. 写入表头和数据（前缀列已在用户配置中按勾选状态过滤）
                if field_display_names:
                    headers = [field_display_names.get(f, f) for f in field_names]
                else:
                    headers = list(field_names)
                ws.append(headers)

                for row_data in merged_rows:
                    row = []
                    for col in field_names:
                        row.append(row_data.get(col, ""))
                    ws.append(row)
                    # 合并模式：行内已无险种，多按公司匹配公式；命中则写活公式，否则保留静态值
                    _postprocess_row(ws.max_row, row_data)
            else:
                # ---- 普通导出（不合并，直接按用户勾选的列导出） ----
                if field_display_names:
                    headers = [field_display_names.get(f, f) for f in field_names]
                else:
                    headers = list(field_names)
                ws.append(headers)

                # 排序：按车牌，相同车牌按险种（商业险→交强险→驾意险）
                _type_order = {"商业险": 0, "交强险": 1, "驾意险": 2}
                def _normal_sort_key(inv):
                    fields = inv.get("fields", {}) if isinstance(inv, dict) else {}
                    plate = fields.get("车牌号", "") or fields.get("车牌", "") or ""
                    policy_type = fields.get("险种", "") or fields.get("险种类型", "") or ""
                    return (plate, _type_order.get(policy_type, 99))
                invoices.sort(key=_normal_sort_key)
                for inv in invoices:
                    fields = inv.get("fields", {}) if isinstance(inv, dict) else {}
                    if has_config and not skip_config:
                        fields = apply_user_config_to_fields(user_config, fields)
                    # 注入交强到期时间（原始终保日期，按配置格式格式化；手动填写过的不覆盖）
                    _inject_compulsory_end(fields)
                    _inject_start_combo(fields)
                    row = []
                    for col in field_names:
                        if col == "文件名":
                            row.append(fields.get("文件名", "") or inv.get("file_name", inv.get("filename", "")))
                        else:
                            row.append(fields.get(col, ""))
                    ws.append(row)
                    # 写入活公式 / 费率列转百分比数字 / 公式引用的金额列转数字
                    _postprocess_row(ws.max_row, fields)

        # 设置所有行高为 22
        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 22

        # 指定列固定宽度（用户要求）：承保公司、被保人列固定 24.01 字符
        FIXED_COL_WIDTH = {"承保公司": 24.01, "投保公司": 24.01, "被保人": 24.01, "被保险人": 24.01}
        # 自动列宽（根据内容，中文字符按 2 倍宽度计算）
        for col_idx, _ in enumerate(ws[1], start=1):
            col_letter = get_column_letter(col_idx)
            header_val = str(ws.cell(row=1, column=col_idx).value or "")
            if header_val in FIXED_COL_WIDTH:
                ws.column_dimensions[col_letter].width = FIXED_COL_WIDTH[header_val]
                continue
            max_len = 0
            for cell in ws[col_letter]:
                try:
                    if cell.value:
                        val = str(cell.value)
                        # 中文等宽字符按 2 倍计算
                        cell_len = sum(2 if ord(c) > 127 else 1 for c in val)
                        max_len = max(max_len, cell_len)
                except Exception:
                    pass
            adjusted_width = max_len + 2
            ws.column_dimensions[col_letter].width = adjusted_width

        # 保存到内存流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        download_name = _build_export_filename(body, export_label)
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=download_name,
        )
    except Exception as e:
        logger.exception("导出 Excel 失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


# ---- 批量下载任务管理 ----
import threading
import time as _time
import uuid
import zipfile

# 内存任务池（task_id → 任务状态），自动清理已完成超过 10 分钟的任务
_download_tasks = {}
_TASK_EXPIRE_SECONDS = 600


def _cleanup_expired_tasks():
    """清理过期任务（已完成超过 10 分钟）"""
    now = _time.time()
    expired = [tid for tid, t in _download_tasks.items()
               if t.get("finished_at") and now - t["finished_at"] > _TASK_EXPIRE_SECONDS]
    for tid in expired:
        _download_tasks.pop(tid, None)


def _do_batch_download(task_id, files_to_zip):
    """后台线程：逐个从 COS 下载文件并打包 ZIP"""
    import requests as http_requests

    task = _download_tasks[task_id]
    zip_buffer = io.BytesIO()
    used_names = {}

    try:
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for i, item in enumerate(files_to_zip):
                task["current"] = i + 1
                task["current_file"] = item["filename"]
                try:
                    resp = http_requests.get(item["cos_url"], timeout=30)
                    if resp.status_code != 200:
                        task["failed"] += 1
                        continue
                    # 处理重名文件
                    name = item["filename"]
                    if name in used_names:
                        used_names[name] += 1
                        base, ext = (name.rsplit(".", 1) + [""])[:2]
                        name = f"{base}_{used_names[name]}.{ext}" if ext else f"{base}_{used_names[name]}"
                    else:
                        used_names[name] = 0
                    zf.writestr(name, resp.content)
                    task["success"] += 1
                except Exception:
                    task["failed"] += 1

        zip_buffer.seek(0)
        task["zip_data"] = zip_buffer
        task["status"] = "done"
    except Exception as e:
        task["status"] = "error"
        task["error"] = str(e)[:200]
    finally:
        task["finished_at"] = _time.time()


@insurance_bp.route("/api/insurance/batch-download/start", methods=["POST"])
def batch_download_start():
    """
    启动批量下载任务（异步）。
    请求体: {record_ids: [int, ...]}
    返回 task_id，前端轮询进度。
    """
    _cleanup_expired_tasks()

    try:
        body = request.get_json(force=True) or {}
        record_ids = body.get("record_ids", [])

        if not record_ids:
            return jsonify({"code": 400, "msg": "缺少 record_ids 参数"}), 400
        if len(record_ids) > 1000:
            return jsonify({"code": 400, "msg": "单次下载不超过 1000 条"}), 400

        # 收集文件信息
        files_to_zip = []
        for rid in record_ids:
            record = get_insurance_record(_db, rid)
            if not record:
                continue
            cos_url = record.get("cos_url", "")
            if not cos_url:
                continue
            filename = record.get("filename", f"record_{rid}.pdf")
            files_to_zip.append({"id": rid, "cos_url": cos_url, "filename": filename})

        if not files_to_zip:
            return jsonify({"code": 400, "msg": "没有可下载的文件"}), 400

        # 创建任务
        task_id = uuid.uuid4().hex[:12]
        _download_tasks[task_id] = {
            "status": "processing",
            "total": len(files_to_zip),
            "current": 0,
            "success": 0,
            "failed": 0,
            "current_file": "",
            "zip_data": None,
            "error": "",
            "finished_at": None,
        }

        # 启动后台线程
        t = threading.Thread(target=_do_batch_download, args=(task_id, files_to_zip), daemon=True)
        t.start()

        return jsonify({"code": 0, "data": {"task_id": task_id, "total": len(files_to_zip)}})
    except Exception as e:
        logger.exception("启动批量下载任务失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/batch-download/status/<task_id>", methods=["GET"])
def batch_download_status(task_id):
    """查询批量下载任务进度"""
    task = _download_tasks.get(task_id)
    if not task:
        return jsonify({"code": 404, "msg": "任务不存在或已过期"}), 404

    return jsonify({"code": 0, "data": {
        "status": task["status"],
        "total": task["total"],
        "current": task["current"],
        "success": task["success"],
        "failed": task["failed"],
        "current_file": task["current_file"],
        "error": task.get("error", ""),
    }})


@insurance_bp.route("/api/insurance/batch-download/file/<task_id>", methods=["GET"])
def batch_download_file(task_id):
    """下载已完成的 ZIP 文件"""
    task = _download_tasks.get(task_id)
    if not task:
        return jsonify({"code": 404, "msg": "任务不存在或已过期"}), 404
    if task["status"] != "done":
        return jsonify({"code": 400, "msg": "任务尚未完成"}), 400

    zip_data = task.get("zip_data")
    if not zip_data or zip_data.getbuffer().nbytes <= 22:
        return jsonify({"code": 400, "msg": "所有文件下载失败，无可下载内容"}), 400

    zip_data.seek(0)

    # 取出后清理任务数据（只允许下载一次，释放内存）
    _download_tasks.pop(task_id, None)

    return send_file(
        zip_data,
        mimetype="application/zip",
        as_attachment=True,
        download_name="保单文件批量下载.zip",
    )


@insurance_bp.route("/api/insurance/extraction-rules", methods=["GET"])
def get_extraction_rules_api():
    """获取保单字段提取规则（供前端展示）"""
    try:
        rules = get_extraction_rules()
        return jsonify({"code": 0, "data": rules})
    except Exception as e:
        logger.exception("获取字段提取规则失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


# ============================================================
# 模板列表
# ============================================================

@insurance_bp.route('/api/insurance/templates', methods=['GET'])
def list_templates():
    """获取已注册的模板列表"""
    import os
    templates_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "templates")
    registry_path = os.path.join(templates_dir, "registry.json")
    if not os.path.exists(registry_path):
        return jsonify({"success": True, "data": []})
    with open(registry_path, "r", encoding="utf-8") as f:
        registry = json.load(f)
    # registry 结构为 {template_id: {info...}, ...}，转为列表返回
    result = []
    for template_id, info in registry.items():
        result.append({
            "template_id": template_id,
            **info,
        })
    return jsonify({"success": True, "data": result})


@insurance_bp.route('/api/insurance/train_records', methods=['GET'])
def list_train_records():
    """获取训练记录列表"""
    if not _db:
        return jsonify({"success": False, "error": "数据库未初始化"})
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    conn = _db.pool.connection()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) AS cnt FROM template_train_records")
        total = cursor.fetchone()["cnt"]
        offset = (page - 1) * page_size
        cursor.execute("""
            SELECT id, company_id, policy_type, version, sample_count,
                   discovered_fields, overall_hit_rate, passed, unstable_fields,
                   registered, trigger_source, created_at
            FROM template_train_records
            ORDER BY created_at DESC
            LIMIT %s OFFSET %s
        """, (page_size, offset))
        records = cursor.fetchall()
        # 转换 datetime 为字符串
        for r in records:
            if r.get("created_at"):
                r["created_at"] = str(r["created_at"])
        return jsonify({"success": True, "data": {"records": records, "total": total}})
    finally:
        conn.close()


# ============================================================
# 监控配置管理
# ============================================================

def _add_id_fields(config):
    """将 rooms/users 对象数组转为前端期望的 room_ids/user_ids 纯ID数组"""
    rooms = config.get("rooms") or []
    users = config.get("users") or []
    # rooms 可能是 [{"id":"wr...","name":"群名"}] 或 ["wr..."]
    config["room_ids"] = [
        r["id"] if isinstance(r, dict) else r for r in rooms
    ]
    config["user_ids"] = [
        u["id"] if isinstance(u, dict) else u for u in users
    ]


@insurance_bp.route("/api/monitor-configs", methods=["GET"])
def list_monitor_configs_api():
    """获取全部监控配置"""
    try:
        configs = list_monitor_configs(_db)
        # 批量查询关联账号信息
        from auth.db import get_user_by_id as _get_user, get_enterprise_by_id as _get_ent
        user_cache = {}
        ent_cache = {}
        for c in configs:
            for ts_field in ("created_at", "updated_at"):
                if c.get(ts_field) and not isinstance(c[ts_field], str):
                    c[ts_field] = str(c[ts_field])
            # 前端期望 room_ids/user_ids（纯ID数组）
            _add_id_fields(c)
            # 填充关联账号信息
            uid = c.get("user_id")
            if uid:
                c["bind_user_id"] = uid
                if uid not in user_cache:
                    u = _get_user(_db, uid)
                    user_cache[uid] = {"name": u.get("name", ""), "phone": u.get("phone", "")} if u else {}
                c["bind_user_info"] = user_cache[uid]
            else:
                c["bind_user_id"] = None
                c["bind_user_info"] = {}
            eid = c.get("enterprise_id")
            if eid:
                if eid not in ent_cache:
                    ent = _get_ent(_db, eid)
                    ent_cache[eid] = ent.get("name", "") if ent else ""
                c["bind_enterprise_name"] = ent_cache[eid]
            else:
                c["bind_enterprise_name"] = ""
        return jsonify({"code": 0, "data": configs})
    except Exception as e:
        logger.exception("获取监控配置列表失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/monitor-configs", methods=["POST"])
def create_monitor_config_api():
    """新增监控配置"""
    try:
        body = request.get_json(force=True) or {}
        name = body.get("name", "").strip()
        if not name:
            return jsonify({"code": 400, "msg": "监控名称不能为空"}), 400

        doc_url = body.get("dingtalk_doc_url", "").strip()
        base_id, sheet_id = "", ""
        if doc_url:
            base_id, sheet_id = parse_dingtalk_doc_url(doc_url)

        # 绑定企业：优先使用前端直接传入的 bind_enterprise_id
        bind_enterprise_id = body.get("bind_enterprise_id")
        if bind_enterprise_id:
            bind_enterprise_id = int(bind_enterprise_id)

        # 绑定账号：从 bind_user_id 获取 user_id，若未单独指定企业则从账号推导
        bind_user_id = body.get("bind_user_id")
        if bind_user_id:
            from auth.db import get_user_by_id as _get_bind_user
            bind_user = _get_bind_user(_db, int(bind_user_id))
            if bind_user and not bind_enterprise_id:
                bind_enterprise_id = bind_user.get("parent_id")

        config_id = create_monitor_config(_db, {
            "name": name,
            "enabled": 1 if body.get("enabled", True) else 0,
            "rooms": body.get("room_ids", body.get("rooms", [])),
            "users": body.get("user_ids", body.get("users", [])),
            "dingtalk_doc_url": doc_url,
            "dingtalk_base_id": base_id,
            "dingtalk_sheet_id": sheet_id,
            "field_mapping": body.get("field_mapping", {}),
            "user_id": int(bind_user_id) if bind_user_id else None,
            "enterprise_id": bind_enterprise_id,
        })

        config = get_monitor_config(_db, config_id)
        _add_id_fields(config)

        # 回填历史记录（绑定了账号或企业时）
        if bind_user_id or bind_enterprise_id:
            src_rooms = body.get("room_ids", body.get("rooms", []))
            src_users = body.get("user_ids", body.get("users", []))
            # 提取纯 ID
            r_ids = [r["id"] if isinstance(r, dict) else r for r in src_rooms]
            u_ids = [u["id"] if isinstance(u, dict) else u for u in src_users]
            backfill_records_by_sources(
                _db, r_ids, u_ids,
                new_user_id=int(bind_user_id) if bind_user_id else None,
                new_enterprise_id=bind_enterprise_id,
            )

        return jsonify({"code": 0, "data": config, "msg": "监控配置已创建"})
    except Exception as e:
        logger.exception("创建监控配置失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/monitor-configs/<int:config_id>", methods=["PUT"])
def update_monitor_config_api(config_id):
    """更新监控配置"""
    try:
        body = request.get_json(force=True) or {}
        if not body:
            return jsonify({"code": 400, "msg": "请求体不能为空"}), 400

        updates = {}
        if "name" in body:
            updates["name"] = body["name"].strip()
        if "enabled" in body:
            updates["enabled"] = 1 if body["enabled"] else 0
        if "room_ids" in body or "rooms" in body:
            updates["rooms"] = body.get("room_ids", body.get("rooms"))
        if "user_ids" in body or "users" in body:
            updates["users"] = body.get("user_ids", body.get("users"))
        if "field_mapping" in body:
            updates["field_mapping"] = body["field_mapping"]
        if "dingtalk_doc_url" in body:
            doc_url = body["dingtalk_doc_url"].strip()
            updates["dingtalk_doc_url"] = doc_url
            if doc_url:
                base_id, sheet_id = parse_dingtalk_doc_url(doc_url)
                updates["dingtalk_base_id"] = base_id
                updates["dingtalk_sheet_id"] = sheet_id
            else:
                updates["dingtalk_base_id"] = ""
                updates["dingtalk_sheet_id"] = ""

        # 绑定企业
        if "bind_enterprise_id" in body:
            eid = body["bind_enterprise_id"]
            updates["enterprise_id"] = int(eid) if eid else None

        # 绑定账号
        if "bind_user_id" in body:
            bind_user_id = body["bind_user_id"]
            if bind_user_id:
                from auth.db import get_user_by_id as _get_bind_user
                bind_user = _get_bind_user(_db, int(bind_user_id))
                updates["user_id"] = int(bind_user_id)
                # 若未单独指定企业，则从账号推导
                if "bind_enterprise_id" not in body:
                    updates["enterprise_id"] = bind_user.get("parent_id") if bind_user else None
            else:
                updates["user_id"] = None
                # 若未单独指定企业，则清空
                if "bind_enterprise_id" not in body:
                    updates["enterprise_id"] = None

        ok = update_monitor_config(_db, config_id, updates)
        if not ok:
            return jsonify({"code": 404, "msg": "监控配置不存在"}), 404

        config = get_monitor_config(_db, config_id)
        _add_id_fields(config)

        # 绑定账号或企业变更时回填历史记录
        if "bind_user_id" in body or "bind_enterprise_id" in body:
            src_rooms = config.get("rooms", [])
            src_users = config.get("users", [])
            r_ids = [r["id"] if isinstance(r, dict) else r for r in src_rooms]
            u_ids = [u["id"] if isinstance(u, dict) else u for u in src_users]
            backfill_records_by_sources(
                _db, r_ids, u_ids,
                new_user_id=config.get("user_id"),
                new_enterprise_id=config.get("enterprise_id"),
            )

        return jsonify({"code": 0, "data": config, "msg": "监控配置已更新"})
    except Exception as e:
        logger.exception("更新监控配置 %d 失败", config_id)
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/monitor-configs/<int:config_id>", methods=["DELETE"])
def delete_monitor_config_api(config_id):
    """删除监控配置"""
    try:
        ok = delete_monitor_config(_db, config_id)
        if not ok:
            return jsonify({"code": 404, "msg": "监控配置不存在"}), 404
        return jsonify({"code": 0, "msg": "监控配置已删除"})
    except Exception as e:
        logger.exception("删除监控配置 %d 失败", config_id)
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/monitor-configs/<int:config_id>/toggle", methods=["PUT"])
def toggle_monitor_config_api(config_id):
    """启停监控配置"""
    try:
        config = get_monitor_config(_db, config_id)
        if not config:
            return jsonify({"code": 404, "msg": "监控配置不存在"}), 404

        new_enabled = 0 if config["enabled"] else 1
        update_monitor_config(_db, config_id, {"enabled": new_enabled})
        return jsonify({"code": 0, "data": {"enabled": bool(new_enabled)}, "msg": "已更新"})
    except Exception as e:
        logger.exception("切换监控配置 %d 状态失败", config_id)
        return jsonify({"code": 500, "msg": str(e)}), 500


# ============================================================
# 补扫历史保单
# ============================================================

@insurance_bp.route("/api/insurance/rescan", methods=["POST"])
def rescan_history():
    """
    补扫历史消息中的保单 PDF。
    根据当前监控配置，查找历史消息中符合条件但未识别的 PDF，重新入队识别。
    """
    from flask import current_app

    data = request.get_json(force=True) if request.data else {}
    lookback_days = data.get("lookback_days", 5)
    config_ids = data.get("config_ids")  # 可选，指定监控配置 ID
    room_ids = data.get("room_ids")  # 可选，指定群 ID
    force = data.get("force", False)  # 是否强制重新识别

    # 参数校验
    if not isinstance(lookback_days, int) or lookback_days < 1 or lookback_days > 365:
        return jsonify({"code": 400, "msg": "回溯天数需为 1-365 之间的整数"}), 400

    # 获取所有 fetcher 实例（主企业 + 额外企业）
    wxbot_ext = current_app.extensions.get("wxbot", {})
    fetcher = wxbot_ext.get("fetcher")
    if not fetcher:
        return jsonify({"code": 500, "msg": "fetcher 未初始化"}), 500

    try:
        # 主企业补扫
        enqueued = fetcher.rescan_missed_insurance(
            lookback_days=lookback_days,
            config_ids=config_ids,
            room_ids=room_ids,
            force=force,
        )
        # 额外企业补扫（各企业使用各自的 SDK）
        for ef in wxbot_ext.get("extra_fetchers", []):
            extra_f = ef.get("fetcher")
            if extra_f and hasattr(extra_f, "rescan_missed_insurance"):
                enqueued += extra_f.rescan_missed_insurance(
                    lookback_days=lookback_days,
                    config_ids=config_ids,
                    room_ids=room_ids,
                    force=force,
                )
        return jsonify({
            "code": 0,
            "msg": f"补扫完成，已入队 {enqueued} 条待识别",
            "data": {"enqueued": enqueued}
        })
    except Exception as e:
        logger.exception("补扫历史保单失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


# ============================================================
# 字段配置（系统设置）相关路由
# ============================================================

@insurance_bp.route("/insurance/settings")
def field_config_page():
    """系统设置页面"""
    return render_template("field_config.html")


@insurance_bp.route("/api/insurance/field-config/templates", methods=["GET"])
def get_field_config_templates():
    """获取模板列表（我的 + 企业可见；超管可查看所有）"""
    try:
        from flask import g
        user = g.current_user
        data = list_field_config_templates(_db, user["user_id"], user["role"], user.get("parent_id"))
        active = get_active_template(_db, user["user_id"])
        result = {**data, "active": active}

        # 超管额外返回系统所有模板
        if user["role"] == ROLE_SUPER_ADMIN:
            result["all"] = list_all_templates(_db)

        return jsonify({"code": 0, "data": result})
    except Exception as e:
        logger.exception("获取模板列表失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/field-config/templates/manage", methods=["PUT"])
def manage_field_config_templates():
    """批量管理模板"""
    try:
        from flask import g
        body = request.get_json(force=True) or {}
        actions = body.get("actions", [])
        user = g.current_user

        # 超管代管模式
        target_scope = body.get("target_scope")
        target_scope_id = body.get("target_scope_id")
        if user["role"] == ROLE_SUPER_ADMIN and target_scope:
            scope = target_scope
            scope_id = int(target_scope_id) if target_scope_id else None
        else:
            scope, scope_id = _get_user_scope()

        for action in actions:
            act_type = action.get("type")
            if act_type == "rename":
                rename_template(_db, scope, scope_id, action["old_name"], action["new_name"])
            elif act_type == "visibility":
                update_template_visibility(_db, scope, scope_id, action["name"], action["visible"])
            elif act_type == "delete":
                delete_template_db(_db, scope, scope_id, action["name"])

        return jsonify({"code": 0, "msg": "模板已更新"})
    except Exception as e:
        logger.exception("管理模板失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/field-config", methods=["GET"])
def get_field_config():
    """获取指定模板的全部配置"""
    try:
        from flask import g
        template_name = request.args.get("template", "默认模板")
        source = request.args.get("source", "own")
        user = g.current_user

        # 超管代管模式：通过 target_scope + target_scope_id 读取任意模板
        target_scope = request.args.get("target_scope")
        target_scope_id = request.args.get("target_scope_id")
        if user["role"] == ROLE_SUPER_ADMIN and target_scope:
            sid = int(target_scope_id) if target_scope_id else None
            config = get_template_config(_db, target_scope, sid, template_name)
        elif source == "system":
            config = get_template_config(_db, "global", None, template_name)
        elif source == "enterprise" and user.get("parent_id"):
            config = get_template_config(_db, "enterprise", user["parent_id"], template_name)
        else:
            scope, scope_id = _get_user_scope()
            config = get_template_config(_db, scope, scope_id, template_name)

        return jsonify({"code": 0, "data": config})
    except Exception as e:
        logger.exception("获取字段配置失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/field-config/save-and-apply", methods=["POST"])
def save_and_apply_field_config():
    """保存模板配置并设为启用"""
    try:
        from flask import g
        body = request.get_json(force=True) or {}
        template_name = body.get("template_name", "默认模板")
        config = body.get("config", {})
        visible = body.get("visible", False)

        user = g.current_user

        # 超管代管模式：保存到指定 scope
        target_scope = body.get("target_scope")
        target_scope_id = body.get("target_scope_id")
        if user["role"] == ROLE_SUPER_ADMIN and target_scope:
            scope = target_scope
            scope_id = int(target_scope_id) if target_scope_id else None
        else:
            scope, scope_id = _get_user_scope()

        save_full_template(_db, scope, scope_id, template_name, config, visible)

        # 非代管模式才设为自己的启用模板
        if not target_scope:
            set_active_template(_db, user["user_id"], "own", template_name)
            # 同步列配置到该模板
            for col_type in ("list_columns", "export_columns"):
                cur_cols = get_column_config(_db, col_type, user["user_id"], user["role"], user.get("parent_id"))
                if cur_cols.get("source") != "default" and cur_cols.get("columns"):
                    save_column_config(_db, col_type, scope, scope_id, cur_cols["columns"],
                                       template_name=template_name)

        return jsonify({"code": 0, "msg": "已保存并应用"})
    except Exception as e:
        logger.exception("保存配置失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/field-config/apply", methods=["POST"])
def apply_field_config_template():
    """切换启用模板（不修改内容）"""
    try:
        from flask import g
        body = request.get_json(force=True) or {}
        template_name = body.get("template_name", "默认模板")
        source = body.get("source", "own")

        set_active_template(_db, g.current_user["user_id"], source, template_name)
        return jsonify({"code": 0, "msg": "已切换模板"})
    except Exception as e:
        logger.exception("切换模板失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/column-config", methods=["GET"])
@login_required
def get_column_config_api():
    """获取列配置（list_columns 或 export_columns）"""
    try:
        config_type = request.args.get("config_type", "list_columns")
        user = g.current_user
        source = request.args.get("source", "own")

        # 超管代管模式：模拟目标用户的角色查询
        target_scope = request.args.get("target_scope")
        target_scope_id = request.args.get("target_scope_id")
        if user["role"] == ROLE_SUPER_ADMIN and target_scope:
            sid = int(target_scope_id) if target_scope_id else None
            if target_scope == "enterprise":
                result = get_column_config(_db, config_type, sid, "enterprise", sid)
            elif target_scope == "user":
                from auth.db import get_user_by_id
                target_user = get_user_by_id(_db, sid)
                result = get_column_config(_db, config_type, sid, "employee", target_user.get("parent_id") if target_user else None)
            else:
                result = get_column_config(_db, config_type, 0, "super_admin", None)
        elif source == "system":
            result = get_column_config(_db, config_type, 0, "super_admin", None)
        elif source == "enterprise" and user.get("parent_id"):
            result = get_column_config(_db, config_type, user["parent_id"], "enterprise", user["parent_id"])
        else:
            result = get_column_config(_db, config_type, user["user_id"], user["role"], user.get("parent_id"))
        return jsonify({"code": 0, "data": result})
    except Exception as e:
        logger.exception("获取列配置失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/column-config", methods=["POST"])
@login_required
def save_column_config_api():
    """保存列配置"""
    try:
        body = request.get_json(force=True) or {}
        config_type = body.get("config_type", "list_columns")
        columns = body.get("columns", [])

        if config_type not in ("list_columns", "export_columns"):
            return jsonify({"code": 400, "msg": "无效的 config_type"}), 400
        if not columns:
            return jsonify({"code": 400, "msg": "columns 不能为空"}), 400

        # 超管代管模式
        user = g.current_user
        target_scope = body.get("target_scope")
        target_scope_id = body.get("target_scope_id")
        if user["role"] == ROLE_SUPER_ADMIN and target_scope:
            scope = target_scope
            scope_id = int(target_scope_id) if target_scope_id else None
        else:
            scope, scope_id = _get_user_scope()

        template_name = body.get("template_name")
        save_column_config(_db, config_type, scope, scope_id, columns,
                           template_name=template_name)
        return jsonify({"code": 0, "msg": "列配置已保存"})
    except Exception as e:
        logger.exception("保存列配置失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/column-config", methods=["DELETE"])
@login_required
def delete_column_config_api():
    """
    重置列配置。
    reset_to=enterprise: 删除当前 scope 配置，回退到上级（企业/默认）
    reset_to=system: 删除当前 scope 配置，强制写入系统默认值
    恢复前自动将现有配置备份到新模板。
    """
    try:
        config_type = request.args.get("config_type", "list_columns")
        reset_to = request.args.get("reset_to", "enterprise")
        user = g.current_user

        # 超管代管模式
        target_scope = request.args.get("target_scope")
        target_scope_id = request.args.get("target_scope_id")
        if user["role"] == ROLE_SUPER_ADMIN and target_scope:
            scope = target_scope
            scope_id = int(target_scope_id) if target_scope_id else None
        else:
            scope, scope_id = _get_user_scope()

        # 先备份当前全部配置到新模板（列配置 + 别名/日期/公式）
        current = get_column_config(_db, config_type, user["user_id"], user["role"], user.get("parent_id"))
        if current.get("source") != "default" and current.get("columns"):
            from datetime import datetime
            now = datetime.now()
            backup_name = f"备份{now.month}-{now.day} {now.strftime('%H:%M')}"
            # 备份列配置
            save_column_config(_db, config_type, scope, scope_id, current["columns"],
                               template_name=backup_name)
            # 备份另一种列配置
            other_type = "list_columns" if config_type == "export_columns" else "export_columns"
            other_cols = get_column_config(_db, other_type, user["user_id"], user["role"], user.get("parent_id"))
            if other_cols.get("source") != "default" and other_cols.get("columns"):
                save_column_config(_db, other_type, scope, scope_id, other_cols["columns"],
                                   template_name=backup_name)
            # 备份别名/日期/公式配置
            cur_config = get_effective_config(_db, user["user_id"], user["role"], user.get("parent_id"))
            if cur_config and any(cur_config.get(k) for k in cur_config):
                save_full_template(_db, scope, scope_id, backup_name, cur_config)
            logger.info("全部配置已备份到模板「%s」", backup_name)

        # 删除当前配置
        delete_column_config(_db, config_type, scope, scope_id)

        # 恢复系统默认时，显式保存 DEFAULT_COLUMNS 覆盖上级回退
        if reset_to == "system":
            save_column_config(_db, config_type, scope, scope_id, list(DEFAULT_COLUMNS))

        return jsonify({"code": 0, "msg": "已恢复默认"})
    except Exception as e:
        logger.exception("重置列配置失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/merge-config", methods=["GET"])
def get_merge_config_api():
    """获取合并导出开关"""
    try:
        user = g.current_user
        # 超管代管模式
        target_scope = request.args.get("target_scope")
        target_scope_id = request.args.get("target_scope_id")
        if user["role"] == ROLE_SUPER_ADMIN and target_scope:
            sid = int(target_scope_id) if target_scope_id else None
            if target_scope == "enterprise":
                enabled = get_merge_by_plate(_db, sid, "enterprise", sid)
            elif target_scope == "user":
                from auth.db import get_user_by_id
                target_user = get_user_by_id(_db, sid)
                enabled = get_merge_by_plate(_db, sid, "employee", target_user.get("parent_id") if target_user else None)
            else:
                enabled = get_merge_by_plate(_db, 0, "super_admin", None)
        else:
            source = request.args.get("source", "own")
            if source == "system":
                enabled = get_merge_by_plate(_db, 0, "super_admin", None)
            elif source == "enterprise" and user.get("parent_id"):
                enabled = get_merge_by_plate(_db, user["parent_id"], "enterprise", user["parent_id"])
            else:
                enabled = get_merge_by_plate(_db, user["user_id"], user["role"], user.get("parent_id"))
        return jsonify({"code": 0, "data": {"enabled": enabled}})
    except Exception as e:
        logger.exception("获取合并配置失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/merge-config", methods=["POST"])
def save_merge_config_api():
    """保存合并导出开关"""
    try:
        body = request.get_json(force=True) or {}
        enabled = body.get("enabled", False)

        # 超管代管模式
        user = g.current_user
        target_scope = body.get("target_scope")
        target_scope_id = body.get("target_scope_id")
        if user["role"] == ROLE_SUPER_ADMIN and target_scope:
            scope = target_scope
            scope_id = int(target_scope_id) if target_scope_id else None
        else:
            scope, scope_id = _get_user_scope()

        save_merge_by_plate(_db, scope, scope_id, enabled)
        return jsonify({"code": 0, "msg": "已保存"})
    except Exception as e:
        logger.exception("保存合并配置失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/plate-format-config", methods=["GET"])
@login_required
def get_plate_format_config_api():
    """获取平安车牌格式化开关"""
    try:
        user = g.current_user
        target_scope = request.args.get("target_scope")
        target_scope_id = request.args.get("target_scope_id")
        if user["role"] == ROLE_SUPER_ADMIN and target_scope:
            sid = int(target_scope_id) if target_scope_id else None
            if target_scope == "enterprise":
                enabled = get_plate_format_pingan(_db, sid, "enterprise", sid)
            elif target_scope == "user":
                from auth.db import get_user_by_id
                target_user = get_user_by_id(_db, sid)
                enabled = get_plate_format_pingan(_db, sid, "employee", target_user.get("parent_id") if target_user else None)
            else:
                enabled = get_plate_format_pingan(_db, 0, "super_admin", None)
        else:
            source = request.args.get("source", "own")
            if source == "system":
                enabled = get_plate_format_pingan(_db, 0, "super_admin", None)
            elif source == "enterprise" and user.get("parent_id"):
                enabled = get_plate_format_pingan(_db, user["parent_id"], "enterprise", user["parent_id"])
            else:
                enabled = get_plate_format_pingan(_db, user["user_id"], user["role"], user.get("parent_id"))
        return jsonify({"code": 0, "data": {"enabled": enabled}})
    except Exception as e:
        logger.exception("获取平安车牌格式化配置失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/plate-format-config", methods=["POST"])
@login_required
def save_plate_format_config_api():
    """保存平安车牌格式化开关"""
    try:
        body = request.get_json(force=True) or {}
        enabled = body.get("enabled", False)

        user = g.current_user
        target_scope = body.get("target_scope")
        target_scope_id = body.get("target_scope_id")
        if user["role"] == ROLE_SUPER_ADMIN and target_scope:
            scope = target_scope
            scope_id = int(target_scope_id) if target_scope_id else None
        else:
            scope, scope_id = _get_user_scope()

        save_plate_format_pingan(_db, scope, scope_id, enabled)
        return jsonify({"code": 0, "msg": "已保存"})
    except Exception as e:
        logger.exception("保存平安车牌格式化配置失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/fixed-values", methods=["GET"])
@login_required
def get_fixed_values_api():
    """获取当前用户的自定义字段固定值（独立于模板）"""
    try:
        user = g.current_user
        values = get_user_fixed_values(_db, user["user_id"])
        return jsonify({"code": 0, "data": values})
    except Exception as e:
        logger.exception("获取固定值失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/fixed-values", methods=["POST"])
@login_required
def save_fixed_values_api():
    """保存当前用户的自定义字段固定值（独立于模板）"""
    try:
        user = g.current_user
        body = request.get_json(force=True) or {}
        values = body.get("values", {})
        save_user_fixed_values(_db, user["user_id"], values)
        return jsonify({"code": 0, "msg": "固定值已保存"})
    except Exception as e:
        logger.exception("保存固定值失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


@insurance_bp.route("/api/insurance/field-config/formula/validate", methods=["POST"])
def validate_formula():
    """验证公式合法性并返回示例计算结果"""
    try:
        body = request.get_json(force=True) or {}
        formula = body.get("formula", "")
        if not formula:
            return jsonify({"code": 400, "msg": "公式不能为空"}), 400

        sample = {"保费": 5000, "车船税": 300, "佣金率": 0.15, "佣金": 750,
                  "采购费率": 0.05, "公司利润": 200, "跟单利润": 100}

        import re
        expr = formula.replace("×", "*").replace("÷", "/")
        for field_name, val in sample.items():
            expr = expr.replace(field_name, str(val))

        if not re.match(r'^[\d\s\+\-\*/\.\(\)]+$', expr.strip()):
            return jsonify({"code": 400, "msg": "公式包含非法字符"}), 400

        result = eval(expr)
        return jsonify({"code": 0, "data": {"result": round(result, 2), "expression": expr}})
    except ZeroDivisionError:
        return jsonify({"code": 400, "msg": "公式存在除零错误"}), 400
    except Exception as e:
        return jsonify({"code": 400, "msg": f"公式错误：{str(e)}"}), 400


@insurance_bp.route("/api/insurance/field-config/defaults", methods=["GET"])
def get_field_config_defaults():
    """
    获取系统全局的公司/险种默认映射（全称→简称）。
    数据来源：数据库全局去重 + 代码内置补充（合并去重）。
    """
    try:
        import pymysql
        from insurance.policy_parser import COMPANY_BASES, COMPANY_SHORT_MAP, _get_policy_type_code

        # ---- 公司 ----
        # 内部简称 → 更友好的显示简称
        _COMPANY_DISPLAY = {
            "人保PICC": "人保", "人民财产": "人保", "中华联合": "中华",
            "前海联合": "前海", "安盛天平": "安盛", "泰康在线": "泰康",
            "京东安联": "京东安联", "国寿财产": "人寿", "大家财险": "大家",
        }

        # 1) 从数据库全局查去重的公司（不区分企业/账号）
        company_defaults = []
        seen_companies = set()
        conn = _db.pool.connection()
        try:
            cursor = conn.cursor(pymysql.cursors.DictCursor)
            # 有简称的排前面，保证优先取到非空简称
            cursor.execute(
                "SELECT DISTINCT company, company_short FROM insurance_policy_fields "
                "WHERE company IS NOT NULL AND company != '' "
                "ORDER BY (company_short IS NULL OR company_short = ''), company_short"
            )
            for r in cursor.fetchall():
                if not r["company"] or r["company"] in seen_companies:
                    continue
                short = _COMPANY_DISPLAY.get(r["company_short"], r["company_short"]) or ""
                company_defaults.append({"key": r["company"], "value": short})
                seen_companies.add(r["company"])
        finally:
            conn.close()

        # 2) 补充代码内置的公司（数据库里还没出现过的）
        keyword_to_full = {}
        for full_name in COMPANY_BASES:
            for keyword in COMPANY_SHORT_MAP:
                if keyword in full_name and keyword not in keyword_to_full:
                    keyword_to_full[keyword] = full_name
        seen_shorts = {c["value"] for c in company_defaults}
        for keyword, short in COMPANY_SHORT_MAP.items():
            full_name = keyword_to_full.get(keyword, keyword)
            if full_name not in seen_companies and short not in seen_shorts:
                company_defaults.append({"key": full_name, "value": short})
                seen_companies.add(full_name)
                seen_shorts.add(short)

        # ---- 险种 ----
        _SHORT_NAME_DISPLAY = {"驾乘/意外险": "驾意险"}
        policy_type_defaults = []
        seen_policy_types = set()
        # 1) 从数据库全局查
        conn = _db.pool.connection()
        try:
            cursor = conn.cursor(pymysql.cursors.DictCursor)
            cursor.execute(
                "SELECT DISTINCT policy_type FROM insurance_policy_fields "
                "WHERE policy_type IS NOT NULL AND policy_type != '' "
                "ORDER BY policy_type"
            )
            for r in cursor.fetchall():
                full_name = r["policy_type"]
                if full_name not in seen_policy_types:
                    _, short_name = _get_policy_type_code(full_name)
                    short_name = _SHORT_NAME_DISPLAY.get(short_name, short_name)
                    policy_type_defaults.append({"key": full_name, "value": short_name})
                    seen_policy_types.add(full_name)
        finally:
            conn.close()
        # 2) 补充常见险种（数据库里还没出现过的）
        for full_name, short_name in [
            ("机动车交通事故责任强制保险", "交强险"),
            ("机动车商业保险", "商业险"),
            ("驾乘人员意外伤害保险", "驾乘险"),
            ("新能源汽车商业保险", "新能源商业险"),
        ]:
            if full_name not in seen_policy_types:
                policy_type_defaults.append({"key": full_name, "value": short_name})
                seen_policy_types.add(full_name)

        return jsonify({"code": 0, "data": {
            "company_alias": company_defaults,
            "policy_type_alias": policy_type_defaults,
        }})
    except Exception as e:
        logger.exception("获取默认值失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


# ==================== 报错反馈 ====================

@insurance_bp.route("/api/insurance/error-report", methods=["POST"])
@login_required
def api_submit_error_report():
    """用户提交报错"""
    from auth.jwt_utils import verify_token
    from auth.db import get_user_by_id
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    payload = verify_token(token)
    if not payload:
        return jsonify({"code": 401, "msg": "未授权"}), 401
    user = get_user_by_id(_db, payload["user_id"])
    if not user:
        return jsonify({"code": 401, "msg": "未授权"}), 401

    data = request.get_json() or {}
    record_id = data.get("record_id")
    description = (data.get("description") or "").strip()
    filename = data.get("filename", "")
    policy_no = data.get("policy_no", "")
    enterprise = data.get("enterprise", "")
    if not description:
        return jsonify({"code": 400, "msg": "请描述问题"}), 400

    conn = _db.pool.connection()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO error_reports (record_id, user_id, user_name, enterprise, filename, policy_no, description, status) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, 'unread')",
            (record_id, user["id"], user.get("name") or user.get("phone"), enterprise, filename, policy_no, description)
        )
        conn.commit()
        # 入库后发钉钉群（复用「新保司通知」群）；通知失败不影响提交结果
        try:
            from core.notify import notify_error_report
            ntf = get_insurance_config(_db, "new_company_notify", {}) or {}
            _parts = []
            _uname = user.get("name") or user.get("phone") or ""
            if _uname:
                _parts.append(f"反馈人: {_uname}")
            if enterprise:
                _parts.append(f"企业: {enterprise}")
            if filename:
                _parts.append(f"文件: {filename}")
            if policy_no:
                _parts.append(f"保单号: {policy_no}")
            if record_id:
                _parts.append(f"record_id={record_id}")
            notify_error_report(description, " | ".join(_parts),
                                ntf.get("webhook", ""), ntf.get("secret", ""))
        except Exception:
            logger.exception("报错反馈钉钉通知失败")
        return jsonify({"code": 0, "msg": "已提交"})
    except Exception as e:
        logger.exception("提交报错失败")
        return jsonify({"code": 500, "msg": str(e)}), 500
    finally:
        conn.close()


@insurance_bp.route("/api/insurance/report-feedback", methods=["POST"])
def api_report_feedback():
    """经营报告页的「反馈与建议」→ 钉钉（复用新保司通知群，不强制登录，便于分享的报告也能反馈）"""
    try:
        data = request.get_json(force=True) or {}
        content = (data.get("content") or "").strip()
        if not content:
            return jsonify({"code": 400, "msg": "请填写反馈内容"}), 400
        _parts = []
        if data.get("enterprise_name"):
            _parts.append("企业: " + str(data["enterprise_name"])[:60])
        elif data.get("enterprise_id"):
            _parts.append("企业ID=" + str(data["enterprise_id"])[:20])
        if data.get("contact"):
            _parts.append("联系方式: " + str(data["contact"])[:60])
        if data.get("date_start"):
            _parts.append("报告周期 %s ~ %s" % (data.get("date_start", ""), data.get("date_end", "")))
        try:
            from core.notify import notify_feedback
            ntf = get_insurance_config(_db, "new_company_notify", {}) or {}
            notify_feedback(content, " | ".join(_parts), ntf.get("webhook", ""), ntf.get("secret", ""))
        except Exception:
            logger.exception("报告反馈钉钉通知失败")
        return jsonify({"code": 0, "msg": "已提交"})
    except Exception as e:
        logger.exception("提交报告反馈失败")
        return jsonify({"code": 500, "msg": str(e)}), 500


# 站内通知（报错列表/未读数/标记已读）已下线：报错反馈改为提交时直接发钉钉群。
# error_reports 表仍保留入库做存档，但不再提供超管读取/已读接口。
